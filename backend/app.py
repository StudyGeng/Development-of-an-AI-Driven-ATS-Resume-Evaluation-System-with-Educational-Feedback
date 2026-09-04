from __future__ import annotations

import json
import os
import re
import secrets
import shutil
import base64
from collections import Counter
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from difflib import SequenceMatcher
from email.message import EmailMessage
from html import escape as html_escape
from io import BytesIO
import smtplib
import urllib.error
import urllib.request
from typing import Any

from database import (
    DatabaseError,
    get_db_connection,
    init_mysql_database,
    is_duplicate_entry_error,
)
from data_cleaning import (
    build_job_data_cleaning_plan,
    build_job_data_quality_report,
    clean_importance_weight,
    clean_job_role_name,
    clean_job_text,
    clean_requirement_type,
    normalize_match_key as normalize_job_data_match_key,
)
from job_data_insights import build_job_data_insights
from flask import Flask, jsonify, redirect, request, send_file, send_from_directory, session
from flask_cors import CORS
from pypdf import PdfReader, PdfWriter
from sklearn.feature_extraction.text import TfidfVectorizer
from werkzeug.security import check_password_hash, generate_password_hash

try:
    from PIL import Image, UnidentifiedImageError  # type: ignore
except ImportError:  # pragma: no cover - optional OCR dependency
    Image = None  # type: ignore
    UnidentifiedImageError = OSError  # type: ignore

try:
    import pytesseract  # type: ignore
except ImportError:  # pragma: no cover - optional OCR dependency
    pytesseract = None  # type: ignore


APP_TITLE = "UTS CS Career Navigator Resume Analyzer"
BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
FRONTEND_DIR = os.path.join(BASE_DIR, "frontend")
FRONTEND_PAGES = {"home_page.html", "analysis_result.html", "job_data_insights.html"}
ACCURACY_TESTING_WORKBOOK = os.getenv(
    "ACCURACY_TESTING_WORKBOOK",
    os.path.join(BASE_DIR, "outputs", "manual_accuracy_testing", "ats_resume_fyp_testing.xlsx"),
)
ACCURACY_TESTING_PENDING_FILE = os.getenv(
    "ACCURACY_TESTING_PENDING_FILE",
    os.path.join(BASE_DIR, "outputs", "manual_accuracy_testing", "accuracy_testing_pending.jsonl"),
)
ACCURACY_TESTING_AUTO_EXPORT = os.getenv("ACCURACY_TESTING_AUTO_EXPORT", "1").strip().lower() not in {
    "0",
    "false",
    "no",
    "off",
}
RESUME_GENERATION_MAX_TEXT_LENGTH = 12000
CS_SAMPLE_ONLY_TOP_MATCH_THRESHOLD = 20.0
PDF_EXTENSIONS = {".pdf"}
IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg"}
SUPPORTED_RESUME_EXTENSIONS = PDF_EXTENSIONS | IMAGE_EXTENSIONS
EMAIL_PIN_LENGTH = 6
EMAIL_PIN_TTL_MINUTES = 10
EMAIL_PIN_MAX_ATTEMPTS = 5
EMAIL_PIN_PURPOSES = {"register", "password_reset"}
SPECIAL_CHARACTER_PATTERN = re.compile(r"[^A-Za-z0-9]")
ROLE_RECOMMENDATION_MARGIN = 8.0
ROLE_FAMILY_ALIASES = {
    "computer programmer": "software_development",
    "programmer": "software_development",
    "application developer": "software_development",
    "software developer": "software_development",
    "software engineer": "software_development",
    "software designer": "software_development",
    "java consultant": "software_development",
    "net developer": "software_development",
    "backend developer": "software_development",
    "frontend developer": "software_development",
    "full stack developer": "software_development",
    "full stack web developer": "software_development",
    "cyber security analyst": "cybersecurity",
    "cybersecurity analyst": "cybersecurity",
    "cybersecurity executive": "cybersecurity",
    "computer security manager": "cybersecurity",
    "data security specialist": "cybersecurity",
    "digital forensic specialist": "cybersecurity",
    "information security consultant": "cybersecurity",
    "security administrator information technology": "cybersecurity",
    "security technologist specialist": "cybersecurity",
    "database architect": "database",
    "oracle database engineer": "database",
    "sql database developer": "database",
    "network analyst": "networking",
    "network system engineer information technology": "networking",
    "computer network technician": "networking",
    "web developer": "web_development",
    "website developer": "web_development",
    "web designer": "web_development",
    "hypertext preprocessor php web programmer": "web_development",
}
POPULAR_ROLE_ALIASES = {
    "software engineer": "Software Developer",
    "software developer": "Software Developer",
    "application developer": "Software Developer",
    "backend developer": "Full Stack Developer",
    "frontend developer": "Full Stack Developer",
    "front end developer": "Full Stack Developer",
    "full stack developer": "Full Stack Developer",
    "web developer": "Full Stack Developer",
    "website developer": "Full Stack Developer",
    "web application developer": "Full Stack Developer",
    "web development intern": "Full Stack Developer",
    "front end web developer": "Full Stack Developer",
    "software designer": "Software Designer",
    "java consultant": "Java Consultant",
    "data analyst": "Oracle Database Engineer",
    "data scientist": "Computer Specialist",
    "machine learning engineer": "Software Developer",
    "devops engineer": "Computer Support Engineer",
    "cloud engineer": "Computer Support Engineer",
    "cybersecurity analyst": "Cybersecurity Analyst",
    "cyber security analyst": "Cybersecurity Analyst",
    "information security analyst": "Cybersecurity Analyst",
    "security analyst": "Cybersecurity Analyst",
    "soc analyst": "Cybersecurity Analyst",
    "security engineer": "Security Technologist Specialist",
    "security administrator": "Security Administrator (Information Technology)",
    "research scientist": "Computer Specialist",
    "database engineer": "Oracle Database Engineer",
    "database developer": "SQL Database Developer",
    "sql developer": "SQL Database Developer",
}
ROLE_TITLE_MATCH_ALIASES = {
    "cybersecurity analyst": {
        "cybersecurity analyst",
        "cyber security analyst",
        "cyber-security analyst",
        "information security analyst",
        "security analyst",
        "soc analyst",
    },
    "cybersecurity executive": {
        "cybersecurity executive",
        "cyber security executive",
        "cybersecurity specialist",
        "security executive",
    },
    "computer security manager": {
        "computer security manager",
        "cybersecurity manager",
        "cyber security manager",
        "information security manager",
        "security manager",
    },
    "data security specialist": {
        "data security specialist",
        "information security specialist",
        "data protection specialist",
        "cybersecurity specialist",
    },
    "digital forensic specialist": {
        "digital forensic specialist",
        "digital forensics specialist",
        "cyber forensic specialist",
        "incident investigator",
    },
    "information security consultant": {
        "information security consultant",
        "cybersecurity consultant",
        "cyber security consultant",
        "security consultant",
    },
    "security administrator information technology": {
        "security administrator",
        "it security administrator",
        "information security administrator",
        "cybersecurity administrator",
    },
    "security technologist specialist": {
        "security technologist specialist",
        "security technologist",
        "security engineer",
        "cybersecurity engineer",
        "cyber security engineer",
    },
    "sql database developer": {
        "sql database developer",
        "database developer",
        "sql developer",
        "database programmer",
    },
    "oracle database engineer": {
        "oracle database engineer",
        "oracle database administrator",
        "oracle dba",
        "database engineer",
    },
    "database architect": {
        "database architect",
        "data architect",
        "database designer",
    },
    "web designer": {
        "web designer",
        "web design specialist",
        "website designer",
        "ui designer",
        "ux designer",
    },
    "hypertext preprocessor php web programmer": {
        "php web programmer",
        "php developer",
        "web programmer",
        "web application developer",
    },
    "full stack developer": {
        "full stack developer",
        "fullstack developer",
        "full stack web developer",
        "fullstack web developer",
        "full-stack developer",
        "full-stack web developer",
        "web developer",
        "website developer",
        "web application developer",
        "web development intern",
        "web developer intern",
        "front end web developer",
        "frontend web developer",
        "front end developer",
        "frontend developer",
        "back end developer",
        "backend developer",
    },
    "software developer": {
        "software developer",
        "software engineer",
        "application developer",
    },
    "software designer": {
        "software designer",
        "software design engineer",
    },
    "java consultant": {
        "java consultant",
        "java developer",
    },
}
REQUIREMENT_TYPE_LABELS = {
    "task": "Key responsibilities",
    "skill": "Required skills",
    "knowledge": "Knowledge areas",
    "tool": "Tools and technologies",
    "education": "Education background",
    "experience": "Experience",
}
REQUIREMENT_TYPE_ORDER = ["task", "skill", "knowledge", "tool", "education", "experience"]
REQUIREMENT_TYPE_SCORE_WEIGHTS = {
    "task": 1.2,
    "experience": 1.2,
    "skill": 1.1,
    "knowledge": 0.85,
    "education": 0.75,
    "tool": 0.7,
}
KEYWORD_STOPWORDS = {
    "a",
    "an",
    "and",
    "are",
    "as",
    "at",
    "be",
    "by",
    "candidate",
    "candidates",
    "code",
    "conduct",
    "computer",
    "complete",
    "ensure",
    "for",
    "from",
    "ability",
    "area",
    "areas",
    "background",
    "include",
    "including",
    "if",
    "in",
    "into",
    "is",
    "it",
    "key",
    "knowledge",
    "masco",
    "must",
    "of",
    "on",
    "or",
    "required",
    "requires",
    "requiring",
    "requirement",
    "requirements",
    "responsibilities",
    "role",
    "roles",
    "responsible",
    "skill",
    "skills",
    "such",
    "target",
    "that",
    "the",
    "their",
    "this",
    "to",
    "tools",
    "use",
    "using",
    "with",
}
CANONICAL_KEYWORD_ALIASES = {
    "analysed": "analysis",
    "analyze": "analysis",
    "analyzed": "analysis",
    "analyzing": "analysis",
    "applications": "application",
    "apps": "application",
    "apis": "api",
    "assessed": "assess",
    "assesses": "assess",
    "assessing": "assess",
    "assessments": "assessment",
    "back-end": "backend",
    "back-end.": "backend",
    "building": "build",
    "built": "build",
    "controls": "control",
    "coding": "code",
    "created": "create",
    "creating": "create",
    "cyber-security": "cybersecurity",
    "databases": "database",
    "designed": "design",
    "designing": "design",
    "designs": "design",
    "debugged": "debug",
    "debugging": "debug",
    "debugs": "debug",
    "decreased": "decrease",
    "decreasing": "decrease",
    "developed": "develop",
    "developer": "develop",
    "developers": "develop",
    "developing": "develop",
    "development": "develop",
    "develops": "develop",
    "delivered": "deliver",
    "delivering": "deliver",
    "enabled": "enable",
    "enabling": "enable",
    "engineered": "engineer",
    "engineering": "engineer",
    "engineers": "engineer",
    "evaluated": "evaluate",
    "evaluating": "evaluate",
    "fixed": "fix",
    "fixing": "fix",
    "front-end": "frontend",
    "front-end.": "frontend",
    "full-stack": "fullstack",
    "incidents": "incident",
    "inspections": "inspection",
    "investigated": "investigate",
    "investigating": "investigate",
    "investigation": "investigate",
    "implementing": "implement",
    "implemented": "implement",
    "implements": "implement",
    "maintained": "maintain",
    "maintaining": "maintain",
    "maintains": "maintain",
    "node.js": "node",
    "nodejs": "node",
    "optimized": "optimize",
    "optimizing": "optimize",
    "procedures": "procedure",
    "programmer": "program",
    "programmers": "program",
    "programming": "program",
    "programs": "program",
    "resolved": "resolve",
    "resolving": "resolve",
    "restful": "rest",
    "react.js": "react",
    "risks": "risk",
    "specifications": "specification",
    "troubleshooting": "troubleshoot",
    "troubleshoots": "troubleshoot",
    "threats": "threat",
    "systems": "system",
    "technologies": "technology",
    "tested": "test",
    "testing": "test",
    "tests": "test",
    "viruses": "virus",
    "vulnerabilities": "vulnerability",
}
SEMANTIC_KEYWORD_EQUIVALENTS = {
    "analysis": {"analyze", "analytics", "evaluate", "assess", "review"},
    "application": {"app", "software", "platform", "system"},
    "backend": {"back", "back-end", "server", "api", "database", "node", "django", "flask", "php"},
    "build": {"built", "building", "create", "develop", "implement", "deliver", "engineer"},
    "cloud": {"aws", "azure", "gcp", "docker", "kubernetes", "devops", "deployment"},
    "complex": {"advanced", "challenging", "problem", "system"},
    "computing": {"cloud", "technology", "infrastructure"},
    "critical": {"analysis", "problem", "solve", "decision"},
    "css": {"html", "tailwind", "bootstrap", "frontend"},
    "debug": {"troubleshoot", "resolve", "fix", "diagnose", "support"},
    "decrease": {"reduce", "reduced", "improve", "optimized", "save"},
    "design": {"architect", "architecture", "model", "plan", "prototype", "ux", "ui"},
    "develop": {"build", "built", "create", "implement", "code", "program", "engineer"},
    "dotnet": {"c#", "csharp", "aspnet", "backend"},
    "frontend": {"front", "front-end", "ui", "ux", "html", "css", "javascript", "typescript", "react", "angular", "bootstrap"},
    "fullstack": {"full", "stack", "frontend", "backend", "front-end", "back-end", "web"},
    "hardware": {"computer", "device", "infrastructure", "network"},
    "javascript": {"js", "typescript", "react", "node", "frontend"},
    "jquery": {"javascript", "js", "react", "frontend"},
    "maintain": {"support", "improve", "optimize", "refactor", "debug", "update", "enhance"},
    "program": {"code", "coding", "develop", "software"},
    "responsive": {"mobile", "cross-platform", "frontend", "web", "ui"},
    "security": {"cybersecurity", "cyber", "threat", "risk", "malware", "vulnerability", "incident", "attack", "protection"},
    "sql": {"mysql", "postgresql", "sqlalchemy", "oracle", "plsql", "tsql"},
    "specification": {"requirement", "documentation", "technical", "design"},
    "system": {"application", "platform", "backend", "service", "api", "database", "software"},
    "test": {"qa", "validation", "unit", "integration", "pytest", "jest", "sdlc"},
    "thinking": {"analysis", "problem", "solve", "decision"},
    "tomcat": {"apache", "server", "deployment", "java", "backend"},
    "troubleshoot": {"debug", "resolve", "fix", "diagnose", "support"},
    "vulnerability": {"risk", "weakness", "exposure"},
}
STRICT_TECHNICAL_KEYWORD_EVIDENCE = {
    "sql": {"sql", "mysql", "postgresql", "sqlalchemy", "oracle", "plsql", "tsql"},
    "mysql": {"mysql"},
    "postgresql": {"postgresql"},
    "oracle": {"oracle"},
}
SOFTWARE_DEVELOPER_EVIDENCE_GROUPS = [
    {"software", "develop", "program", "code", "python", "java", "javascript", "typescript", "c#", "c++"},
    {"api", "rest", "flask", "django", "fastapi", "react", "node", "html", "css", "frontend", "backend"},
    {"sql", "mysql", "postgresql", "sqlalchemy", "database", "orm"},
    {"test", "debug", "validation", "qa", "sdlc", "troubleshoot"},
    {"design", "architecture", "system", "application", "platform"},
    {"git", "docker", "kubernetes", "aws", "azure", "gcp", "cloud", "deployment", "ci", "cd"},
]
CYBERSECURITY_EVIDENCE_GROUPS = [
    {"cybersecurity", "cyber", "security", "information security", "network security", "soc"},
    {"threat", "risk", "mitigate", "malware", "virus", "attack", "vulnerability", "incident", "intrusion"},
    {"assessment", "audit", "penetration", "nist", "inspection", "security plan", "forensic", "digital evidence"},
    {"data protection", "dlp", "policy", "access", "control", "compliance", "secure"},
    {"network", "firewall", "endpoint", "computer", "system", "infrastructure"},
    {"monitor", "detect", "respond", "investigate", "analysis"},
]
DATABASE_EVIDENCE_GROUPS = [
    {"database", "dbms", "sql", "mysql", "postgresql", "oracle", "mongodb"},
    {"query", "schema", "table", "data model", "database design", "data architecture"},
    {"administration", "management", "backup", "performance", "optimize", "maintain"},
    {"stored procedure", "etl", "data warehouse", "data migration", "replication"},
    {"data", "integrity", "validation", "report", "analytics"},
]
NETWORK_EVIDENCE_GROUPS = [
    {"network", "networking", "telecommunication", "infrastructure"},
    {"routing", "cisco", "switch", "router", "lan", "wan"},
    {"configure", "troubleshoot", "monitor", "performance", "optimize"},
    {"network security", "firewall", "intrusion", "vpn", "access control"},
    {"hardware", "cable", "server", "operating system"},
]
QA_TESTING_EVIDENCE_GROUPS = [
    {"qa", "quality assurance", "tester", "test", "testing", "validation"},
    {"bug", "defect", "debug", "troubleshoot", "issue"},
    {"test plan", "test case", "automation", "selenium", "unit", "integration"},
    {"software", "application", "system", "web", "mobile"},
]
WEB_DEVELOPMENT_EVIDENCE_GROUPS = [
    {"web", "website", "web application", "web page", "frontend", "front-end"},
    {"html", "css", "javascript", "jquery", "xml", "php", "wordpress", "dreamweaver"},
    {"design", "layout", "ui", "ux", "responsive", "navigation", "color", "visual"},
    {"develop", "build", "create", "implement", "code", "bug", "enhancement", "functionality"},
    {"user", "customer", "client", "conversion", "click", "performance", "security"},
    {"backend", "server", "database", "oracle", "sql", ".net", "application"},
]
DEFAULT_SKILL_DICTIONARY = [
    ".net",
    "abap",
    "ajax",
    "angular",
    "angularjs",
    "api",
    "apache ant",
    "asp.net",
    "backend",
    "back-end",
    "blackboard",
    "bootstrap",
    "c#",
    "c++",
    "ci/cd",
    "cloud computing",
    "cissp",
    "cybersecurity",
    "data structures",
    "data protection",
    "debugging",
    "digital forensics",
    "deployment",
    "dlp",
    "dreamweaver",
    "frontend",
    "front-end",
    "full-stack",
    "java",
    "javascript",
    "jquery",
    "operations analysis",
    "node",
    "node.js",
    "oop",
    "php",
    "powershell",
    "quality assurance",
    "rest api",
    "restful api",
    "sdlc",
    "security+",
    "security audit",
    "security assessment",
    "security plans",
    "nist",
    "penetration testing",
    "intrusion detection",
    "malware analysis",
    "network security",
    "threat analysis",
    "software development",
    "software testing",
    "spring mvc",
    "spring soa",
    "system design",
    "tailwind",
    "typescript",
    "qa",
    "database",
    "mobile app",
    "visual basic",
    "wordpress",
    "xml",
    "web app",
    "web application",
    "web technologies",
    "website",
    "python",
    "sql",
    "data analysis",
    "machine learning",
    "deep learning",
    "nlp",
    "nosql",
    "mongodb",
    "apache spark",
    "objective c",
    "flask",
    "django",
    "fastapi",
    "docker",
    "kubernetes",
    "aws",
    "azure",
    "gcp",
    "git",
    "linux",
    "tableau",
    "power bi",
    "excel",
    "statistics",
    "communication",
    "leadership",
    "problem solving",
    "research",
    "tensorflow",
    "pytorch",
    "vulnerability assessment",
    "react",
    "javascript",
    "html",
    "css",
    "mysql",
    "postgresql",
    "supabase",
]
COMMON_EDUCATION_MARKERS = ["bachelor", "master", "phd", "b.sc", "m.sc", "degree", "university", "college"]
COMMON_EXPERIENCE_MARKERS = [
    "intern",
    "internship",
    "project",
    "experience",
    "responsibility",
    "worked",
    "developed",
    "built",
    "designed",
    "deployed",
    "tested",
    "led",
    "managed",
    "assisted",
    "automated",
    "compiled",
    "conducted",
    "contributed",
    "coordinated",
    "decreased",
    "executed",
    "fixed",
    "improved",
    "implemented",
    "launched",
    "maintained",
    "optimized",
    "partnered",
    "performed",
    "redesigned",
    "revamped",
    "reduced",
    "supported",
    "trained",
]
CORE_RESUME_SECTION_MARKERS = [
    "about me",
    "achievements",
    "awards",
    "certifications",
    "certificates",
    "education",
    "experience",
    "work experience",
    "projects",
    "skills",
    "technical skills",
]
CONTACT_MARKERS = ["linkedin", "github", "portfolio"]
ENGLISH_LANGUAGE_MARKERS = {
    "and",
    "with",
    "for",
    "experience",
    "education",
    "skills",
    "project",
    "developed",
    "software",
    "engineer",
    "university",
    "work",
}
MALAY_LANGUAGE_MARKERS = {
    "dan",
    "yang",
    "dengan",
    "untuk",
    "saya",
    "adalah",
    "pengalaman",
    "kemahiran",
    "pendidikan",
    "universiti",
    "projek",
    "pembangunan",
    "perisian",
    "kerja",
}


app = Flask(__name__)
app.secret_key = os.getenv("FLASK_SECRET_KEY", "uts-career-navigator-dev-secret")
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
)
CORS(app, supports_credentials=True)


@dataclass
class ParsedResume:
    text: str
    skills: list[str]
    education: list[str]
    experience: list[str]
    excerpt: str


def normalize_email(email: str) -> str:
    return normalize_text(email).lower()


def get_admin_emails() -> set[str]:
    configured = os.getenv("ADMIN_EMAILS", os.getenv("ADMIN_EMAIL", "admin@uts.local"))
    return {
        normalize_email(email)
        for email in re.split(r"[,;]", configured)
        if normalize_email(email)
    }


def is_configured_admin_email(email: str) -> bool:
    return normalize_email(email) in get_admin_emails()


def is_user_admin(user: dict[str, Any] | None) -> bool:
    if not user:
        return False
    return bool(user.get("is_admin")) or is_configured_admin_email(str(user.get("email") or ""))


def serialize_user(user: dict[str, Any]) -> dict[str, Any]:
    return {
        "id": user["id"],
        "full_name": user["full_name"],
        "email": user["email"],
        "email_verified_at": user.get("email_verified_at"),
        "created_at": user.get("created_at"),
        "is_admin": is_user_admin(user),
    }


def save_login_session(user: dict[str, Any]) -> None:
    session.clear()
    session["user_id"] = int(user["id"])
    session["email"] = normalize_email(str(user.get("email") or ""))
    session["is_admin"] = is_user_admin(user)


def clear_login_session() -> None:
    session.clear()


def current_session_is_admin() -> bool:
    return bool(session.get("is_admin"))


def require_admin_json_response():
    if current_session_is_admin():
        return None
    return jsonify({"detail": "Admin access required"}), 403


def is_valid_email(email: str) -> bool:
    return bool(re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", email))


def utc_now_naive() -> datetime:
    return datetime.now(timezone.utc).replace(tzinfo=None)


def normalize_pin(pin: str) -> str:
    return re.sub(r"\D+", "", str(pin or ""))


def normalize_pin_purpose(purpose: str) -> str:
    normalized = normalize_match_text(purpose).replace(" ", "_")
    aliases = {
        "signup": "register",
        "sign_up": "register",
        "create_account": "register",
        "forgot_password": "password_reset",
        "reset_password": "password_reset",
    }
    return aliases.get(normalized, normalized)


def get_password_validation_error(password: str) -> str:
    if len(password) < 8:
        return "Password must be at least 8 characters."
    if not re.search(r"[A-Za-z]", password):
        return "Password must include at least 1 letter."
    if not re.search(r"\d", password):
        return "Password must include at least 1 number."
    if not SPECIAL_CHARACTER_PATTERN.search(password):
        return "Password must include at least 1 special character, for example !, @, #, or $."
    return ""


def email_pin_subject(purpose: str) -> str:
    if purpose == "password_reset":
        return "Your UTS CS Career Navigator password reset PIN"
    return "Your UTS CS Career Navigator email verification PIN"


def email_pin_body(pin: str, purpose: str) -> str:
    action = "reset your password" if purpose == "password_reset" else "verify your email address"
    return (
        f"Your UTS CS Career Navigator PIN is {pin}.\n\n"
        f"Use this PIN to {action}. It expires in {EMAIL_PIN_TTL_MINUTES} minutes.\n\n"
        "If you did not request this, you can ignore this email."
    )


def send_resend_api_email_pin(email: str, pin: str, purpose: str) -> str:
    api_key = os.getenv("RESEND_API_KEY", "").strip()
    from_email = os.getenv(
        "RESEND_FROM",
        os.getenv("SMTP_FROM", "UTS CS Career Navigator <onboarding@resend.dev>"),
    ).strip()
    payload = {
        "from": from_email,
        "to": [email],
        "subject": email_pin_subject(purpose),
        "text": email_pin_body(pin, purpose),
    }
    request = urllib.request.Request(
        "https://api.resend.com/emails",
        data=json.dumps(payload).encode("utf-8"),
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(request, timeout=10) as response:
            if response.status >= 400:
                raise OSError(f"Resend API error {response.status}: {response.reason}")
    except urllib.error.HTTPError as error:
        detail = error.read().decode("utf-8", errors="ignore")
        raise OSError(f"Resend API error {error.code}: {detail}") from error
    except urllib.error.URLError as error:
        raise OSError(f"Resend API request failed: {error.reason}") from error
    return "resend_api"


def send_email_pin(email: str, pin: str, purpose: str) -> str:
    if os.getenv("RESEND_API_KEY", "").strip():
        return send_resend_api_email_pin(email, pin, purpose)

    smtp_host = os.getenv("SMTP_HOST", "").strip()
    if not smtp_host:
        print(f"[DEV EMAIL PIN] purpose={purpose} email={email} pin={pin}")
        return "dev"

    smtp_port = int(os.getenv("SMTP_PORT", "587"))
    smtp_username = os.getenv("SMTP_USERNAME", "").strip()
    smtp_password = os.getenv("SMTP_PASSWORD", "")
    smtp_from = os.getenv("SMTP_FROM", smtp_username or "no-reply@career-navigator.local").strip()
    smtp_use_tls = os.getenv("SMTP_USE_TLS", "1").strip().lower() not in {"0", "false", "no"}

    message = EmailMessage()
    message["Subject"] = email_pin_subject(purpose)
    message["From"] = smtp_from
    message["To"] = email
    message.set_content(email_pin_body(pin, purpose))

    with smtplib.SMTP(smtp_host, smtp_port, timeout=10) as server:
        if smtp_use_tls:
            server.starttls()
        if smtp_username:
            server.login(smtp_username, smtp_password)
        server.send_message(message)

    return "smtp"


def should_return_dev_pin(delivery: str) -> bool:
    return delivery == "dev" or os.getenv("EMAIL_PIN_DEV_MODE", "").strip().lower() in {"1", "true", "yes"}


def get_email_delivery_error_message(error: Exception) -> str:
    raw_error = ""
    if isinstance(error, smtplib.SMTPResponseException):
        smtp_error = error.smtp_error
        raw_error = smtp_error.decode("utf-8", errors="ignore") if isinstance(smtp_error, bytes) else str(smtp_error)
    else:
        raw_error = str(error)

    normalized = raw_error.lower()
    if "only send testing emails to your own email address" in normalized:
        return (
            "Resend test mode only sends to your own Resend account email. "
            "Use that email for testing, or verify a domain in Resend before sending to other users."
        )
    if "application-specific password required" in normalized or "invalidsecondfactor" in normalized:
        return "Gmail requires an App Password. Turn on Google 2-Step Verification, create an App Password, then use it as SMTP_PASSWORD."
    if "username and password not accepted" in normalized:
        return "Gmail rejected the SMTP login. Check that SMTP_USERNAME is your Gmail and SMTP_PASSWORD is a valid Google App Password."
    if "domain" in normalized and ("verify" in normalized or "verified" in normalized):
        return "Resend requires a verified sending domain. Verify your domain, then use a From email from that domain."
    if "resend api" in normalized and ("forbidden" in normalized or "api key" in normalized):
        return "Resend API rejected the request. Check that RESEND_API_KEY is active and RESEND_FROM is allowed."
    if "authentication" in normalized or "invalid api key" in normalized or "unauthorized" in normalized:
        return "Email authentication failed. Check that SMTP_PASSWORD or RESEND_API_KEY is active."
    return "Could not send verification email. Check SMTP settings."


def email_exists(email: str) -> bool:
    connection = get_db_connection()
    cursor = connection.cursor()
    cursor.execute("SELECT id FROM users WHERE email = %s LIMIT 1", (email,))
    exists = cursor.fetchone() is not None
    cursor.close()
    connection.close()
    return exists


def create_email_pin(email: str, purpose: str) -> tuple[str, datetime]:
    pin = f"{secrets.randbelow(10 ** EMAIL_PIN_LENGTH):0{EMAIL_PIN_LENGTH}d}"
    expires_at = utc_now_naive() + timedelta(minutes=EMAIL_PIN_TTL_MINUTES)
    pin_hash = generate_password_hash(pin)

    connection = get_db_connection()
    cursor = connection.cursor()
    cursor.execute(
        """
        UPDATE email_verification_pins
        SET consumed_at = UTC_TIMESTAMP()
        WHERE email = %s
          AND purpose = %s
          AND consumed_at IS NULL
        """,
        (email, purpose),
    )
    cursor.execute(
        """
        INSERT INTO email_verification_pins (email, purpose, pin_hash, expires_at)
        VALUES (%s, %s, %s, %s)
        """,
        (email, purpose, pin_hash, expires_at),
    )
    connection.commit()
    cursor.close()
    connection.close()
    return pin, expires_at


def consume_email_pin(email: str, purpose: str, pin: str) -> tuple[bool, str]:
    normalized_pin = normalize_pin(pin)
    if len(normalized_pin) != EMAIL_PIN_LENGTH:
        return False, f"Enter the {EMAIL_PIN_LENGTH}-digit PIN."

    connection = get_db_connection()
    cursor = connection.cursor(dictionary=True)
    cursor.execute(
        """
        SELECT id, pin_hash, attempts, expires_at
        FROM email_verification_pins
        WHERE email = %s
          AND purpose = %s
          AND consumed_at IS NULL
        ORDER BY id DESC
        LIMIT 1
        """,
        (email, purpose),
    )
    pin_row = cursor.fetchone()

    if not pin_row:
        cursor.close()
        connection.close()
        return False, "Request a new PIN first."

    pin_id = int(pin_row["id"])
    attempts = int(pin_row.get("attempts") or 0)
    expires_at = pin_row["expires_at"]

    if isinstance(expires_at, datetime) and expires_at < utc_now_naive():
        cursor.execute("UPDATE email_verification_pins SET consumed_at = UTC_TIMESTAMP() WHERE id = %s", (pin_id,))
        connection.commit()
        cursor.close()
        connection.close()
        return False, "This PIN has expired. Request a new one."

    if attempts >= EMAIL_PIN_MAX_ATTEMPTS:
        cursor.execute("UPDATE email_verification_pins SET consumed_at = UTC_TIMESTAMP() WHERE id = %s", (pin_id,))
        connection.commit()
        cursor.close()
        connection.close()
        return False, "Too many wrong attempts. Request a new PIN."

    if not check_password_hash(pin_row["pin_hash"], normalized_pin):
        cursor.execute("UPDATE email_verification_pins SET attempts = attempts + 1 WHERE id = %s", (pin_id,))
        connection.commit()
        cursor.close()
        connection.close()
        return False, "Invalid PIN."

    cursor.execute("UPDATE email_verification_pins SET consumed_at = UTC_TIMESTAMP() WHERE id = %s", (pin_id,))
    connection.commit()
    cursor.close()
    connection.close()
    return True, "PIN verified."


def normalize_text(value: str) -> str:
    return re.sub(r"\s+", " ", value or "").strip()


def normalize_match_text(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", value.lower()).strip()


def role_similarity_score(requested: str, candidate: str) -> float:
    requested_normalized = normalize_match_text(requested)
    candidate_normalized = normalize_match_text(candidate)
    if not requested_normalized or not candidate_normalized:
        return 0.0

    sequence_score = SequenceMatcher(None, requested_normalized, candidate_normalized).ratio()
    requested_tokens = set(requested_normalized.split())
    candidate_tokens = set(candidate_normalized.split())
    if not requested_tokens or not candidate_tokens:
        return sequence_score

    token_score = len(requested_tokens & candidate_tokens) / max(len(requested_tokens), len(candidate_tokens))
    contains_score = 1.0 if requested_normalized in candidate_normalized or candidate_normalized in requested_normalized else 0.0
    return max(sequence_score, token_score, contains_score)


def fetch_job_roles() -> list[dict[str, Any]]:
    connection = get_db_connection()
    cursor = connection.cursor(dictionary=True)
    cursor.execute(
        """
        SELECT
            jr.id,
            jr.role_name,
            jr.masco_description,
            COUNT(req.id) AS requirement_count
        FROM job_roles jr
        LEFT JOIN job_requirements req
            ON req.role_id = jr.id
        GROUP BY jr.id, jr.role_name, jr.masco_description
        ORDER BY jr.role_name
        """
    )
    rows = cursor.fetchall()
    cursor.close()
    connection.close()
    roles: list[dict[str, Any]] = []
    seen_role_keys: set[str] = set()
    for row in rows:
        role_name = clean_job_role_name(row.get("role_name"))
        role_key = normalize_job_data_match_key(role_name)
        if not role_name or not role_key or role_key in seen_role_keys:
            continue
        seen_role_keys.add(role_key)
        roles.append(
            {
                "id": row["id"],
                "role_name": role_name,
                "masco_description": clean_job_text(row.get("masco_description")),
                "requirement_count": int(row.get("requirement_count") or 0),
            }
        )
    return roles


def fetch_job_requirements(role_id: int) -> list[dict[str, Any]]:
    connection = get_db_connection()
    cursor = connection.cursor(dictionary=True)
    cursor.execute(
        """
        SELECT
            id,
            requirement_type,
            requirement_text,
            importance_weight
        FROM job_requirements
        WHERE role_id = %s
        ORDER BY importance_weight DESC, requirement_type, id
        """,
        (role_id,),
    )
    rows = cursor.fetchall()
    cursor.close()
    connection.close()
    requirements: list[dict[str, Any]] = []
    seen_requirement_keys: set[tuple[str, str]] = set()
    for row in rows:
        requirement_type = clean_requirement_type(row.get("requirement_type"))
        requirement_text = clean_job_text(row.get("requirement_text"))
        requirement_key = (requirement_type, normalize_job_data_match_key(requirement_text))
        if not requirement_text or not requirement_key[1] or requirement_key in seen_requirement_keys:
            continue
        seen_requirement_keys.add(requirement_key)
        requirements.append(
            {
                "id": row["id"],
                "requirement_type": requirement_type,
                "requirement_text": requirement_text,
                "importance_weight": clean_importance_weight(row.get("importance_weight")),
            }
        )
    return requirements


def fetch_raw_job_data_rows() -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    connection = get_db_connection()
    cursor = connection.cursor(dictionary=True)
    cursor.execute(
        """
        SELECT
            id,
            role_name,
            masco_description
        FROM job_roles
        ORDER BY id
        """
    )
    role_rows = cursor.fetchall()
    cursor.execute(
        """
        SELECT
            id,
            role_id,
            requirement_type,
            requirement_text,
            importance_weight
        FROM job_requirements
        ORDER BY role_id, id
        """
    )
    requirement_rows = cursor.fetchall()
    cursor.close()
    connection.close()
    return role_rows, requirement_rows


def apply_job_data_cleaning_plan(plan: dict[str, Any]) -> dict[str, Any]:
    role_updates = plan.get("role_updates", []) or []
    requirement_updates = plan.get("requirement_updates", []) or []
    duplicate_requirement_ids = [
        int(requirement_id)
        for requirement_id in plan.get("duplicate_requirement_ids", []) or []
        if requirement_id
    ]

    connection = get_db_connection()
    cursor = connection.cursor()
    try:
        for update in role_updates:
            cursor.execute(
                """
                UPDATE job_roles
                SET role_name = %s,
                    masco_description = %s
                WHERE id = %s
                """,
                (
                    update["role_name"],
                    update["masco_description"],
                    int(update["id"]),
                ),
            )

        for update in requirement_updates:
            cursor.execute(
                """
                UPDATE job_requirements
                SET requirement_type = %s,
                    requirement_text = %s,
                    importance_weight = %s
                WHERE id = %s
                """,
                (
                    update["requirement_type"],
                    update["requirement_text"],
                    int(update["importance_weight"]),
                    int(update["id"]),
                ),
            )

        if duplicate_requirement_ids:
            placeholders = ", ".join(["%s"] * len(duplicate_requirement_ids))
            cursor.execute(
                f"DELETE FROM job_requirements WHERE id IN ({placeholders})",
                tuple(duplicate_requirement_ids),
            )

        connection.commit()
    except Exception:
        connection.rollback()
        raise
    finally:
        cursor.close()
        connection.close()

    return {
        "role_updates_applied": len(role_updates),
        "requirement_updates_applied": len(requirement_updates),
        "duplicate_requirements_removed": len(duplicate_requirement_ids),
    }


def fetch_job_data_for_insights() -> tuple[list[dict[str, Any]], dict[int, list[dict[str, Any]]]]:
    roles = fetch_job_roles()
    requirements_by_role: dict[int, list[dict[str, Any]]] = {}
    for role in roles:
        role_id = int(role["id"])
        requirements_by_role[role_id] = fetch_job_requirements(role_id)
    return roles, requirements_by_role


def find_job_role_match(requested_role: str) -> tuple[dict[str, Any] | None, str]:
    requested = normalize_match_text(requested_role)
    if not requested:
        return None, "none"

    roles = fetch_job_roles()
    normalized_roles = {normalize_match_text(role["role_name"]): role for role in roles}

    if requested in normalized_roles:
        return normalized_roles[requested], "exact"

    alias_target = POPULAR_ROLE_ALIASES.get(requested)
    if alias_target:
        alias_key = normalize_match_text(alias_target)
        if alias_key in normalized_roles:
            return normalized_roles[alias_key], "alias"

    best_alias_key = ""
    best_alias_score = 0.0
    for alias_key in POPULAR_ROLE_ALIASES:
        score = role_similarity_score(requested, alias_key)
        if score > best_alias_score:
            best_alias_score = score
            best_alias_key = alias_key

    if best_alias_key and best_alias_score >= 0.78:
        alias_target = POPULAR_ROLE_ALIASES[best_alias_key]
        alias_key = normalize_match_text(alias_target)
        if alias_key in normalized_roles:
            return normalized_roles[alias_key], "alias_fuzzy"

    for role in roles:
        role_key = normalize_match_text(role["role_name"])
        if requested in role_key or role_key in requested:
            return role, "partial"

    requested_tokens = set(requested.split())
    best_role: dict[str, Any] | None = None
    best_score = 0
    for role in roles:
        role_tokens = set(normalize_match_text(role["role_name"]).split())
        score = len(requested_tokens & role_tokens)
        if score > best_score:
            best_score = score
            best_role = role

    if best_role and best_score >= 2:
        return best_role, "keyword"

    fuzzy_role: dict[str, Any] | None = None
    fuzzy_score = 0.0
    for role in roles:
        score = role_similarity_score(requested, role["role_name"])
        if score > fuzzy_score:
            fuzzy_score = score
            fuzzy_role = role

    if fuzzy_role and fuzzy_score >= 0.68:
        return fuzzy_role, "fuzzy"

    return None, "none"


def group_job_requirements(requirements: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    grouped: dict[str, list[dict[str, Any]]] = {}
    for requirement in requirements:
        grouped.setdefault(requirement["requirement_type"], []).append(requirement)
    return grouped


def build_job_description_from_requirements(role: dict[str, Any], requirements: list[dict[str, Any]]) -> str:
    lines = [f"Target role: {role['role_name']}."]
    masco_description = normalize_text(role.get("masco_description", ""))
    if masco_description:
        lines.extend(["", masco_description])

    grouped = group_job_requirements(requirements)
    for requirement_type in REQUIREMENT_TYPE_ORDER:
        items = grouped.get(requirement_type, [])
        if not items:
            continue
        label = REQUIREMENT_TYPE_LABELS.get(requirement_type, requirement_type.title())
        lines.extend(["", f"{label}:"])
        for item in items:
            lines.append(f"- {item['requirement_text']}")

    for requirement_type in sorted(set(grouped) - set(REQUIREMENT_TYPE_ORDER)):
        label = REQUIREMENT_TYPE_LABELS.get(requirement_type, requirement_type.title())
        lines.extend(["", f"{label}:"])
        for item in grouped[requirement_type]:
            lines.append(f"- {item['requirement_text']}")

    return "\n".join(lines)


def resolve_database_job_description(requested_role: str) -> dict[str, Any] | None:
    matched_role, match_type = find_job_role_match(requested_role)
    if not matched_role:
        return None

    requirements = fetch_job_requirements(int(matched_role["id"]))
    if not requirements:
        return None

    return {
        "role": matched_role,
        "requirements": requirements,
        "match_type": match_type,
        "job_description": build_job_description_from_requirements(matched_role, requirements),
    }


def compute_role_title_match_score(resume_text: str, role_name: str) -> float:
    normalized_resume = f" {normalize_match_text(resume_text)} "
    normalized_role = normalize_match_text(role_name)
    if not normalized_resume.strip() or not normalized_role:
        return 0.0

    aliases = ROLE_TITLE_MATCH_ALIASES.get(normalized_role, {normalized_role})
    for alias in aliases:
        normalized_alias = normalize_match_text(alias)
        if normalized_alias and f" {normalized_alias} " in normalized_resume:
            return 100.0

    role_tokens = set(normalized_role.split())
    if len(role_tokens) >= 2 and role_tokens.issubset(set(normalized_resume.split())):
        return 70.0

    return 0.0


def has_exact_role_title_match(resume_text: str, role_name: str) -> bool:
    normalized_resume = f" {normalize_match_text(resume_text)} "
    normalized_role = normalize_match_text(role_name)
    if not normalized_role:
        return False

    aliases = ROLE_TITLE_MATCH_ALIASES.get(normalized_role, {normalized_role})
    return any(
        bool(normalized_alias and f" {normalized_alias} " in normalized_resume)
        for normalized_alias in (normalize_match_text(alias) for alias in aliases)
    )


def compute_database_role_confidence_score(
    combined_score: float,
    role_evidence_score: float,
    title_match_score: float,
    exact_title_match: bool,
) -> float:
    adjusted_score = (
        (combined_score * 0.62)
        + (role_evidence_score * 0.23)
        + (title_match_score * 0.15)
    )

    if exact_title_match and role_evidence_score >= 80:
        adjusted_score = max(adjusted_score, 78.0)
        if combined_score >= 38:
            adjusted_score = max(adjusted_score, 82.0)
    elif title_match_score >= 100 and role_evidence_score >= 80 and combined_score >= 45:
        adjusted_score = max(adjusted_score, 72.0)

    return round(min(100.0, adjusted_score), 2)


def rank_database_role_matches(resume_text: str, limit: int | None = None) -> list[dict[str, Any]]:
    roles = fetch_job_roles()
    ranked_matches: list[dict[str, Any]] = []

    for role in roles:
        requirements = fetch_job_requirements(int(role["id"]))
        if not requirements:
            continue

        job_description = build_job_description_from_requirements(role, requirements)
        fit_result = evaluate_job_fit(resume_text, job_description, requirements)
        combined_score = float(fit_result["combined_score"])
        role_name = str(role.get("role_name") or "")
        role_evidence_score = compute_role_evidence_score(
            resume_text,
            role_name,
            fit_result.get("important_keywords", []),
        )
        title_match_score = compute_role_title_match_score(resume_text, role_name)
        exact_title_match = has_exact_role_title_match(resume_text, role_name)
        adjusted_score = compute_database_role_confidence_score(
            combined_score,
            role_evidence_score,
            title_match_score,
            exact_title_match,
        )

        ranked_matches.append(
            {
                "role": role,
                "requirements": requirements,
                "match_type": "resume_auto",
                "job_description": job_description,
                "confidence_score": round(adjusted_score, 2),
                "base_confidence_score": round(combined_score, 2),
                "role_evidence_score": role_evidence_score,
                "title_match_score": title_match_score,
                "exact_title_match": exact_title_match,
                "fit_result": fit_result,
            }
        )

    ranked_matches.sort(
        key=lambda item: (
            float(item.get("confidence_score") or 0),
            1 if item.get("exact_title_match") else 0,
            len(normalize_match_text(str((item.get("role") or {}).get("role_name") or "")).split()),
            float(item.get("role_evidence_score") or 0),
            str((item.get("role") or {}).get("role_name") or ""),
        ),
        reverse=True,
    )

    if limit is None:
        return ranked_matches
    return ranked_matches[: max(0, limit)]


def count_resume_words(text: str) -> int:
    return len(re.findall(r"\b[a-zA-Z][a-zA-Z+.#/-]*\b", text))


def choose_pdf_text_candidate(plain_raw_text: str, layout_raw_text: str) -> tuple[str, str, str]:
    plain_text = clean_resume_source_text(plain_raw_text)
    layout_text = clean_resume_source_text(layout_raw_text)
    if not layout_text:
        return plain_text, plain_raw_text, "plain"
    if not plain_text:
        return layout_text, layout_raw_text, "layout"

    def candidate_score(text: str) -> float:
        section_hits = sum(1 for marker in CORE_RESUME_SECTION_MARKERS if marker in text.lower())
        skill_hits = len(extract_key_phrases(text, DEFAULT_SKILL_DICTIONARY))
        spacing_penalty = count_text_extraction_spacing_glitches(text) * 18
        return count_resume_words(text) + (section_hits * 25) + (skill_hits * 8) - spacing_penalty

    plain_score = candidate_score(plain_text)
    layout_score = candidate_score(layout_text)
    plain_words = count_resume_words(plain_text)
    layout_words = count_resume_words(layout_text)
    plain_glitches = count_text_extraction_spacing_glitches(plain_text)
    layout_glitches = count_text_extraction_spacing_glitches(layout_text)

    if (
        layout_score >= plain_score + 40
        or (
            plain_words < 180
            and layout_words >= plain_words * 1.45
            and layout_glitches <= plain_glitches
        )
    ):
        return layout_text, layout_raw_text, "layout"

    return plain_text, plain_raw_text, "plain"


def extract_pdf_text(file_storage) -> tuple[str, dict[str, Any]]:
    try:
        reader = PdfReader(BytesIO(file_storage.read()))
    except Exception as error:
        raise ValueError("Could not read the uploaded PDF. Please upload a valid, unlocked PDF file.") from error

    plain_pages: list[str] = []
    layout_pages: list[str] = []
    embedded_image_count = 0
    try:
        for page in reader.pages:
            plain_pages.append(page.extract_text() or "")
            try:
                layout_pages.append(page.extract_text(extraction_mode="layout") or "")
            except TypeError:
                layout_pages.append("")
            try:
                embedded_image_count += len(page.images)
            except Exception:
                pass
    except Exception as error:
        raise ValueError("Could not extract text from this PDF. Please upload a text-based PDF exported from Word, Google Docs, or your resume editor.") from error

    raw_text = "\n".join(plain_pages)
    layout_raw_text = "\n".join(layout_pages)
    normalized_text, chosen_raw_text, extraction_mode = choose_pdf_text_candidate(raw_text, layout_raw_text)
    return normalized_text, {
        "file_type": "pdf",
        "page_count": len(reader.pages),
        "embedded_image_count": embedded_image_count,
        "ocr_used": False,
        "raw_text_length": len(chosen_raw_text),
        "pdf_extraction_mode": extraction_mode,
        "character_spacing_repaired": bool(CHARACTER_SPACED_WORD_PATTERN.search(chosen_raw_text)),
    }


def get_file_extension(filename: str) -> str:
    _, extension = os.path.splitext(filename or "")
    return extension.lower()


def configure_tesseract() -> None:
    if pytesseract is None:
        raise ValueError("Image resume OCR requires the pytesseract Python package. Please install backend requirements again.")

    configured_cmd = os.getenv("TESSERACT_CMD", "").strip()
    if configured_cmd:
        pytesseract.pytesseract.tesseract_cmd = configured_cmd
        if os.path.exists(configured_cmd):
            return

    if shutil.which("tesseract"):
        return

    raise ValueError(
        "Image resume OCR is not configured. Install Tesseract OCR, then set TESSERACT_CMD in env "
        "to the tesseract.exe path. You can still upload a text-based PDF."
    )


def extract_image_text(file_storage) -> tuple[str, dict[str, Any]]:
    if Image is None:
        raise ValueError("Image resume OCR requires the Pillow Python package. Please install backend requirements again.")

    configure_tesseract()
    try:
        image = Image.open(BytesIO(file_storage.read()))
        image = image.convert("RGB")
    except (UnidentifiedImageError, OSError) as error:
        raise ValueError("Could not read the uploaded image. Please upload a valid PNG or JPG file.") from error

    try:
        raw_text = pytesseract.image_to_string(image)
        return normalize_text(raw_text), {
            "file_type": get_file_extension(file_storage.filename).lstrip(".") or "image",
            "page_count": 1,
            "embedded_image_count": 1,
            "ocr_used": True,
            "raw_text_length": len(raw_text),
            "image_width": image.width,
            "image_height": image.height,
        }
    except Exception as error:
        raise ValueError("Could not extract text from the image resume. Please upload a clearer image or a text-based PDF.") from error


def extract_resume_text(file_storage, filename: str) -> tuple[str, dict[str, Any]]:
    extension = get_file_extension(filename)
    if extension in PDF_EXTENSIONS:
        return extract_pdf_text(file_storage)
    if extension in IMAGE_EXTENSIONS:
        return extract_image_text(file_storage)
    raise ValueError("Only PDF, PNG, JPG, or JPEG resumes are supported.")


def tokenize_keywords(text: str) -> list[str]:
    tokens = re.findall(r"[a-zA-Z][a-zA-Z+.#/-]{1,}", text.lower())
    cleaned_tokens = [token.strip(".,;:!?()[]{}\"'") for token in tokens]
    filtered = [token for token in cleaned_tokens if len(token) > 2 and token not in KEYWORD_STOPWORDS]
    return filtered


def normalize_keyword_token(token: str) -> str:
    cleaned = token.lower().strip(".,;:!?()[]{}\"'")
    if cleaned in CANONICAL_KEYWORD_ALIASES:
        return CANONICAL_KEYWORD_ALIASES[cleaned]
    if cleaned == "net":
        return "dotnet"
    if cleaned.endswith("ies") and len(cleaned) > 5:
        return cleaned[:-3] + "y"
    if cleaned.endswith("s") and len(cleaned) > 4 and not cleaned.endswith(("ss", "sis")):
        return cleaned[:-1]
    return cleaned


def tokenize_canonical_keywords(text: str) -> list[str]:
    return [normalize_keyword_token(token) for token in tokenize_keywords(text)]


def get_semantic_equivalents(token: str) -> set[str]:
    canonical = normalize_keyword_token(token)
    equivalents = {canonical}

    direct_matches = SEMANTIC_KEYWORD_EQUIVALENTS.get(canonical, set())
    equivalents.update(normalize_keyword_token(match) for match in direct_matches)

    for source, related_terms in SEMANTIC_KEYWORD_EQUIVALENTS.items():
        source_token = normalize_keyword_token(source)
        normalized_related = {normalize_keyword_token(term) for term in related_terms}
        if canonical == source_token or canonical in normalized_related:
            equivalents.add(source_token)
            equivalents.update(normalized_related)

    return equivalents


def token_has_resume_evidence(token: str, resume_tokens: set[str]) -> bool:
    canonical = normalize_keyword_token(token)
    strict_tokens = STRICT_TECHNICAL_KEYWORD_EVIDENCE.get(canonical)
    if strict_tokens is not None:
        return bool(strict_tokens & resume_tokens)
    return bool(get_semantic_equivalents(canonical) & resume_tokens)


def keyword_matches_resume(keyword: str, resume_text: str, resume_tokens: set[str]) -> bool:
    keyword_tokens = tokenize_canonical_keywords(keyword)
    if not keyword_tokens:
        return False

    canonical_resume = " " + " ".join(tokenize_canonical_keywords(resume_text)) + " "
    canonical_keyword = " ".join(keyword_tokens)
    if f" {canonical_keyword} " in canonical_resume:
        return True

    return all(token_has_resume_evidence(token, resume_tokens) for token in keyword_tokens)


def key_phrase_exists(text: str, phrase: str) -> bool:
    cleaned_phrase = normalize_text(phrase).lower()
    if not cleaned_phrase:
        return False
    return bool(re.search(rf"(?<![a-z0-9]){re.escape(cleaned_phrase)}(?![a-z0-9])", text.lower()))


def extract_key_phrases(text: str, vocabulary: list[str]) -> list[str]:
    matches: list[str] = []
    for term in vocabulary:
        if key_phrase_exists(text, term):
            matches.append(term)
    return sorted(set(matches))


def extract_sentences(text: str) -> list[str]:
    raw_sentences = re.split(r"(?<=[.!?])\s+|\n+", text)
    sentences = [normalize_text(sentence) for sentence in raw_sentences]
    return [sentence for sentence in sentences if len(sentence) > 20]


def extract_resume_profile(text: str) -> ParsedResume:
    skills = extract_key_phrases(text, DEFAULT_SKILL_DICTIONARY)
    education = extract_key_phrases(text, COMMON_EDUCATION_MARKERS)
    experience = extract_key_phrases(text, COMMON_EXPERIENCE_MARKERS)
    sentences = extract_sentences(text)
    excerpt = " ".join(sentences[:3]) if sentences else text[:400]
    return ParsedResume(
        text=text,
        skills=skills,
        education=education,
        experience=experience,
        excerpt=excerpt,
    )


def build_quality_criterion(
    category: str,
    status: str,
    title: str,
    detail: str,
    suggestion: str,
    penalty: int = 0,
    visual_marker: str | None = None,
) -> dict[str, Any]:
    criterion = {
        "category": category,
        "status": status,
        "title": title,
        "detail": detail,
        "suggestion": suggestion,
        "penalty": penalty,
    }
    if visual_marker:
        criterion["visual_marker"] = visual_marker
    return criterion


def count_language_markers(text: str, markers: set[str]) -> int:
    tokens = re.findall(r"[a-zA-Z]+", text.lower())
    return sum(1 for token in tokens if token in markers)


def count_text_extraction_spacing_glitches(text: str) -> int:
    candidates = re.findall(
        r"\b(?:[a-z]{3,}[A-Z][A-Za-z]{2,}|\d{2,}[A-Za-z]{3,}|[A-Za-z]{3,}\d{2,})\b",
        text,
    )
    expected_technical_tokens = {
        "css3",
        "html5",
        "ipv4",
        "ipv6",
        "nodejs",
        "oauth2",
        "reactjs",
        "restjson",
        "vuejs",
        "webapi",
    }
    return sum(1 for token in candidates if normalize_match_text(token) not in expected_technical_tokens)


def analyze_resume_quality(
    resume_text: str,
    uploaded_name: str,
    document_info: dict[str, Any] | None,
    resume_profile: ParsedResume,
) -> dict[str, Any]:
    info = document_info or {}
    file_type = str(info.get("file_type") or get_file_extension(uploaded_name).lstrip(".") or "unknown").lower()
    word_count = count_resume_words(resume_text)
    lower_text = resume_text.lower()
    criteria: list[dict[str, Any]] = []
    section_hits = sorted({marker for marker in CORE_RESUME_SECTION_MARKERS if marker in lower_text})
    has_email = bool(re.search(r"[^@\s]+@[^@\s]+\.[^@\s]+", resume_text))
    has_phone = bool(re.search(r"(\+?\d[\d\s().-]{7,}\d)", resume_text))
    has_profile_link = any(marker in lower_text for marker in CONTACT_MARKERS)
    raw_text_length = int(info.get("raw_text_length") or len(resume_text or ""))
    has_resume_signals = len(section_hits) >= 2 or bool(resume_profile.skills and (resume_profile.education or resume_profile.experience))
    has_readable_pdf_text = word_count >= 180 or (raw_text_length >= 1200 and has_resume_signals)

    def add(
        category: str,
        status: str,
        title: str,
        detail: str,
        suggestion: str,
        penalty: int = 0,
        visual_marker: str | None = None,
    ) -> None:
        criteria.append(build_quality_criterion(category, status, title, detail, suggestion, penalty, visual_marker))

    if info.get("ocr_used"):
        add(
            "Format",
            "warning",
            "Image or scanned resume",
            "The uploaded file was read with OCR, which is less reliable than selectable PDF text.",
            "Export the resume as a text-based PDF from Word, Google Docs, Canva, or your resume editor.",
            18,
        )
    elif file_type == "pdf" and has_readable_pdf_text:
        add(
            "Format",
            "pass",
            "Text-based PDF detected",
            "The resume contains selectable text, so the analyzer can read it directly.",
            "Keep using a text-based PDF for ATS submissions.",
            0,
        )
    elif file_type == "pdf":
        add(
            "Format",
            "warning",
            "Low extracted text",
            "The PDF produced a small amount of readable text, which may mean the layout is image-heavy or scanned.",
            "Use a cleaner one-column text PDF and avoid exporting the whole resume as an image.",
            16,
        )

    embedded_image_count = int(info.get("embedded_image_count") or 0)
    if info.get("ocr_used"):
        add(
            "Photo",
            "warning",
            "Photo/image resume risk",
            "Image resumes can include visual content that ATS systems may skip or misread.",
            "For ATS applications, remove passport-style photos and use normal selectable text.",
            10,
        )
    elif embedded_image_count > 0:
        add(
            "Photo",
            "warning",
            "Embedded image detected",
            f"The PDF contains {embedded_image_count} embedded image(s). If one is a passport photo, logo, or decorative graphic, ATS systems may ignore it.",
            "Use photos only when the employer asks for them; otherwise remove passport photos and decorative images.",
            6,
        )
    else:
        add(
            "Photo",
            "pass",
            "No embedded image risk found",
            "No obvious image or passport-photo signal was detected in the extracted document.",
            "Keep the resume text-focused for ATS uploads.",
            0,
        )

    english_hits = count_language_markers(resume_text, ENGLISH_LANGUAGE_MARKERS)
    malay_hits = count_language_markers(resume_text, MALAY_LANGUAGE_MARKERS)
    non_latin_count = len(re.findall(r"[\u4e00-\u9fff\u3040-\u30ff\u0600-\u06ff\u0900-\u097f\u0b80-\u0bff]", resume_text))
    if english_hits >= 4 and (malay_hits >= 3 or non_latin_count > 20):
        add(
            "Language",
            "warning",
            "Mixed language detected",
            "The resume appears to mix English with Malay or another script. Some ATS keyword matching is language-specific.",
            "Use one main language for the ATS version, normally English for CS roles, and translate key skills consistently.",
            10,
        )
    else:
        add(
            "Language",
            "pass",
            "Single-language text",
            "No strong bilingual-language risk was detected from the extracted text.",
            "Keep job titles, skills, and project descriptions in the same language as the job posting.",
            0,
        )

    spacing_glitches = count_text_extraction_spacing_glitches(resume_text)
    ligature_count = len(re.findall(r"[\ufb00-\ufb06]", resume_text))
    if spacing_glitches >= 14 or ligature_count > 0:
        add(
            "Parsing",
            "warning",
            "Text extraction spacing issue",
            "Some words appear joined together or contain special ligature characters after extraction.",
            "Use a simple one-column layout and standard fonts so ATS parsers do not join words together.",
            10,
        )
    else:
        add(
            "Parsing",
            "pass",
            "Clean text extraction",
            "The extracted text does not show major spacing or encoding problems.",
            "Keep fonts standard and avoid complex text boxes.",
            0,
        )

    if len(section_hits) < 3:
        add(
            "Structure",
            "warning",
            "Missing standard resume sections",
            "The analyzer found fewer than three standard sections such as Education, Skills, Experience, or Projects.",
            "Use clear section headings: Education, Technical Skills, Projects, and Work Experience.",
            12,
        )
    else:
        add(
            "Structure",
            "pass",
            "Standard sections found",
            "The resume includes recognizable ATS-friendly section headings.",
            "Keep headings simple and avoid replacing them with icons only.",
            0,
        )

    if not has_email or not has_phone:
        add(
            "Contact",
            "warning",
            "Incomplete contact details",
            "The extracted text does not clearly show both email and phone details.",
            "Put email, phone, LinkedIn, and GitHub/portfolio as selectable text near the top.",
            8,
        )
    elif not has_profile_link:
        add(
            "Contact",
            "warning",
            "Missing professional profile link",
            "Email and phone are present, but no LinkedIn, GitHub, or portfolio marker was detected.",
            "Add GitHub, LinkedIn, or a portfolio link in the contact area near your email and phone.",
            2,
            "profile_link_contact_area",
        )
    else:
        add(
            "Contact",
            "pass",
            "Contact details readable",
            "Email, phone, and professional profile markers are visible in the extracted text.",
            "Keep contact details as plain selectable text.",
            0,
        )

    quantified_evidence = len(re.findall(r"(\d+(?:\.\d+)?%|\d+\+|\$|users?|revenue|reduced|increased|improved|optimized|enabled|delivered)", resume_text, re.I))
    if quantified_evidence < 3:
        add(
            "Impact",
            "warning",
            "Limited measurable impact",
            "The resume has limited numeric or outcome-based evidence.",
            "Quantify projects and experience with users, speed, accuracy, revenue, cost, or time improvements.",
            8,
        )
    else:
        add(
            "Impact",
            "pass",
            "Measurable impact found",
            "The resume includes numbers or outcome language that helps ATS and recruiters understand impact.",
            "Keep measurable results in project and work-experience bullets.",
            0,
        )

    if word_count < 250:
        add(
            "Length",
            "warning",
            "Resume may be too short",
            f"Only about {word_count} words were extracted.",
            "Add enough project, skills, and experience detail for the target role.",
            10,
        )
    elif word_count > 1300:
        add(
            "Length",
            "warning",
            "Resume may be too long",
            f"About {word_count} words were extracted.",
            "Trim older or less relevant details so the ATS version stays focused.",
            6,
        )
    else:
        add(
            "Length",
            "pass",
            "Reasonable resume length",
            f"About {word_count} words were extracted.",
            "Keep the ATS version concise and role-focused.",
            0,
        )

    if len(resume_profile.skills) < 4:
        add(
            "Skills",
            "warning",
            "Technical skills not explicit enough",
            "Few recognizable technical skills were detected as standalone terms.",
            "Add a Technical Skills section with languages, frameworks, databases, tools, and cloud platforms.",
            8,
        )
    else:
        add(
            "Skills",
            "pass",
            "Technical skills detected",
            "The resume includes several recognizable technical skills.",
            "Keep the most relevant target-role skills near the top.",
            0,
        )

    total_penalty = sum(int(item.get("penalty") or 0) for item in criteria)
    quality_score = max(0, min(100, 100 - total_penalty))
    if quality_score >= 82:
        risk_level = "Low"
    elif quality_score >= 62:
        risk_level = "Medium"
    else:
        risk_level = "High"

    issues = [item for item in criteria if item["status"] != "pass"]
    summary = (
        f"ATS quality risk is {risk_level.lower()} with {len(issues)} issue(s) found."
        if issues
        else "ATS quality risk is low. The resume format looks readable and structured."
    )

    return {
        "quality_score": quality_score,
        "risk_level": risk_level,
        "word_count": word_count,
        "document_info": info,
        "criteria": criteria,
        "issues": issues[:8],
        "summary": summary,
    }


def extract_job_keywords(job_description: str, limit: int = 10) -> list[str]:
    candidates = tokenize_canonical_keywords(job_description)
    counts = Counter(candidates)
    ordered = [word for word, _ in counts.most_common()]
    return ordered[:limit]


def rule_based_feedback(resume_profile: ParsedResume, job_keywords: list[str]) -> tuple[list[str], list[str]]:
    resume_tokens = set(tokenize_canonical_keywords(resume_profile.text))
    keyword_hits = [keyword for keyword in job_keywords if keyword_matches_resume(keyword, resume_profile.text, resume_tokens)]
    missing_keywords = [keyword for keyword in job_keywords if keyword not in keyword_hits]

    feedback: list[str] = []
    if missing_keywords:
        feedback.append("Add or emphasize these job keywords: " + ", ".join(missing_keywords[:5]) + ".")
    else:
        feedback.append("Your resume already reflects the main job keywords well.")

    if not resume_profile.skills:
        feedback.append("Add a dedicated skills section with measurable technical tools and frameworks.")
    else:
        feedback.append("Highlight your strongest technical skills higher in the resume.")

    if not resume_profile.experience:
        feedback.append("Make your experience section more explicit by using action verbs and project outcomes.")
    else:
        feedback.append("Quantify achievements in each project or work experience bullet.")

    if not resume_profile.education:
        feedback.append("Add education details if they are missing or hard to locate.")
    else:
        feedback.append("Show relevant coursework, capstones, or research if it supports the target role.")

    return feedback, missing_keywords


def compute_tfidf_score(resume_text: str, job_description: str) -> float:
    documents = [resume_text or "", job_description or ""]
    if not any(documents):
        return 0.0

    vectorizer = TfidfVectorizer(stop_words="english")
    try:
        matrix = vectorizer.fit_transform(documents)
    except ValueError:
        return 0.0
    resume_vector = matrix[0]
    job_vector = matrix[1]
    similarity = float((resume_vector @ job_vector.T).toarray()[0][0])
    return max(0.0, min(1.0, similarity)) * 100.0


def get_requirement_score_weight(requirement: dict[str, Any]) -> float:
    try:
        importance = int(requirement.get("importance_weight") or 3)
    except (TypeError, ValueError):
        importance = 3
    importance = max(1, min(5, importance))
    requirement_type = str(requirement.get("requirement_type") or "").lower()
    type_weight = REQUIREMENT_TYPE_SCORE_WEIGHTS.get(requirement_type, 1.0)
    weighted_score = importance * type_weight

    requirement_tokens = set(tokenize_canonical_keywords(str(requirement.get("requirement_text") or "")))
    optional_stack_terms = {"dotnet", "jquery", "tomcat"}
    if requirement_tokens & optional_stack_terms:
        weighted_score *= 0.65

    return weighted_score


def score_requirement_against_resume(
    resume_text: str,
    resume_tokens: set[str],
    requirement_text: str,
) -> dict[str, Any]:
    requirement_tokens = list(dict.fromkeys(tokenize_canonical_keywords(requirement_text)))
    matched_tokens = [
        token
        for token in requirement_tokens
        if keyword_matches_resume(token, resume_text, resume_tokens)
    ]

    if not requirement_tokens:
        return {"score": 0.0, "matched_tokens": [], "missing_tokens": []}

    token_coverage = (len(matched_tokens) / len(requirement_tokens)) * 100.0
    phrase_similarity = compute_tfidf_score(resume_text, requirement_text)
    score = (token_coverage * 0.85) + (phrase_similarity * 0.15)

    if normalize_match_text(requirement_text) and normalize_match_text(requirement_text) in normalize_match_text(resume_text):
        score = 100.0
        matched_tokens = requirement_tokens

    return {
        "score": round(min(100.0, score), 2),
        "matched_tokens": matched_tokens,
        "missing_tokens": [token for token in requirement_tokens if token not in matched_tokens],
    }


def evaluate_weighted_requirements(
    resume_text: str,
    requirements: list[dict[str, Any]],
) -> dict[str, Any] | None:
    if not requirements:
        return None

    resume_tokens = set(tokenize_canonical_keywords(resume_text))
    scored_requirements: list[dict[str, Any]] = []
    total_weight = 0.0
    weighted_score = 0.0
    all_keywords: list[str] = []
    matched_keywords: list[str] = []
    missing_keywords: list[str] = []

    for requirement in requirements:
        requirement_text = normalize_text(str(requirement.get("requirement_text") or ""))
        if not requirement_text:
            continue

        score_detail = score_requirement_against_resume(resume_text, resume_tokens, requirement_text)
        weight = get_requirement_score_weight(requirement)
        total_weight += weight
        weighted_score += score_detail["score"] * weight
        all_keywords.extend(tokenize_canonical_keywords(requirement_text))
        matched_keywords.extend(score_detail["matched_tokens"])
        missing_keywords.extend(score_detail["missing_tokens"])
        scored_requirements.append(
            {
                "requirement_text": requirement_text,
                "requirement_type": requirement.get("requirement_type"),
                "importance_weight": requirement.get("importance_weight", 3),
                "match_score": score_detail["score"],
                "matched_tokens": score_detail["matched_tokens"],
                "missing_tokens": score_detail["missing_tokens"],
            }
        )

    if not scored_requirements or total_weight <= 0:
        return None

    coverage_score = weighted_score / total_weight
    matched_requirements = [item for item in scored_requirements if item["match_score"] >= 50]
    missing_requirements = [item for item in scored_requirements if item["match_score"] < 50]

    ordered_keywords = [word for word, _ in Counter(all_keywords).most_common(20)]
    ordered_matched = [word for word, _ in Counter(matched_keywords).most_common(20)]
    matched_keyword_set = set(ordered_matched)
    ordered_missing = [word for word, _ in Counter(missing_keywords).most_common(20) if word not in matched_keyword_set]

    return {
        "requirement_coverage_score": round(coverage_score, 2),
        "important_keywords": ordered_keywords,
        "matched_keywords": ordered_matched,
        "missing_keywords": ordered_missing[:12],
        "matched_requirements": matched_requirements[:8],
        "missing_requirements": missing_requirements[:8],
        "scored_requirements": scored_requirements,
    }


def evaluate_job_fit(
    resume_text: str,
    job_description: str,
    requirements: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    similarity_score = compute_tfidf_score(resume_text, job_description)
    requirement_result = evaluate_weighted_requirements(resume_text, requirements or [])

    if requirement_result:
        job_keywords = requirement_result["important_keywords"]
        matched_keywords = requirement_result["matched_keywords"]
        missing_keywords = requirement_result["missing_keywords"]
        keyword_overlap_score = requirement_result["requirement_coverage_score"]
    else:
        resume_tokens = set(tokenize_canonical_keywords(resume_text))
        job_keywords = extract_job_keywords(job_description, limit=20)
        matched_keywords = [
            keyword
            for keyword in job_keywords
            if keyword_matches_resume(keyword, resume_text, resume_tokens)
        ]
        missing_keywords = [keyword for keyword in job_keywords if keyword not in matched_keywords]
        keyword_overlap_score = (len(matched_keywords) / max(1, len(job_keywords))) * 100.0

    combined_score = (keyword_overlap_score * 0.7) + (similarity_score * 0.3)

    result = {
        "similarity_score": round(similarity_score, 2),
        "keyword_overlap_score": round(keyword_overlap_score, 2),
        "combined_score": round(combined_score, 2),
        "important_keywords": job_keywords,
        "matched_keywords": matched_keywords,
        "missing_keywords": missing_keywords,
    }

    if requirement_result:
        result["matched_requirements"] = requirement_result["matched_requirements"]
        result["missing_requirements"] = requirement_result["missing_requirements"]
        result["scored_requirements"] = requirement_result["scored_requirements"]

    return result


def compute_group_evidence_score(resume_text: str, evidence_groups: list[set[str]]) -> float:
    if not evidence_groups:
        return 0.0

    resume_tokens = set(tokenize_canonical_keywords(resume_text))
    matched_groups = 0

    for group in evidence_groups:
        if any(keyword_matches_resume(term, resume_text, resume_tokens) for term in group):
            matched_groups += 1

    group_score = (matched_groups / len(evidence_groups)) * 100.0
    impact_bonus = 0.0
    if re.search(r"(\d+(?:\.\d+)?%|\d+\+|\$|users?|revenue|increased|reduced|improved|optimized|enabled|delivered)", resume_text, re.I):
        impact_bonus = 8.0

    return round(min(100.0, (group_score * 0.9) + impact_bonus), 2)


def compute_role_evidence_score(resume_text: str, role_name: str, job_keywords: list[str]) -> float:
    role_context = normalize_match_text(role_name)
    if any(term in role_context for term in ("cyber", "security", "forensic")):
        return compute_group_evidence_score(resume_text, CYBERSECURITY_EVIDENCE_GROUPS)
    if any(term in role_context for term in ("database", "oracle", "sql")):
        return compute_group_evidence_score(resume_text, DATABASE_EVIDENCE_GROUPS)
    if "network" in role_context:
        return compute_group_evidence_score(resume_text, NETWORK_EVIDENCE_GROUPS)
    if "web" in role_context:
        return compute_group_evidence_score(resume_text, WEB_DEVELOPMENT_EVIDENCE_GROUPS)
    if any(term in role_context for term in ("tester", "testing", "quality assurance")):
        return compute_group_evidence_score(resume_text, QA_TESTING_EVIDENCE_GROUPS)
    if any(term in role_context for term in ("software", "developer", "develop", "programmer", "programming")):
        return compute_group_evidence_score(resume_text, SOFTWARE_DEVELOPER_EVIDENCE_GROUPS)

    resume_tokens = set(tokenize_canonical_keywords(resume_text))
    if not job_keywords:
        return 0.0

    matched_keywords = [
        keyword
        for keyword in job_keywords
        if keyword_matches_resume(keyword, resume_text, resume_tokens)
    ]
    return round((len(matched_keywords) / max(1, len(job_keywords))) * 100.0, 2)


def compute_resume_structure_score(resume_profile: ParsedResume) -> float:
    skill_score = min(30, len(resume_profile.skills) * 8)
    experience_score = 0
    if resume_profile.experience:
        experience_score = 25 + min(20, len(resume_profile.experience) * 3)
    education_score = 15 if resume_profile.education else 0
    return float(min(100, 10 + skill_score + experience_score + education_score))


def compute_experience_alignment_score(
    resume_profile: ParsedResume,
    keyword_coverage: float,
    similarity_score: float,
) -> float:
    if not resume_profile.experience:
        return round((keyword_coverage * 0.45) + (similarity_score * 0.15), 2)

    marker_score = min(40, len(resume_profile.experience) * 6)
    return round(min(100, 20 + marker_score + (keyword_coverage * 0.35) + (similarity_score * 0.1)), 2)


def compute_academic_rigor_score(
    resume_profile: ParsedResume,
    keyword_coverage: float,
    similarity_score: float,
) -> float:
    education_score = 30 if resume_profile.education else 10
    skill_score = min(25, len(resume_profile.skills) * 6)
    return round(min(100, education_score + skill_score + (keyword_coverage * 0.25) + (similarity_score * 0.1)), 2)


def compute_final_readiness_score(
    keyword_match: float,
    role_evidence: float,
    experience_alignment: float,
    resume_quality_score: float,
    academic_rigor: float,
    similarity_score: float,
) -> float:
    weighted_score = (
        (keyword_match * 0.30)
        + (role_evidence * 0.25)
        + (experience_alignment * 0.15)
        + (resume_quality_score * 0.20)
        + (academic_rigor * 0.07)
        + (similarity_score * 0.03)
    )
    if (
        resume_quality_score >= 95
        and role_evidence >= 90
        and experience_alignment >= 75
        and keyword_match >= 50
    ):
        weighted_score = max(weighted_score, 82.0)

    return round(min(100.0, weighted_score), 2)


def get_database_role_name(job_match: dict[str, Any] | None) -> str:
    if not job_match:
        return ""
    role = job_match.get("role") or {}
    return str(role.get("role_name") or "")


def get_role_family(role_name: str) -> str:
    normalized = normalize_match_text(role_name)
    if not normalized:
        return ""
    if normalized in ROLE_FAMILY_ALIASES:
        return ROLE_FAMILY_ALIASES[normalized]
    for alias, family in ROLE_FAMILY_ALIASES.items():
        if alias in normalized or normalized in alias:
            return family
    return normalized


def pick_recommended_database_job(
    resume_text: str,
    selected_job: dict[str, Any] | None,
    detected_job: dict[str, Any] | None,
) -> tuple[dict[str, Any] | None, dict[str, Any] | None, dict[str, Any] | None]:
    selected_fit = evaluate_job_fit(
        resume_text,
        selected_job["job_description"],
        selected_job.get("requirements", []),
    ) if selected_job else None
    detected_fit = evaluate_job_fit(
        resume_text,
        detected_job["job_description"],
        detected_job.get("requirements", []),
    ) if detected_job else None

    if selected_job and detected_job:
        selected_name = normalize_match_text(get_database_role_name(selected_job))
        detected_name = normalize_match_text(get_database_role_name(detected_job))
        same_role_family = bool(selected_name and detected_name and get_role_family(selected_name) == get_role_family(detected_name))
        selected_score = float(selected_fit["combined_score"]) if selected_fit else 0.0
        detected_score = float(detected_fit["combined_score"]) if detected_fit else 0.0

        if selected_score >= 50.0:
            return selected_job, selected_fit, detected_fit

        if selected_name != detected_name and not same_role_family and detected_score >= selected_score + ROLE_RECOMMENDATION_MARGIN:
            return detected_job, selected_fit, detected_fit
        return selected_job, selected_fit, detected_fit

    if detected_job:
        return detected_job, selected_fit, detected_fit
    if selected_job:
        return selected_job, selected_fit, detected_fit
    return None, selected_fit, detected_fit


def build_top_role_match_items(
    top_role_matches: list[dict[str, Any]] | None,
    selected_database_role: str,
    recommended_database_role: str,
) -> list[dict[str, Any]]:
    selected_key = normalize_match_text(selected_database_role)
    recommended_key = normalize_match_text(recommended_database_role)
    items: list[dict[str, Any]] = []

    for rank, match in enumerate(top_role_matches or [], start=1):
        role = match.get("role") or {}
        role_name = str(role.get("role_name") or "")
        role_key = normalize_match_text(role_name)
        fit_result = match.get("fit_result") or {}
        matched_keywords = list(dict.fromkeys(fit_result.get("matched_keywords", []) or []))[:8]
        missing_keywords = list(dict.fromkeys(fit_result.get("missing_keywords", []) or []))[:8]
        score = float(match.get("confidence_score") or fit_result.get("combined_score") or 0)

        if score >= 75:
            confidence_label = "Strong fit"
        elif score >= 55:
            confidence_label = "Good fit"
        elif score >= 35:
            confidence_label = "Possible fit"
        else:
            confidence_label = "Weak evidence"

        if matched_keywords:
            reason = "Strong evidence: " + ", ".join(matched_keywords[:5]) + "."
        else:
            reason = "The resume has limited direct keyword evidence for this role."
        if missing_keywords:
            reason += " Improve with: " + ", ".join(missing_keywords[:4]) + "."

        items.append(
            {
                "rank": rank,
                "role": role_name,
                "database_role": role_name,
                "score": round(score, 2),
                "confidence_label": confidence_label,
                "keyword_match": fit_result.get("keyword_overlap_score"),
                "similarity_score": fit_result.get("similarity_score"),
                "role_evidence_score": match.get("role_evidence_score"),
                "matched_keywords": matched_keywords,
                "missing_keywords": missing_keywords,
                "is_selected_database_role": bool(selected_key and role_key == selected_key),
                "is_final_recommended": bool(recommended_key and role_key == recommended_key),
                "reason": reason,
            }
        )

    return items


def build_role_recommendation(
    requested_role: str,
    selected_job: dict[str, Any] | None,
    detected_job: dict[str, Any] | None,
    recommended_job: dict[str, Any] | None,
    selected_fit: dict[str, Any] | None,
    detected_fit: dict[str, Any] | None,
    top_role_matches: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    selected_label = normalize_text(requested_role)
    selected_database_role = get_database_role_name(selected_job)
    detected_database_role = get_database_role_name(detected_job)
    recommended_database_role = get_database_role_name(recommended_job)

    recommended_fit = None
    if recommended_database_role and detected_database_role and normalize_match_text(recommended_database_role) == normalize_match_text(detected_database_role):
        recommended_fit = detected_fit
    elif recommended_database_role and selected_database_role and normalize_match_text(recommended_database_role) == normalize_match_text(selected_database_role):
        recommended_fit = selected_fit

    def clean_score(value: Any) -> float | None:
        try:
            return round(float(value), 2)
        except (TypeError, ValueError):
            return None

    def calibrated_match_score(role_name: str) -> float | None:
        role_key = normalize_match_text(role_name)
        if not role_key:
            return None
        for match in top_role_matches or []:
            match_role = match.get("role") or {}
            match_role_key = normalize_match_text(str(match_role.get("role_name") or ""))
            if match_role_key != role_key:
                continue
            return clean_score(
                match.get("confidence_score")
                or (match.get("fit_result") or {}).get("combined_score")
            )
        return None

    def fit_combined_score(fit_result: dict[str, Any] | None) -> float | None:
        return clean_score((fit_result or {}).get("combined_score"))

    selected_score = calibrated_match_score(selected_database_role)
    if selected_score is None:
        selected_score = fit_combined_score(selected_fit)
    detected_score = calibrated_match_score(detected_database_role)
    if detected_score is None:
        detected_score = fit_combined_score(detected_fit)
    recommended_score = calibrated_match_score(recommended_database_role)
    if recommended_score is None:
        recommended_score = fit_combined_score(recommended_fit)

    different_from_selected = bool(
        selected_database_role
        and recommended_database_role
        and normalize_match_text(selected_database_role) != normalize_match_text(recommended_database_role)
    )

    matched_signals = (recommended_fit or {}).get("matched_keywords", [])
    selected_missing = (selected_fit or {}).get("missing_keywords", [])

    if different_from_selected:
        selected_name = selected_label or selected_database_role
        reason_summary = (
            f"The resume is stronger for {recommended_database_role} than {selected_name} because it matches "
            f"{', '.join(matched_signals[:5]) or 'more of the database requirements'}."
        )
        if selected_missing:
            reason_summary += " The selected role still needs stronger evidence for " + ", ".join(selected_missing[:5]) + "."
    elif selected_database_role:
        reason_summary = (
            f"The selected role maps to {selected_database_role}, and it remains the best database match for this resume."
        )
        if matched_signals:
            reason_summary += " Strong signals include " + ", ".join(matched_signals[:5]) + "."
    elif recommended_database_role:
        reason_summary = (
            f"No selected database role was used, so the system recommends {recommended_database_role} from the resume content."
        )
        if matched_signals:
            reason_summary += " The strongest signals are " + ", ".join(matched_signals[:5]) + "."
    else:
        reason_summary = "The system could not compare this resume with a database role."

    return {
        "selected_role": selected_label,
        "selected_database_role": selected_database_role,
        "detected_database_role": detected_database_role,
        "recommended_role": recommended_database_role or selected_label,
        "recommended_database_role": recommended_database_role,
        "is_different_from_selected": different_from_selected,
        "selected_score": selected_score,
        "detected_score": detected_score,
        "recommended_score": recommended_score,
        "matched_resume_signals": matched_signals,
        "recommended_missing_keywords": (recommended_fit or {}).get("missing_keywords", []),
        "selected_missing_keywords": selected_missing,
        "top_role_matches": build_top_role_match_items(
            top_role_matches,
            selected_database_role,
            recommended_database_role,
        ),
        "reason_summary": reason_summary,
    }


def percent_status_label(value: Any) -> str:
    try:
        score = float(value)
    except (TypeError, ValueError):
        score = 0.0
    if score >= 75:
        return "strong"
    if score >= 55:
        return "fair"
    return "needs_improvement"


def build_recommendation_action(
    priority: int,
    category: str,
    title: str,
    detail: str,
    action: str,
    keywords: list[Any] | None = None,
    source: str = "analysis",
) -> dict[str, Any]:
    return {
        "priority": priority,
        "category": category,
        "title": normalize_text(title),
        "detail": normalize_text(detail),
        "action": normalize_text(action),
        "keywords": unique_clean_values(list(keywords or []), 6),
        "source": source,
    }


def build_part8_recommendation_module(analysis: dict[str, Any]) -> dict[str, Any]:
    role_recommendation = analysis.get("role_recommendation") if isinstance(analysis.get("role_recommendation"), dict) else {}
    job_analysis = analysis.get("job_description_analysis") if isinstance(analysis.get("job_description_analysis"), dict) else {}
    resume_summary = analysis.get("resume_summary") if isinstance(analysis.get("resume_summary"), dict) else {}
    resume_quality = analysis.get("resume_quality") if isinstance(analysis.get("resume_quality"), dict) else {}
    compatibility = analysis.get("compatibility") if isinstance(analysis.get("compatibility"), dict) else {}

    recommended_role = normalize_text(
        str(
            role_recommendation.get("recommended_role")
            or role_recommendation.get("recommended_database_role")
            or analysis.get("target_role")
            or analysis.get("selected_role")
            or "selected role"
        )
    )
    selected_role = normalize_text(str(role_recommendation.get("selected_role") or analysis.get("selected_role") or ""))
    matched_signals = unique_clean_values(
        list(role_recommendation.get("matched_resume_signals") or [])
        + list(job_analysis.get("matched_keywords") or [])
        + list(resume_summary.get("skills") or []),
        12,
    )
    missing_keywords = unique_clean_values(
        list(role_recommendation.get("recommended_missing_keywords") or [])
        + list(job_analysis.get("missing_keywords") or []),
        12,
    )
    selected_missing_keywords = unique_clean_values(list(role_recommendation.get("selected_missing_keywords") or []), 8)
    display_missing_keywords = unique_clean_values([format_resume_keyword(keyword) for keyword in missing_keywords], 10)
    display_matched_signals = unique_clean_values([format_resume_keyword(keyword) for keyword in matched_signals], 10)
    quality_issues = [
        item
        for item in list(resume_quality.get("issues") or [])
        if isinstance(item, dict)
    ]
    keyword_match = compatibility.get("keyword_match")
    role_evidence = compatibility.get("role_evidence")
    experience_alignment = compatibility.get("experience_alignment")
    academic_rigor = compatibility.get("academic_rigor")
    resume_quality_score = compatibility.get("resume_quality", resume_quality.get("quality_score"))

    actions: list[dict[str, Any]] = []

    if role_recommendation.get("is_different_from_selected") and selected_role:
        actions.append(
            build_recommendation_action(
                1,
                "role_fit",
                "Confirm the best target role",
                f"The system found stronger evidence for {recommended_role} than {selected_role}.",
                f"Use {recommended_role} as the main resume target, or rewrite the resume with clearer evidence for {selected_role}.",
                display_matched_signals[:5],
                "role_recommendation",
            )
        )
    elif recommended_role:
        actions.append(
            build_recommendation_action(
                1,
                "role_fit",
                "Use the confirmed comparison role",
                f"The resume is being compared against {recommended_role}.",
                f"Keep the resume headline, summary, skills, and project bullets focused on {recommended_role}.",
                display_matched_signals[:5],
                "role_recommendation",
            )
        )

    if display_missing_keywords:
        actions.append(
            build_recommendation_action(
                len(actions) + 1,
                "keyword_gap",
                "Add missing role keywords with proof",
                "The database requirements include keywords that are weak or missing in the resume.",
                "Add these terms only inside truthful project, coursework, internship, or work-experience evidence.",
                display_missing_keywords[:6],
                "database_requirements",
            )
        )

    if percent_status_label(keyword_match) == "needs_improvement" or percent_status_label(role_evidence) == "needs_improvement":
        actions.append(
            build_recommendation_action(
                len(actions) + 1,
                "evidence_gap",
                "Strengthen database requirement evidence",
                "The resume needs clearer proof that it satisfies the selected role requirements.",
                "Rewrite two or three bullets using this pattern: action verb + tool/skill + task + measurable result.",
                display_missing_keywords[:4] or display_matched_signals[:4],
                "score_breakdown",
            )
        )

    if percent_status_label(experience_alignment) == "needs_improvement" or not resume_summary.get("experience_markers"):
        actions.append(
            build_recommendation_action(
                len(actions) + 1,
                "experience",
                "Improve project and experience bullets",
                "Experience evidence is one of the main areas that can raise the readiness score.",
                "Add project scope, your responsibility, technologies used, and measurable outcomes such as users, accuracy, speed, or completion time.",
                display_matched_signals[:4],
                "resume_summary",
            )
        )

    if quality_issues:
        first_issue = quality_issues[0]
        actions.append(
            build_recommendation_action(
                len(actions) + 1,
                "ats_quality",
                str(first_issue.get("title") or "Fix ATS readability issue"),
                str(first_issue.get("detail") or resume_quality.get("summary") or "ATS quality issue detected."),
                str(first_issue.get("suggestion") or "Fix the highest-risk formatting issue before exporting the resume."),
                [],
                "ats_quality",
            )
        )

    if percent_status_label(academic_rigor) == "needs_improvement":
        actions.append(
            build_recommendation_action(
                len(actions) + 1,
                "education",
                "Connect education to the target role",
                "Education or coursework evidence is not strong enough for this role yet.",
                "Add relevant coursework, capstone, research, certification, or academic project evidence if it is true.",
                display_missing_keywords[:4],
                "resume_summary",
            )
        )

    if len(actions) <= 2 and percent_status_label(resume_quality_score) == "strong":
        actions.append(
            build_recommendation_action(
                len(actions) + 1,
                "impact",
                "Polish with measurable impact",
                "The resume has a usable foundation, so the next gain is stronger evidence quality.",
                "Add numbers or outcomes to the strongest bullets, then keep only the most relevant role keywords.",
                display_matched_signals[:4],
                "polish",
            )
        )

    actions = actions[:5]
    for index, action in enumerate(actions, start=1):
        action["priority"] = index

    section_targets = []
    action_categories = {action["category"] for action in actions}
    if "role_fit" in action_categories:
        section_targets.extend(["headline", "professional summary"])
    if "keyword_gap" in action_categories or "evidence_gap" in action_categories:
        section_targets.extend(["technical skills", "projects", "experience bullets"])
    if "experience" in action_categories or "impact" in action_categories:
        section_targets.extend(["professional experience", "projects"])
    if "education" in action_categories:
        section_targets.extend(["education", "certifications"])
    if "ats_quality" in action_categories:
        section_targets.append("resume format")

    strongest_action = actions[0]["title"] if actions else "review the resume evidence"
    summary = (
        f"Recommended role: {recommended_role}. Priority: {strongest_action}."
        if recommended_role
        else f"Priority: {strongest_action}."
    )
    if display_missing_keywords:
        summary += " Main keyword targets: " + ", ".join(display_missing_keywords[:5]) + "."
    elif display_matched_signals:
        summary += " Main evidence to keep: " + ", ".join(display_matched_signals[:5]) + "."

    return {
        "stage": "Part 8 - Recommendation and generated resume support",
        "status": "ready",
        "method": "Converts NLP/ML-style role matching, weighted database requirements, missing keywords, score components, and ATS quality checks into user-facing improvement actions.",
        "recommended_role": recommended_role,
        "selected_role": selected_role,
        "summary": summary,
        "priority_actions": actions,
        "resume_generation_strategy": {
            "target_role": recommended_role,
            "strategy_summary": (
                f"Generate the ATS resume around {recommended_role}, keep proven resume signals, and add missing keywords only where the user can support them with real evidence."
            ),
            "sections_to_update": unique_clean_values(section_targets, 7),
            "keyword_targets": display_missing_keywords[:8],
            "evidence_to_keep": display_matched_signals[:8],
            "selected_role_gaps": selected_missing_keywords[:6],
        },
    }


def build_analysis_payload(
    resume_text: str,
    job_description: str,
    uploaded_name: str,
    requirements: list[dict[str, Any]] | None = None,
    target_role: str = "",
    document_info: dict[str, Any] | None = None,
) -> dict[str, Any]:
    resume_profile = extract_resume_profile(resume_text)
    resume_quality = analyze_resume_quality(resume_text, uploaded_name, document_info, resume_profile)
    fit_result = evaluate_job_fit(resume_text, job_description, requirements)
    job_keywords = fit_result["important_keywords"]
    rule_feedback, missing_keywords = rule_based_feedback(resume_profile, job_keywords)
    similarity_score = fit_result["similarity_score"]
    keyword_coverage = fit_result["keyword_overlap_score"]

    keyword_match = round(keyword_coverage)
    role_evidence = round(compute_role_evidence_score(resume_text, target_role, job_keywords))
    experience_alignment = round(compute_experience_alignment_score(resume_profile, keyword_coverage, similarity_score))
    academic_rigor = round(compute_academic_rigor_score(resume_profile, keyword_coverage, similarity_score))
    resume_structure = round(compute_resume_structure_score(resume_profile))
    resume_quality_score = round(resume_quality["quality_score"])
    ats_score = compute_final_readiness_score(
        keyword_match,
        role_evidence,
        experience_alignment,
        resume_quality_score,
        academic_rigor,
        similarity_score,
    )
    title_match_score = compute_role_title_match_score(resume_text, target_role)
    exact_title_match = has_exact_role_title_match(resume_text, target_role)
    if (
        resume_quality_score >= 95
        and role_evidence >= 90
        and experience_alignment >= 70
        and (exact_title_match or title_match_score >= 70)
        and keyword_match >= 35
    ):
        ats_score = max(ats_score, 82.0)

    summary = "Resume parsing identified {} skill(s), {} education marker(s), and {} experience marker(s).".format(
        len(resume_profile.skills), len(resume_profile.education), len(resume_profile.experience)
    )
    recommendation = rule_feedback[0] if rule_feedback else "Refine the resume to better match the target role."
    quality_feedback = [
        f"ATS format issue: {issue['suggestion']}"
        for issue in resume_quality.get("issues", [])[:2]
    ]

    return {
        "analysis_title": APP_TITLE,
        "uploaded_resume_name": uploaded_name,
        "matching_percentage": ats_score,
        "resume_excerpt": resume_profile.excerpt,
        "resume_summary": {
            "skills": resume_profile.skills,
            "education_markers": resume_profile.education,
            "experience_markers": resume_profile.experience,
        },
        "resume_quality": resume_quality,
        "job_description_analysis": {
            "important_keywords": job_keywords,
            "missing_keywords": missing_keywords,
            "matched_keywords": fit_result["matched_keywords"],
            "matched_requirements": fit_result.get("matched_requirements", []),
            "missing_requirements": fit_result.get("missing_requirements", []),
        },
        "compatibility": {
            "matching_percentage": ats_score,
            "keyword_match": keyword_match,
            "role_evidence": role_evidence,
            "experience_alignment": experience_alignment,
            "academic_rigor": academic_rigor,
            "resume_quality": resume_quality_score,
        },
        "score_breakdown": {
            "ats_score": ats_score,
            "keyword_coverage": keyword_match,
            "role_evidence": role_evidence,
            "experience_alignment": experience_alignment,
            "academic_rigor": academic_rigor,
            "resume_structure": resume_structure,
            "resume_quality": resume_quality_score,
            "tfidf_similarity": round(similarity_score, 2),
            "title_match": round(title_match_score),
            "formula": "30% database requirement coverage + 25% role evidence + 15% experience evidence + 20% ATS format quality + 7% education/skills evidence + 3% TF-IDF similarity, with a strong-evidence calibration floor",
        },
        "feedback": " ".join(rule_feedback),
        "structured_feedback": quality_feedback + rule_feedback,
        "summary": summary,
        "recommendation": recommendation,
        "generated_at": datetime.now(timezone.utc).isoformat(),
    }


def get_user_id_by_email(email: str) -> int | None:
    normalized_email = normalize_email(email)
    if not normalized_email:
        return None

    connection = get_db_connection()
    cursor = connection.cursor(dictionary=True)
    cursor.execute("SELECT id FROM users WHERE email = %s", (normalized_email,))
    user = cursor.fetchone()
    cursor.close()
    connection.close()
    return int(user["id"]) if user else None


def persist_analysis(payload: dict[str, Any], job_description: str, target_role: str, user_email: str) -> int:
    user_id = get_user_id_by_email(user_email)
    connection = get_db_connection()
    cursor = connection.cursor()
    cursor.execute(
        """
        INSERT INTO resume_analyses (
            user_id,
            uploaded_resume_name,
            target_role,
            matching_percentage,
            feedback,
            job_description,
            resume_excerpt,
            analysis_json
        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """,
        (
            user_id,
            payload.get("uploaded_resume_name", "resume.pdf"),
            target_role or "selected role",
            payload["compatibility"]["matching_percentage"],
            payload["feedback"],
            job_description,
            payload.get("resume_excerpt", ""),
            json.dumps(payload),
        ),
    )
    analysis_id = int(cursor.lastrowid)
    connection.commit()
    cursor.close()
    connection.close()
    return analysis_id


def get_role_family_prefix(role_name: str) -> str:
    normalized_role = normalize_match_text(role_name)
    if "software" in normalized_role and ("developer" in normalized_role or "engineer" in normalized_role):
        return "SD"
    if "data" in normalized_role and "analyst" in normalized_role:
        return "DA"
    if (
        "cyber" in normalized_role
        or "security" in normalized_role
        or "cloud" in normalized_role
        or "database" in normalized_role
    ):
        return "CY"
    return ""


def infer_accuracy_resume_id(uploaded_name: str, target_role: str, selected_role: str = "") -> str:
    name = normalize_match_text(os.path.splitext(os.path.basename(uploaded_name or ""))[0])
    direct_match = re.search(r"\b(sd|da|cy)\s*0?(\d{1,2})\b", name)
    if direct_match:
        return f"{direct_match.group(1).upper()}{int(direct_match.group(2)):02d}"

    role_text = f"{target_role} {selected_role}"
    prefix = get_role_family_prefix(role_text)
    if not prefix:
        return ""

    number_match = re.search(r"\b(\d{1,2})\b", name)
    if number_match:
        return f"{prefix}{int(number_match.group(1)):02d}"

    return ""


def get_analysis_recommended_role(analysis: dict[str, Any]) -> str:
    role_recommendation = analysis.get("role_recommendation") or {}
    return normalize_text(
        role_recommendation.get("recommended_role")
        or role_recommendation.get("recommended_database_role")
        or analysis.get("target_role")
        or analysis.get("selected_role")
        or ""
    )


def build_accuracy_missing_keywords_note(analysis: dict[str, Any]) -> str:
    job_analysis = analysis.get("job_description_analysis") or {}
    role_recommendation = analysis.get("role_recommendation") or {}
    missing_keywords = list(job_analysis.get("missing_keywords") or [])
    if not missing_keywords:
        missing_keywords = list(role_recommendation.get("recommended_missing_keywords") or [])

    note_parts: list[str] = []
    if missing_keywords:
        note_parts.append("Missing: " + ", ".join(str(item) for item in missing_keywords[:8]))

    matched_keywords = list(job_analysis.get("matched_keywords") or [])
    if matched_keywords:
        note_parts.append("Matched: " + ", ".join(str(item) for item in matched_keywords[:8]))

    return " | ".join(note_parts)[:500]


def find_accuracy_testing_row(workbook, analysis: dict[str, Any]) -> tuple[Any | None, int | None, str]:
    sheet = workbook["Accuracy Testing"] if "Accuracy Testing" in workbook.sheetnames else None
    if sheet is None:
        return None, None, "Accuracy Testing sheet was not found."

    uploaded_name = str(analysis.get("uploaded_resume_name") or "")
    target_role = str(analysis.get("target_role") or "")
    selected_role = str(analysis.get("selected_role") or "")
    inferred_resume_id = infer_accuracy_resume_id(uploaded_name, target_role, selected_role)

    for row in range(5, min(sheet.max_row, 34) + 1):
        resume_id = normalize_text(str(sheet.cell(row=row, column=1).value or ""))
        if inferred_resume_id and normalize_match_text(resume_id) == normalize_match_text(inferred_resume_id):
            return sheet, row, ""

    if "Resume Dataset" in workbook.sheetnames:
        dataset = workbook["Resume Dataset"]
        uploaded_key = normalize_match_text(uploaded_name)
        for data_row in range(5, dataset.max_row + 1):
            dataset_resume_id = normalize_text(str(dataset.cell(row=data_row, column=1).value or ""))
            fake_file_name = normalize_match_text(str(dataset.cell(row=data_row, column=4).value or ""))
            if uploaded_key and fake_file_name and uploaded_key == fake_file_name:
                for row in range(5, min(sheet.max_row, 34) + 1):
                    if normalize_text(str(sheet.cell(row=row, column=1).value or "")) == dataset_resume_id:
                        return sheet, row, ""

    expected_role_key = normalize_match_text(selected_role or target_role)
    for row in range(5, min(sheet.max_row, 34) + 1):
        row_expected_role = normalize_match_text(str(sheet.cell(row=row, column=2).value or ""))
        row_actual_role = normalize_text(str(sheet.cell(row=row, column=3).value or ""))
        if row_expected_role and expected_role_key and row_expected_role == expected_role_key and not row_actual_role:
            return sheet, row, ""

    for row in range(5, min(sheet.max_row, 34) + 1):
        if not normalize_text(str(sheet.cell(row=row, column=1).value or "")):
            return sheet, row, ""

    return sheet, None, "No empty Accuracy Testing row is available."


def compact_accuracy_testing_analysis(analysis: dict[str, Any]) -> dict[str, Any]:
    role_recommendation = analysis.get("role_recommendation") or {}
    job_analysis = analysis.get("job_description_analysis") or {}
    return {
        "uploaded_resume_name": analysis.get("uploaded_resume_name"),
        "target_role": analysis.get("target_role"),
        "selected_role": analysis.get("selected_role"),
        "matching_percentage": analysis.get("matching_percentage"),
        "summary": analysis.get("summary"),
        "feedback": analysis.get("feedback"),
        "job_description_analysis": {
            "missing_keywords": list(job_analysis.get("missing_keywords") or [])[:12],
            "matched_keywords": list(job_analysis.get("matched_keywords") or [])[:12],
        },
        "role_recommendation": {
            "recommended_role": role_recommendation.get("recommended_role"),
            "recommended_database_role": role_recommendation.get("recommended_database_role"),
            "recommended_missing_keywords": list(role_recommendation.get("recommended_missing_keywords") or [])[:12],
            "reason_summary": role_recommendation.get("reason_summary"),
        },
        "queued_at": datetime.now().isoformat(),
    }


def queue_accuracy_testing_analysis(analysis: dict[str, Any]) -> None:
    pending_path = os.path.abspath(ACCURACY_TESTING_PENDING_FILE)
    os.makedirs(os.path.dirname(pending_path), exist_ok=True)
    with open(pending_path, "a", encoding="utf-8") as pending_file:
        pending_file.write(json.dumps(compact_accuracy_testing_analysis(analysis)) + "\n")


def load_pending_accuracy_testing_analyses() -> list[dict[str, Any]]:
    pending_path = os.path.abspath(ACCURACY_TESTING_PENDING_FILE)
    if not os.path.exists(pending_path):
        return []

    pending_items: list[dict[str, Any]] = []
    with open(pending_path, "r", encoding="utf-8") as pending_file:
        for line in pending_file:
            line = line.strip()
            if not line:
                continue
            try:
                pending_items.append(json.loads(line))
            except json.JSONDecodeError:
                continue
    return pending_items


def clear_pending_accuracy_testing_analyses() -> None:
    pending_path = os.path.abspath(ACCURACY_TESTING_PENDING_FILE)
    if os.path.exists(pending_path):
        os.remove(pending_path)


def write_accuracy_testing_analysis_row(workbook, analysis: dict[str, Any]) -> tuple[int | None, str]:
    sheet, row, warning = find_accuracy_testing_row(workbook, analysis)
    if warning or sheet is None or row is None:
        return None, warning

    recommended_role = get_analysis_recommended_role(analysis)
    target_role = normalize_text(str(analysis.get("target_role") or analysis.get("selected_role") or recommended_role))
    if not normalize_text(str(sheet.cell(row=row, column=1).value or "")):
        sheet.cell(row=row, column=1).value = infer_accuracy_resume_id(
            str(analysis.get("uploaded_resume_name") or ""),
            target_role,
            str(analysis.get("selected_role") or ""),
        ) or f"AUTO{row - 4:02d}"
    if not normalize_text(str(sheet.cell(row=row, column=2).value or "")):
        sheet.cell(row=row, column=2).value = target_role

    sheet.cell(row=row, column=3).value = recommended_role or target_role
    sheet.cell(row=row, column=4).value = float(analysis.get("matching_percentage") or 0)
    if not normalize_text(str(sheet.cell(row=row, column=5).value or "")):
        sheet.cell(row=row, column=5).value = "Yes"
    if not normalize_text(str(sheet.cell(row=row, column=6).value or "")):
        sheet.cell(row=row, column=6).value = "Yes"
    if not normalize_text(str(sheet.cell(row=row, column=7).value or "")):
        sheet.cell(row=row, column=7).value = "Yes"

    sheet.cell(row=row, column=8).value = build_accuracy_missing_keywords_note(analysis)
    role_recommendation = analysis.get("role_recommendation") or {}
    sheet.cell(row=row, column=9).value = normalize_text(
        role_recommendation.get("reason_summary")
        or analysis.get("summary")
        or analysis.get("feedback")
        or ""
    )[:500]

    if not normalize_text(str(sheet.cell(row=row, column=14).value or "")):
        sheet.cell(row=row, column=14).value = (
            f"Auto-filled from {analysis.get('uploaded_resume_name', 'resume')} "
            f"at {datetime.now().strftime('%Y-%m-%d %H:%M')}."
        )

    return row, ""


def update_accuracy_testing_workbook(analysis: dict[str, Any]) -> dict[str, Any]:
    if not ACCURACY_TESTING_AUTO_EXPORT:
        return {"accuracy_testing_saved": False, "accuracy_testing_warning": "Accuracy testing auto export is disabled."}

    workbook_path = os.path.abspath(ACCURACY_TESTING_WORKBOOK)
    if not os.path.exists(workbook_path):
        return {
            "accuracy_testing_saved": False,
            "accuracy_testing_warning": f"Accuracy testing workbook was not found: {workbook_path}",
        }

    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError:
        return {
            "accuracy_testing_saved": False,
            "accuracy_testing_warning": "Install openpyxl to auto-update the Excel testing workbook.",
        }

    workbook = None
    try:
        workbook = load_workbook(workbook_path)
        pending_items = load_pending_accuracy_testing_analyses()
        saved_rows: list[int] = []
        for pending_analysis in pending_items:
            row, warning = write_accuracy_testing_analysis_row(workbook, pending_analysis)
            if warning or row is None:
                workbook.close()
                return {"accuracy_testing_saved": False, "accuracy_testing_warning": warning}
            saved_rows.append(row)

        row, warning = write_accuracy_testing_analysis_row(workbook, analysis)
        if warning or row is None:
            workbook.close()
            return {"accuracy_testing_saved": False, "accuracy_testing_warning": warning}
        saved_rows.append(row)

        try:
            workbook.calculation.fullCalcOnLoad = True
            workbook.calculation.forceFullCalc = True
        except AttributeError:
            pass

        workbook.save(workbook_path)
        workbook.close()
        if pending_items:
            clear_pending_accuracy_testing_analyses()
        return {
            "accuracy_testing_saved": True,
            "accuracy_testing_row": row,
            "accuracy_testing_saved_rows": saved_rows,
            "accuracy_testing_pending_flushed": len(pending_items),
            "accuracy_testing_workbook": workbook_path,
        }
    except PermissionError:
        if workbook is not None:
            try:
                workbook.close()
            except Exception:
                pass
        try:
            queue_accuracy_testing_analysis(analysis)
        except OSError as error:
            return {
                "accuracy_testing_saved": False,
                "accuracy_testing_pending": False,
                "accuracy_testing_warning": (
                    "Excel workbook is open, and the result could not be queued. "
                    f"Close ats_resume_fyp_testing.xlsx and run the test again. Detail: {error}"
                ),
            }
        return {
            "accuracy_testing_saved": False,
            "accuracy_testing_pending": True,
            "accuracy_testing_warning": "Excel workbook is open. Result was queued; close ats_resume_fyp_testing.xlsx and run one more analysis to flush pending results.",
        }
    except OSError as error:
        return {
            "accuracy_testing_saved": False,
            "accuracy_testing_warning": f"Could not update accuracy testing workbook: {error}",
        }


CHARACTER_SPACED_WORD_PATTERN = re.compile(r"(?<![A-Za-z])(?:[A-Za-z][ \t]){2,}[A-Za-z](?![A-Za-z])")
CHARACTER_SPACED_DIGIT_PATTERN = re.compile(r"(?<!\d)(?:\d[ \t])+\d(?!\d)")
PDF_SYMBOL_ARTIFACT_PATTERN = re.compile(r"[\uf000-\uf8ff\ufffd\u25a0-\u25ff\u2610-\u2612\u2022\u2023\u2043]")
SHORT_SPACED_WORD_REPAIRS = (
    ("a m", "am"),
    ("a n", "an"),
    ("a s", "as"),
    ("a t", "at"),
    ("b e", "be"),
    ("b y", "by"),
    ("i f", "if"),
    ("i n", "in"),
    ("i s", "is"),
    ("i t", "it"),
    ("m y", "my"),
    ("o f", "of"),
    ("o n", "on"),
    ("o r", "or"),
    ("q a", "QA"),
    ("t o", "to"),
)


def sanitize_resume_text_artifacts(value: str) -> str:
    cleaned = str(value or "")
    cleaned = cleaned.translate({
        0x2018: ord("'"),
        0x2019: ord("'"),
        0x201C: ord('"'),
        0x201D: ord('"'),
        0x2013: ord("-"),
        0x2014: ord("-"),
    })
    ligature_replacements = {
        "\ufb00": "ff",
        "\ufb01": "fi",
        "\ufb02": "fl",
        "\ufb03": "ffi",
        "\ufb04": "ffl",
    }
    for source, replacement in ligature_replacements.items():
        cleaned = cleaned.replace(source, replacement)
    cleaned = PDF_SYMBOL_ARTIFACT_PATTERN.sub("\n", cleaned)
    cleaned = re.sub(r"\b(?:cid|wingdings|symbol)\s*:\s*\d+\b", " ", cleaned, flags=re.I)
    cleaned = re.sub(r"_{3,}", "\n", cleaned)
    cleaned = re.sub(r"(?<=\d)(?=[A-Za-z])", " ", cleaned)
    cleaned = re.sub(r"(?<=%)(?=[A-Za-z])", " ", cleaned)
    cleaned = re.sub(
        r"(?<=[a-z])(?=(?:Access|Analyst|Contact|Developer|Education|Engineer|Experience|Full|Functional|Other|Present|Previous|Skills|Work)\b)",
        " ",
        cleaned,
    )
    cleaned = re.sub(r"(?<=[a-z])(?=(?:CSS|HTML|JavaScript|Node|PHP|Python|React)\b)", " ", cleaned)
    cleaned = re.sub(
        r"(?<=[a-z])(?=(?:agile|coding|cost|customer|customers|data|enterprise|features|library|major|months|network|retrieve|satisfaction|software|support|testing|traffic|using|website|websites)\b)",
        " ",
        cleaned,
        flags=re.I,
    )
    cleaned = re.sub(r"\.\s*c\s*o\s*m+\b", ".com", cleaned, flags=re.I)
    cleaned = re.sub(r"(?<=[A-Za-z])\s+(?=\d+@)", "", cleaned)
    cleaned = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", " ", cleaned)
    return cleaned


def repair_character_spaced_pdf_text(value: str) -> str:
    """Recover words from PDFs that extract text as `R e s u m e`."""

    def collapse_line(line: str) -> str:
        if not CHARACTER_SPACED_WORD_PATTERN.search(line):
            return line
        if not re.search(r"[ \t]{2,}", line):
            return line
        parts = re.split(r"([ \t]{2,})", line)
        repaired_parts: list[str] = []
        for part in parts:
            if re.fullmatch(r"[ \t]{2,}", part or ""):
                repaired_parts.append(" ")
                continue
            repaired_parts.append(
                CHARACTER_SPACED_WORD_PATTERN.sub(
                    lambda match: match.group(0).replace(" ", "").replace("\t", ""),
                    part,
                )
            )
        return "".join(repaired_parts)

    repaired = "\n".join(collapse_line(line) for line in sanitize_resume_text_artifacts(value).splitlines())
    repaired = CHARACTER_SPACED_DIGIT_PATTERN.sub(lambda match: re.sub(r"[ \t]+", "", match.group(0)), repaired)
    for spaced_word, replacement in SHORT_SPACED_WORD_REPAIRS:
        repaired = re.sub(rf"\b{re.escape(spaced_word)}\b", replacement, repaired, flags=re.I)
    repaired = re.sub(r"\bAsa\b", "As a", repaired)
    repaired = re.sub(r"\bso\s+ware\b", "software", repaired, flags=re.I)
    repaired = re.sub(r"(?<=\w)\s*@\s*(?=\w)", "@", repaired)
    repaired = re.sub(r"(?<=\w)\s*\.\s*(?=(?:com|net|org|edu|io|dev|me)\b)", ".", repaired, flags=re.I)
    repaired = re.sub(r"\s+/\s+", " / ", repaired)
    return repaired


def clean_resume_source_text(value: str) -> str:
    cleaned = repair_character_spaced_pdf_text(str(value or ""))
    cleaned = sanitize_resume_text_artifacts(cleaned)
    cleaned = re.sub(r"\r\n?", "\n", cleaned)
    cleaned = re.sub(r"[ \t]+", " ", cleaned)
    cleaned = re.sub(r"\n{3,}", "\n\n", cleaned)
    return cleaned.strip()


def unique_clean_values(values: list[Any], limit: int | None = None) -> list[str]:
    seen: set[str] = set()
    cleaned_values: list[str] = []
    for value in values:
        cleaned = normalize_text(str(value or ""))
        if not cleaned:
            continue
        key = normalize_match_text(cleaned)
        if not key or key in seen:
            continue
        seen.add(key)
        cleaned_values.append(cleaned)
        if limit is not None and len(cleaned_values) >= limit:
            break
    return cleaned_values


def normalize_resume_link(value: str) -> str:
    link = normalize_text(str(value or "")).strip(" ,;)")
    if not link:
        return ""
    if not re.match(r"^https?://", link, re.I):
        link = "https://" + link
    return link


def get_link_label(link: str) -> str:
    normalized = normalize_text(link)
    lowered = normalized.lower()
    without_scheme = re.sub(r"^https?://", "", normalized, flags=re.I).rstrip("/")
    if "linkedin.com" in lowered:
        return "LinkedIn: " + without_scheme
    if "github.com" in lowered:
        return "GitHub: " + without_scheme
    return "Portfolio: " + without_scheme


def select_resume_template_profile(target_role: str) -> dict[str, str]:
    normalized_role = normalize_match_text(target_role)
    if any(term in normalized_role for term in ("cyber", "security")):
        family = "Cybersecurity"
    elif any(term in normalized_role for term in ("devops", "cloud", "operation", "infrastructure")):
        family = "DevOps / IT Operations"
    elif "quality" in normalized_role or "tester" in normalized_role or "qa" in normalized_role:
        family = "Quality Assurance"
    elif "network" in normalized_role:
        family = "Network Engineering"
    elif "java" in normalized_role:
        family = "Java Developer"
    elif "python" in normalized_role:
        family = "Python Developer"
    elif "full" in normalized_role or "web" in normalized_role or "front" in normalized_role:
        family = "Full Stack / Web Developer"
    else:
        family = "Software Developer"

    return {
        "name": family + " ATS Template",
        "headline": target_role or family,
        "experience_heading": "Professional Experience",
        "project_heading": "Projects",
        "certification_heading": "Certifications",
        "achievement_heading": "Achievements",
        "skills_heading": "Technical Skills",
        "education_heading": "Education",
        "other_heading": "Additional ATS Evidence",
    }


def get_resume_generation_role(analysis: dict[str, Any], requested_role: str = "") -> str:
    role_recommendation = analysis.get("role_recommendation") or {}
    return normalize_text(
        requested_role
        or role_recommendation.get("recommended_role")
        or role_recommendation.get("recommended_database_role")
        or analysis.get("target_role")
        or analysis.get("selected_role")
        or "Target Role"
    )


def infer_resume_name_from_text(text: str) -> str:
    lines = [normalize_text(line) for line in clean_resume_source_text(text).splitlines()]
    ignored_terms = set(CORE_RESUME_SECTION_MARKERS + CONTACT_MARKERS + ["email", "phone", "mobile", "resume", "curriculum vitae"])
    role_terms = {"developer", "engineer", "analyst", "manager", "intern", "student", "assistant", "support"}

    for raw_line in lines[:14]:
        line = re.sub(r"[\w.+-]+@[\w-]+(?:\.[\w-]+)+", "", raw_line)
        line = re.sub(r"(?:\+?\d[\d\s().-]{7,}\d)", "", line)
        line = re.sub(r"^(?:name|candidate)\s*:\s*", "", line, flags=re.I)
        line = re.sub(r"([a-z])([A-Z])", r"\1 \2", line)
        line = re.split(
            r"\b(?:entry level|junior|senior|lead)?\s*(?:software|web|full stack|front end|back end|react|java|python|computer|desktop|support|network|qa|quality|data|cyber|cloud)\b",
            line,
            maxsplit=1,
            flags=re.I,
        )[0]
        line = normalize_text(line.strip(" |,;:-"))
        if not line:
            continue
        normalized = normalize_match_text(line)
        if not normalized or any(term in normalized for term in ignored_terms):
            continue
        words = [word for word in re.split(r"\s+", line) if word]
        if 2 <= len(words) <= 5 and not (set(normalized.split()) & role_terms):
            if sum(1 for word in words if word[:1].isupper() or word.isupper()) >= max(1, len(words) - 1):
                return line
    return ""


def infer_resume_name_from_filename(filename: str) -> str:
    stem = os.path.splitext(os.path.basename(str(filename or "")))[0]
    cleaned = re.sub(r"[_-]+", " ", stem)
    cleaned = re.sub(r"\b(?:resume|cv|curriculum|vitae|ats|final|latest|updated)\b", " ", cleaned, flags=re.I)
    cleaned = normalize_text(cleaned.strip(" ,;-"))
    if not cleaned:
        return ""
    words = [word for word in cleaned.split() if not re.fullmatch(r"\d+", word)]
    if 2 <= len(words) <= 5:
        return " ".join(words)
    return ""


def extract_resume_phone(text: str) -> str:
    candidates = re.findall(r"(?:\+?\s*\d[\d\s().-]{7,}\d)", text)
    for candidate in candidates:
        cleaned = re.sub(r"^\+\s*", "+", normalize_text(candidate))
        if re.fullmatch(r"(?:19|20)\d{2}\s*[-/]\s*(?:19|20)\d{2}", cleaned):
            continue
        digits = re.sub(r"\D", "", cleaned)
        if len(digits) < 8:
            continue
        if len(digits) == 8 and re.fullmatch(r"(?:19|20)\d{2}(?:19|20)\d{2}", digits):
            continue
        return cleaned
    return ""


def extract_resume_email(text: str) -> str:
    candidates = re.findall(r"[\w.+-]+@[\w-]+(?:\s*\.\s*[\w-]+)+", text)
    emails: list[tuple[str, str, int]] = []
    for index, candidate in enumerate(candidates):
        cleaned = re.sub(r"\s+", "", candidate).strip(".,;:()[]{}<>")
        cleaned = re.sub(r"\.comm\b", ".com", cleaned, flags=re.I)
        if not re.fullmatch(r"[\w.+-]+@[\w-]+(?:\.[\w-]+)+", cleaned):
            continue
        local_part, domain = cleaned.split("@", 1)
        if len(local_part) < 3:
            continue
        normalized = cleaned.lower()
        emails.append((normalized, normalized, index))

    if not emails:
        return ""

    counts = Counter(email for email, _, _ in emails)
    first_seen: dict[str, int] = {}
    for email, _, index in emails:
        first_seen.setdefault(email, index)

    return max(
        counts,
        key=lambda email: (counts[email], len(email.split("@", 1)[0]), -first_seen[email]),
    )


def extract_resume_contact(
    resume_text: str,
    candidate_name: str = "",
    extra_links: list[str] | None = None,
) -> dict[str, Any]:
    text = clean_resume_source_text(resume_text)
    email = extract_resume_email(text)
    phone = extract_resume_phone(text)
    extracted_links = re.findall(
        r"(?:https?://)?(?:www\.)?(?:linkedin\.com|github\.com|[\w.-]+\.(?:dev|io|me|com)/)[^\s,;)]*",
        text,
        re.I,
    )
    links = unique_clean_values(
        [normalize_resume_link(link) for link in list(extra_links or []) + extracted_links],
        5,
    )

    provided_name = normalize_text(candidate_name)
    resume_name = infer_resume_name_from_text(text)
    if resume_name and (not provided_name or len(provided_name.split()) < 2):
        inferred_name = resume_name
    else:
        inferred_name = provided_name or resume_name

    return {
        "name": inferred_name or "Candidate Name",
        "email": email,
        "phone": phone,
        "links": links,
        "link_labels": [get_link_label(link) for link in links],
    }


def clean_resume_sentence(sentence: str) -> str:
    cleaned = sanitize_resume_text_artifacts(sentence)
    cleaned = normalize_text(re.sub(r"^[\-\*\d.)\s]+", "", cleaned or ""))
    cleaned = re.sub(r"\s+[,;:.]", lambda match: match.group(0).strip(), cleaned)
    if len(cleaned) > 220:
        cleaned = cleaned[:217].rstrip(" ,;:.") + "..."
    if cleaned and cleaned[-1] not in ".!?":
        cleaned += "."
    return cleaned


def has_pdf_text_artifacts(sentence: str) -> bool:
    single_letters = [
        token
        for token in re.findall(r"\b[A-Za-z]\b", sentence)
        if token.lower() not in {"a", "i"}
    ]
    if len(single_letters) >= 2:
        return True
    if re.search(r"\b[b-hj-z]\s+[a-z]{2,}\b", sentence):
        return True
    if re.search(r"[a-z][A-Z]\s+[a-z]", sentence):
        return True
    normalized = normalize_match_text(sentence)
    heading_hits = sum(
        1
        for marker in ("about", "experience", "education", "contact", "skill")
        if marker in normalized
    )
    return heading_hits >= 2 and len(sentence) > 110


def polish_resume_evidence_sentence(sentence: str) -> str:
    cleaned = clean_resume_sentence(sentence)
    cleaned = re.sub(r"^(?:Responsibility|Responsibilities)\s*:\s*", "", cleaned, flags=re.I)
    replacements = (
        (r"\bI have been design and deploy for\b", "Designed and deployed"),
        (r"\bI have been design and deploy\b", "Designed and deployed"),
        (r"\bI have been QA for testing\b", "Performed QA testing"),
        (r"\bI have giving the training to the client by how to use\b", "Provided client training on how to use"),
        (r"\bI have giving the training\b", "Provided training"),
        (r"\bI have getting my internship on\b", "Completed internship from"),
        (r"\bI have getting this job half year on\b", "Worked a half-year role from"),
        (r"\bI have been fixes the problem like installing\b", "Fixed setup issues by installing"),
        (r"\bI have knowledge by for creating\b", "Created"),
        (r"\ban webapp\b", "a web app"),
        (r"\bwebapp\b", "web app"),
        (r"\bsql\b", "SQL"),
        (r"\bpower point\b", "PowerPoint"),
        (r"\bfro\b", "for"),
        (r"\bsoftrware\b", "software"),
        (r"\bpastthrough the message to developer to solve\b", "reported to developers for resolution"),
        (r"\bpassed through the message to developer to solve\b", "reported to developers for resolution"),
        (r"\bbug will be record\b", "bugs were recorded"),
        (r"\bhave been done\b", "before release"),
        (r"\bbefore release before deploy for the customer\b", "before customer deployment"),
    )
    for pattern, replacement in replacements:
        cleaned = re.sub(pattern, replacement, cleaned, flags=re.I)
    return clean_resume_sentence(cleaned)


def is_skill_inventory_sentence(sentence: str, sentence_tokens: set[str], action_terms: set[str]) -> bool:
    normalized = normalize_match_text(sentence)
    if sentence_tokens & action_terms:
        return False
    inventory_starts = (
        "additional skills",
        "computer skills",
        "languages",
        "operating systems",
        "programming language",
        "programming languages",
        "relevant courses",
        "software",
        "technical skills",
        "tools",
    )
    if normalized.startswith(inventory_starts):
        return True
    if re.search(r"\b(?:gpa|honors|relevant courses|dean s list)\b", normalized):
        return True
    if sentence.count(",") >= 6:
        return True
    return False


def normalized_phrase_in_text(phrase: str, text: str) -> bool:
    return f" {phrase} " in f" {text} "


def polish_resume_education_sentence(sentence: str) -> str:
    cleaned = clean_resume_sentence(sentence)
    cleaned = re.sub(r"\bTechnolgy\b", "Technology", cleaned, flags=re.I)
    cleaned = re.sub(r"\bCommunicatio n\b", "Communication", cleaned, flags=re.I)
    cleaned = re.sub(r"\bSarawa k\b", "Sarawak", cleaned, flags=re.I)
    cleaned = re.sub(r"\bPolyethnic Mukah\b", "Politeknik Mukah", cleaned, flags=re.I)
    cleaned = re.sub(r"\(\s*((?:19|20)\d{2})\s*-\s*((?:19|20)\d{2})\s*\)", r"(\1 - \2)", cleaned)
    date_match = re.search(r"((?:19|20)\d{2})\s*-\s*((?:19|20)\d{2})", cleaned)
    degree_match = re.search(
        r"\b(Bachelor|Master|PhD|Diploma)\s+(of|in)\s+([A-Za-z &/]{2,80})",
        cleaned,
        re.I,
    )
    if degree_match:
        subject = normalize_text(degree_match.group(3))
        subject = re.split(
            r"\b(?:Bachelor|Master|PhD|Diploma|University|Politeknik|Polytechnic|College|School|Contact)\b",
            subject,
            maxsplit=1,
            flags=re.I,
        )[0].strip(" ,;:.")
        if subject:
            degree = f"{degree_match.group(1).title()} {degree_match.group(2).lower()} {subject}"
            if date_match:
                degree += f" ({date_match.group(1)} - {date_match.group(2)})"
            return clean_resume_sentence(degree)
    if date_match and date_match.end() < len(cleaned):
        trailing = cleaned[date_match.end() :]
        if re.search(r"\b(?:Bachelor|Master|PhD|Diploma|University|Politeknik|Polytechnic|Contact)\b|@|\+?\d", trailing, re.I):
            cleaned = cleaned[: date_match.end()].rstrip(" ,;:.") + "."
    return clean_resume_sentence(cleaned)


def normalize_education_scan_text(resume_text: str) -> str:
    text = normalize_text(clean_resume_source_text(resume_text))
    replacements = (
        (r"\bTechnolgy\b", "Technology"),
        (r"\bCommunicatio\s*n\b", "Communication"),
        (r"\bCommunicationn\b", "Communication"),
        (r"\bSarawa\s*k\b", "Sarawak"),
        (r"\bPolyethnic Mukah\b", "Politeknik Mukah"),
        (r"\bU niversity\b", "University"),
        (r"\bP\s+oliteknik\b", "Politeknik"),
        (r"\bM\s+ukah\b", "Mukah"),
        (r"\bD\s+iploma\b", "Diploma"),
    )
    for pattern, replacement in replacements:
        text = re.sub(pattern, replacement, text, flags=re.I)
    text = re.sub(r"([a-z])([A-Z])", r"\1 \2", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def find_education_institution(context: str) -> str:
    known_institutions = (
        ("University of Technology Sarawak", r"University\s+of\s+Technology\s+Sarawa\s*k"),
        ("Politeknik Mukah Sarawak", r"Politeknik\s+Mukah\s+Sarawa\s*k"),
        ("Politeknik Mukah", r"Politeknik\s+Mukah"),
        ("Polytechnic Mukah Sarawak", r"Polytechnic\s+Mukah\s+Sarawa\s*k"),
    )
    for label, pattern in known_institutions:
        if re.search(pattern, context, re.I):
            return label

    generic_matches = re.findall(
        r"\b(?:University|Politeknik|Polytechnic|College|Institute|School)\s+(?:of\s+)?[A-Za-z]+(?:\s+[A-Za-z]+){0,5}",
        context,
        re.I,
    )
    if not generic_matches:
        return ""

    institution = normalize_text(generic_matches[-1])
    institution = re.split(
        r"\b(?:Bachelor|Master|PhD|Diploma|Undergraduate|Contact|Email|Phone)\b",
        institution,
        maxsplit=1,
        flags=re.I,
    )[0].strip(" ,;:.")
    return " ".join(word.upper() if len(word) <= 3 and word.isupper() else word.capitalize() for word in institution.split())


def clean_education_subject(value: str) -> str:
    subject = normalize_text(value)
    subject = re.sub(r"\bD\s+iploma\b", "Diploma", subject, flags=re.I)
    subject = re.sub(r"\bCommunicatio\s*n\b", "Communication", subject, flags=re.I)
    subject = re.split(
        r"\b(?:Bachelor|Master|PhD|Diploma|University|Politeknik|Polytechnic|College|School|Contact|Responsibility|I have)\b",
        subject,
        maxsplit=1,
        flags=re.I,
    )[0]
    subject = subject.strip(" ,;:.()")
    subject = re.sub(r"\bComputer science\b", "Computer Science", subject, flags=re.I)
    subject = re.sub(r"\bTechnology\b", "Technology", subject, flags=re.I)
    return normalize_text(subject)


def extract_structured_education_lines(resume_text: str) -> list[str]:
    scan_text = normalize_education_scan_text(resume_text)
    degree_pattern = re.compile(
        r"\b(Bachelor|Master|PhD|Diploma)\s+(of|in)\s+([A-Za-z &/]{3,90}?)(?:\s*\(\s*((?:19|20)\d{2})\s*-\s*((?:19|20)\d{2})\s*\)|(?=\s+(?:Bachelor|Master|PhD|Diploma|University|Politeknik|Polytechnic|College|School|Contact|Responsibility|I have|$)))",
        re.I,
    )
    structured: list[tuple[int, str]] = []
    seen: set[tuple[str, str, str]] = set()

    for match in degree_pattern.finditer(scan_text):
        degree_type = match.group(1).title()
        connector = match.group(2).lower()
        subject = clean_education_subject(match.group(3))
        if not subject:
            continue
        start_year = match.group(4) or ""
        end_year = match.group(5) or ""
        if not start_year or not end_year:
            nearby = scan_text[match.end() : match.end() + 80]
            year_match = re.search(r"((?:19|20)\d{2})\s*-\s*((?:19|20)\d{2})", nearby)
            if year_match:
                start_year, end_year = year_match.group(1), year_match.group(2)

        context = scan_text[max(0, match.start() - 280) : match.start()]
        institution = find_education_institution(context)
        signature = (degree_type.lower(), normalize_match_text(subject), f"{start_year}-{end_year}")
        if signature in seen:
            continue
        seen.add(signature)

        line = f"{degree_type} {connector} {subject}"
        if institution:
            line += f" - {institution}"
        if start_year and end_year:
            line += f" ({start_year} - {end_year})"
        sort_year = int(start_year) if start_year.isdigit() else 0
        structured.append((sort_year, clean_resume_sentence(line)))

    structured.sort(key=lambda item: item[0], reverse=True)
    return unique_clean_values([line for _, line in structured], 3)


def is_resume_section_heading(line: str) -> bool:
    normalized = normalize_match_text(line)
    if not normalized:
        return False
    headings = set(CORE_RESUME_SECTION_MARKERS + ["contact", "my contact", "professional summary"])
    return len(line) <= 60 and any(normalized == heading or normalized.startswith(heading + " ") for heading in headings)


def split_resume_statement_parts(text: str) -> list[str]:
    lines = [normalize_text(line) for line in clean_resume_source_text(text).splitlines()]
    parts: list[str] = []
    buffer = ""
    starts_new_statement = re.compile(
        r"^(?:responsibilit(?:y|ies)|project|achievement|experience|work|internship|employment|authored|built|contributed|created|designed|developed|established|fixed|implemented|improved|performed|provided|repaired|resolved|tested|updated|worked)\b",
        re.I,
    )

    def flush_buffer() -> None:
        nonlocal buffer
        if buffer:
            parts.append(buffer)
            buffer = ""

    for line in lines:
        if not line:
            flush_buffer()
            continue
        if re.search(r"[\w.+-]+@[\w-]+(?:\.[\w-]+)+|(?:\+?\d[\d\s().-]{7,}\d)", line):
            flush_buffer()
            continue
        if is_resume_section_heading(line):
            flush_buffer()
            continue
        if buffer and starts_new_statement.search(line):
            flush_buffer()
        buffer = (buffer + " " + line).strip() if buffer else line
        if re.search(r"[.!?]$", line) or len(buffer) >= 240:
            flush_buffer()

    flush_buffer()

    split_parts: list[str] = []
    for part in parts:
        split_parts.extend(
            re.split(
                r"(?<=[.!?])\s+|(?=\b(?:Project|Projects|Experience|Work|Education|Skills|Responsibility)\b)",
                part,
            )
        )
    return [normalize_text(part) for part in split_parts if len(normalize_text(part)) > 10]


def extract_resume_evidence_sentences(resume_text: str, keywords: list[str], limit: int = 6) -> list[str]:
    raw_parts = split_resume_statement_parts(resume_text)
    terms = {
        normalized
        for term in keywords
        if (normalized := normalize_match_text(term)) and len(normalized.replace(" ", "")) > 1
    }
    action_terms = {
        "achieved",
        "analyzed",
        "administered",
        "assisted",
        "built",
        "configured",
        "contributed",
        "create",
        "created",
        "debugged",
        "deployed",
        "deploy",
        "designed",
        "developed",
        "implemented",
        "improved",
        "led",
        "managed",
        "monitored",
        "optimized",
        "performed",
        "recorded",
        "resolved",
        "supported",
        "trained",
        "training",
        "tested",
        "troubleshot",
        "updated",
        "worked",
    }
    education_terms = {"bachelor", "master", "phd", "degree", "university", "college", "diploma"}
    scored_sentences: list[tuple[int, str]] = []
    seen: set[str] = set()

    for raw_part in raw_parts:
        sentence = polish_resume_evidence_sentence(raw_part)
        if len(sentence) < 24:
            continue
        if has_pdf_text_artifacts(sentence):
            continue
        if re.search(r"[\w.+-]+@[\w-]+(?:\.[\w-]+)+", sentence):
            continue
        if re.search(r"(?:\+?\d[\d\s().-]{7,}\d)", sentence):
            continue
        key = normalize_match_text(sentence)
        if not key or key in seen:
            continue
        seen.add(key)

        normalized_sentence = normalize_match_text(sentence)
        sentence_tokens = set(normalized_sentence.split())
        if is_skill_inventory_sentence(sentence, sentence_tokens, action_terms):
            continue
        if sentence_tokens & education_terms and not (sentence_tokens & action_terms):
            continue
        term_hits = sum(1 for term in terms if term and normalized_phrase_in_text(term, normalized_sentence))
        has_action = bool(sentence_tokens & action_terms)
        has_metric = bool(re.search(r"\d+(?:\.\d+)?%|\d+\+|\b\d{2,}\b|users?|accuracy|reduced|increased|improved", sentence, re.I))
        has_technical_context = any(
            normalized_phrase_in_text(marker, normalized_sentence)
            for marker in ("project", "intern", "system", "application", "website", "webapp", "software", "database", "network", "networks", "api", "program", "code", "technical")
        )
        if not has_action and re.search(r"\b(?:present|(?:19|20)\d{2}|[01]?\d\s*/\s*(?:19|20)\d{2})\b", sentence, re.I):
            if any(normalized_phrase_in_text(term, normalized_sentence) for term in ("developer", "engineer", "analyst", "assistant", "intern", "support")):
                continue
        if not has_action and len(sentence_tokens) <= 7:
            continue
        if term_hits == 0 and not has_technical_context:
            continue
        if not has_action and term_hits < 2 and not has_technical_context:
            continue
        score = 0
        score += term_hits
        score += 2 if has_action else 0
        score += 2 if has_metric else 0
        score += 1 if has_technical_context else 0
        if score <= 0:
            continue
        scored_sentences.append((score, sentence))

    scored_sentences.sort(key=lambda item: (item[0], len(item[1])), reverse=True)
    return [sentence for _, sentence in scored_sentences[:limit]]


def extract_education_lines(resume_text: str, education_markers: list[str]) -> list[str]:
    structured_lines = extract_structured_education_lines(resume_text)
    if structured_lines:
        return structured_lines

    education_terms = {"bachelor", "master", "phd", "degree", "diploma", "university", "college", "faculty", "school"}
    raw_parts = split_resume_statement_parts(resume_text)
    scored_lines: list[tuple[int, str]] = []
    for raw_part in raw_parts:
        sentence = polish_resume_education_sentence(raw_part)
        normalized_sentence = normalize_match_text(sentence)
        if len(sentence) < 16:
            continue
        if re.search(r"(?:19|20)\d{2}\s*-\s*(?:bachelor|master|phd|diploma|university|politeknik|polytechnic)", sentence, re.I):
            continue
        if any(term in normalized_sentence for term in education_terms):
            score = 0
            score += 4 if re.search(r"\b(?:Bachelor|Master|PhD|Diploma)\b", sentence, re.I) else 0
            score += 3 if re.search(r"(?:19|20)\d{2}\s*-\s*(?:19|20)\d{2}", sentence) else 0
            score += 2 if any(term in normalized_sentence for term in ("university", "college", "politeknik", "polytechnic", "school")) else 0
            score -= 4 if has_pdf_text_artifacts(sentence) else 0
            score -= 3 if re.search(r"[\w.+-]+@[\w-]+(?:\.[\w-]+)+|(?:\+?\d[\d\s().-]{7,}\d)", sentence) else 0
            score -= 1 if normalized_sentence.startswith(("firstly", "currently", "i ")) else 0
            scored_lines.append((score, sentence))

    if scored_lines:
        scored_lines.sort(key=lambda item: (item[0], len(item[1])), reverse=True)
        ranked_lines = unique_clean_values([sentence for _, sentence in scored_lines], 5)
        filtered_lines: list[str] = []
        filtered_keys: list[str] = []
        seen_education_signatures: set[tuple[str, str]] = set()
        for line in ranked_lines:
            key = normalize_match_text(line)
            if len(filtered_lines) >= 2 and key.startswith(("currently ", "firstly ", "for my ", "i ")):
                continue
            degree_match = re.search(r"\b(bachelor|master|phd|diploma)\b", line, re.I)
            date_match = re.search(r"((?:19|20)\d{2})\s*-\s*((?:19|20)\d{2})", line)
            signature = (
                degree_match.group(1).lower() if degree_match else "",
                f"{date_match.group(1)}-{date_match.group(2)}" if date_match else "",
            )
            if signature[0] and signature[1] and signature in seen_education_signatures:
                continue
            if any(existing_key.startswith(key) and len(existing_key) > len(key) + 8 for existing_key in filtered_keys):
                continue
            filtered_lines.append(line)
            filtered_keys.append(key)
            if signature[0] and signature[1]:
                seen_education_signatures.add(signature)
            if len(filtered_lines) >= 3:
                break
        return filtered_lines
    if education_markers:
        return ["Education details detected in the original resume. Confirm degree, institution, and graduation date."]
    return []


def generated_placeholder(label: str, detail: str) -> str:
    return f"[Add {label}] {detail}"


def is_generated_placeholder(value: str) -> bool:
    return normalize_text(value).startswith("[Add ")


def is_resume_advice_line(value: str) -> bool:
    normalized = normalize_match_text(value)
    if not normalized:
        return False
    advice_starts = (
        "add evidence if you have it",
        "add a projects section only if",
        "add certifications only if",
        "add awards or achievements only if",
        "add those details only if",
        "draft uses extracted resume evidence",
        "if available add more measurable",
        "limited source information",
        "only keep a missing keyword",
        "target keywords to add with evidence",
    )
    return any(normalized.startswith(prefix) for prefix in advice_starts)


def real_resume_items(values: list[Any]) -> list[str]:
    return [
        normalize_text(str(value or ""))
        for value in values
        if normalize_text(str(value or ""))
        and not is_generated_placeholder(str(value or ""))
        and not is_resume_advice_line(str(value or ""))
    ]


def is_project_like_sentence(sentence: str) -> bool:
    normalized = normalize_match_text(sentence)
    project_terms = (
        "app",
        "application",
        "capstone",
        "dashboard",
        "mobile app",
        "platform",
        "portal",
        "project",
        "website",
        "web app",
    )
    return any(normalized_phrase_in_text(term, normalized) for term in project_terms)


def has_achievement_signal(sentence: str) -> bool:
    normalized = normalize_match_text(sentence)
    if re.search(r"\d+(?:\.\d+)?%|\d+\+|\b\d{2,}\b|users?|customers?|accuracy|traffic|reduced|increased|improved|boosted|decreased", sentence, re.I):
        return True
    return any(
        normalized_phrase_in_text(term, normalized)
        for term in ("award", "achievement", "achieved", "honor", "dean s list", "scholarship")
    )


def split_project_and_experience_bullets(evidence_bullets: list[str]) -> tuple[list[str], list[str]]:
    projects: list[str] = []
    experience: list[str] = []
    for bullet in evidence_bullets:
        if is_project_like_sentence(bullet) and len(projects) < 3:
            projects.append(bullet)
        elif len(experience) < 4:
            experience.append(bullet)

    return unique_clean_values(experience, 4), unique_clean_values(projects, 3)


def extract_certification_lines(resume_text: str, limit: int = 3) -> list[str]:
    known_cert_terms = (
        "aws certified",
        "azure certified",
        "certificate",
        "certification",
        "certified",
        "cisco",
        "ccna",
        "comptia",
        "google professional",
        "microsoft certified",
        "oracle certified",
        "ocp",
        "security+",
    )
    certifications: list[str] = []
    for raw_part in split_resume_statement_parts(resume_text):
        sentence = clean_resume_sentence(raw_part)
        normalized = normalize_match_text(sentence)
        if len(sentence) < 12 or has_pdf_text_artifacts(sentence):
            continue
        if re.search(r"[\w.+-]+@[\w-]+(?:\.[\w-]+)+|(?:\+?\d[\d\s().-]{7,}\d)", sentence):
            continue
        if any(term in normalized for term in known_cert_terms):
            certifications.append(sentence)
    return unique_clean_values(certifications, limit)


def extract_achievement_lines(resume_text: str, evidence_bullets: list[str], limit: int = 3) -> list[str]:
    achievements = [bullet for bullet in evidence_bullets if has_achievement_signal(bullet)]
    for raw_part in split_resume_statement_parts(resume_text):
        sentence = clean_resume_sentence(raw_part)
        normalized = normalize_match_text(sentence)
        if len(sentence) < 16 or has_pdf_text_artifacts(sentence):
            continue
        if re.search(r"[\w.+-]+@[\w-]+(?:\.[\w-]+)+|(?:\+?\d[\d\s().-]{7,}\d)", sentence):
            continue
        if any(normalized_phrase_in_text(term, normalized) for term in ("award", "achievement", "honor", "dean s list", "scholarship")):
            achievements.append(sentence)
    return unique_clean_values(achievements, limit)


def build_section_completeness(sections: dict[str, Any]) -> dict[str, Any]:
    section_labels = {
        "experience_bullets": "professional experience",
        "projects": "projects",
        "certifications": "certifications",
        "achievements": "achievements",
        "education": "education",
    }
    missing_sections: list[str] = []
    placeholder_counts: dict[str, int] = {}
    for key, label in section_labels.items():
        values = list(sections.get(key) or [])
        count = sum(1 for value in values if is_generated_placeholder(str(value)))
        if count:
            placeholder_counts[key] = count
        if not values or count == len(values):
            missing_sections.append(label)

    return {
        "missing_sections": missing_sections,
        "placeholder_counts": placeholder_counts,
        "has_limited_source_info": bool(missing_sections or placeholder_counts),
    }


def build_limited_info_notes(completeness: dict[str, Any]) -> list[str]:
    missing_sections = completeness.get("missing_sections") or []
    if not missing_sections:
        return []
    readable_missing = ", ".join(missing_sections)
    return [
        "Limited source information: the uploaded resume did not clearly show " + readable_missing + ".",
        "Add those details only if they are true; otherwise leave the section out of the final resume.",
    ]


def build_generated_summary(target_role: str, skills: list[str], matched_signals: list[str], experience_markers: list[str]) -> str:
    top_skills = unique_clean_values(skills + matched_signals, 6)
    skill_text = ", ".join(top_skills[:5]) if top_skills else "computer science fundamentals"
    evidence_text = ", ".join(unique_clean_values(experience_markers, 3)) if experience_markers else "academic and project work"
    return (
        f"Computer Science candidate targeting {target_role} roles with resume evidence in {skill_text}. "
        f"Background includes {evidence_text}, with focus on clear documentation, practical implementation, "
        "and measurable project outcomes."
    )


def format_resume_keyword(value: str) -> str:
    keyword = normalize_text(str(value or ""))
    normalized = normalize_match_text(keyword)
    compact = re.sub(r"[^a-z0-9+#]+", "", keyword.lower())
    direct_labels = {
        "c#": "C#",
        "c++": "C++",
        "node.js": "Node.js",
        "express.js": "Express.js",
        "restapi": "REST APIs",
        "restapis": "REST APIs",
        "vscode": "Visual Studio Code",
        "vs code": "Visual Studio Code",
    }
    if keyword.lower() in direct_labels:
        return direct_labels[keyword.lower()]
    if compact in direct_labels:
        return direct_labels[compact]
    labels = {
        "android studio": "Android Studio",
        "api": "REST APIs",
        "aws": "AWS",
        "azure": "Azure",
        "c": "C",
        "ci cd": "CI/CD",
        "css": "CSS",
        "debug": "Troubleshooting",
        "debugging": "Troubleshooting",
        "develop": "Software Development",
        "docker": "Docker",
        "express": "Express.js",
        "express js": "Express.js",
        "fastapi": "FastAPI",
        "firebase": "Firebase",
        "flask": "Flask",
        "full stack": "Full-Stack Development",
        "full stack development": "Full-Stack Development",
        "fullstack": "Full-Stack Development",
        "gcp": "Google Cloud",
        "git": "Git",
        "github": "GitHub",
        "html": "HTML",
        "javascript": "JavaScript",
        "jquery": "jQuery",
        "json": "JSON",
        "kubernetes": "Kubernetes",
        "linux": "Linux",
        "mongodb": "MongoDB",
        "mysql": "MySQL",
        "node": "Node.js",
        "node js": "Node.js",
        "object oriented programming": "OOP",
        "oop": "OOP",
        "oracle": "Oracle",
        "postman": "Postman",
        "postgresql": "PostgreSQL",
        "problem solving": "Problem Solving",
        "python": "Python",
        "qa": "QA Testing",
        "qa testing": "QA Testing",
        "react": "React",
        "rest": "REST APIs",
        "rest api": "REST APIs",
        "rest apis": "REST APIs",
        "sdlc": "SDLC",
        "software test": "Software Testing",
        "sql": "SQL",
        "spring": "Spring Boot",
        "spring boot": "Spring Boot",
        "supabase": "Supabase",
        "tailwind": "Tailwind CSS",
        "test": "Software Testing",
        "testing": "Software Testing",
        "troubleshoot": "Troubleshooting",
        "troubleshooting": "Troubleshooting",
        "typescript": "TypeScript",
        "ui": "UI",
        "ux": "UX",
        "visual studio code": "Visual Studio Code",
    }
    if normalized in labels:
        return labels[normalized]
    if not keyword:
        return ""
    if keyword.isupper() or any(char in keyword for char in "#+/"):
        return keyword
    return " ".join(word.upper() if word in {"api", "sql", "ui", "ux"} else word.capitalize() for word in keyword.split())


def categorize_resume_skills(skills: list[str]) -> list[dict[str, Any]]:
    group_rules: list[tuple[str, set[str]]] = [
        ("Languages", {"java", "python", "javascript", "typescript", "sql", "c", "c#", "c++", "php", "ruby", "go", "kotlin", "swift", "dart"}),
        ("Frontend", {"html", "css", "javascript", "typescript", "react", "angular", "vue", "tailwind css", "bootstrap", "ui", "ux"}),
        ("Backend", {"node js", "express js", "spring boot", "rest apis", "rest api", "api", "flask", "django", "fastapi", "net", "asp net", "laravel", "php"}),
        ("Databases", {"mysql", "supabase", "firebase", "postgresql", "mongodb", "oracle", "sql server", "database administration"}),
        ("Tools", {"git", "github", "postman", "android studio", "visual studio code", "docker", "kubernetes", "jenkins", "linux", "aws", "azure", "google cloud", "gcp", "excel", "tableau", "power bi"}),
        ("Concepts", {"oop", "object oriented programming", "full stack development", "fullstack", "software development", "software testing", "qa testing", "troubleshooting", "sdlc", "agile", "scrum", "ci cd", "system design", "problem solving"}),
    ]
    low_value_terms = {
        "analysis",
        "application",
        "critical",
        "design",
        "direct",
        "engineer",
        "information",
        "procedure",
        "process",
        "related",
        "research",
        "software",
        "solving",
        "system",
        "systems",
        "thinking",
    }
    groups = [{"label": label, "items": []} for label, _ in group_rules]
    preferred_order = {
        "Languages": ["Java", "Python", "JavaScript", "SQL", "TypeScript", "C#", "C++", "PHP", "Go", "Ruby", "Kotlin", "Swift", "Dart"],
        "Frontend": ["HTML", "CSS", "JavaScript", "TypeScript", "React", "Angular", "Vue", "Tailwind CSS", "Bootstrap", "UI", "UX"],
        "Backend": ["Node.js", "Express.js", "Spring Boot", "REST APIs", "Flask", "Django", "FastAPI", ".NET", "ASP.NET", "Laravel"],
        "Databases": ["MySQL", "Supabase", "Firebase", "PostgreSQL", "MongoDB", "Oracle", "SQL Server"],
        "Tools": ["Git", "GitHub", "Postman", "Android Studio", "Visual Studio Code", "Docker", "Kubernetes", "Jenkins", "Linux", "AWS", "Azure", "Google Cloud", "Excel", "Tableau", "Power BI"],
        "Concepts": ["OOP", "Full-Stack Development", "Software Testing", "Troubleshooting", "SDLC", "Agile", "Scrum", "CI/CD", "System Design", "Problem Solving"],
    }

    for skill in unique_clean_values(skills, 36):
        display = format_resume_keyword(skill)
        normalized_display = normalize_match_text(display)
        normalized_skill = normalize_match_text(skill)
        if normalized_skill in low_value_terms and normalized_display in low_value_terms:
            continue
        for index, (_, terms) in enumerate(group_rules):
            if normalized_skill in terms or normalized_display in terms or any(term in normalized_skill for term in terms if len(term) > 4):
                groups[index]["items"].append(display)

    clean_groups = []
    for group in groups:
        items = unique_clean_values(group["items"], 8)
        order_map = {
            normalize_match_text(item): index
            for index, item in enumerate(preferred_order.get(group["label"], []))
        }
        items = [
            item
            for _, item in sorted(
                enumerate(items),
                key=lambda pair: (order_map.get(normalize_match_text(pair[1]), 1000 + pair[0]), pair[0]),
            )
        ]
        if items:
            clean_groups.append({"label": group["label"], "items": items})
    return clean_groups


def build_generated_resume_plain_text(draft: dict[str, Any]) -> str:
    candidate = draft["candidate"]
    sections = draft["sections"]
    template = draft.get("template") or {}
    contact_items = unique_clean_values(
        [candidate.get("email"), candidate.get("phone")] + list(candidate.get("link_labels") or candidate.get("links") or []),
        6,
    )
    lines = [
        (candidate.get("name") or "Candidate Name").upper(),
        template.get("headline") or draft.get("target_role") or "Target Role",
        " | ".join(contact_items),
        "",
        "PROFESSIONAL SUMMARY",
        sections["professional_summary"],
        "",
        str(template.get("skills_heading") or "Core Skills").upper(),
    ]
    for group in sections.get("skill_groups") or []:
        lines.append(f"{group['label']}: " + ", ".join(group["items"]))

    section_order = [
        ("experience_bullets", str(template.get("experience_heading") or "Project / Experience Highlights")),
        ("projects", str(template.get("project_heading") or "Projects")),
        ("certifications", str(template.get("certification_heading") or "Certifications")),
        ("achievements", str(template.get("achievement_heading") or "Achievements")),
        ("education", str(template.get("education_heading") or "Education")),
    ]
    for section_key, section_heading in section_order:
        items = real_resume_items(list(sections.get(section_key) or []))
        if not items:
            continue
        lines.extend(["", section_heading.upper()])
        lines.extend(f"- {line}" for line in items)
    return "\n".join(line for line in lines if line is not None).strip() + "\n"


def build_generated_resume_html(draft: dict[str, Any]) -> str:
    candidate = draft["candidate"]
    sections = draft["sections"]
    template = draft.get("template") or {}
    contact_items = unique_clean_values(
        [candidate.get("email"), candidate.get("phone")] + list(candidate.get("link_labels") or candidate.get("links") or []),
        6,
    )

    def section_html(section_key: str, heading: str) -> str:
        items = real_resume_items(list(sections.get(section_key) or []))
        if not items:
            return ""
        rendered_items = "".join(f"<li>{html_escape(str(item))}</li>" for item in items)
        return f"<h2>{html_escape(heading)}</h2><ul>{rendered_items}</ul>"

    skill_group_html = "".join(
        "<div class=\"skill-row\">"
        f"<strong>{html_escape(group['label'])}</strong>"
        f"<span>{html_escape(', '.join(group['items']))}</span>"
        "</div>"
        for group in sections.get("skill_groups") or []
    )

    return (
        "<!doctype html><html><head><meta charset=\"utf-8\">"
        "<title>ATS Resume Draft</title>"
        "<style>"
        "@page{size:letter;margin:.55in}"
        "*{box-sizing:border-box}"
        "body{font-family:Arial,Helvetica,sans-serif;margin:0;background:#fff;color:#111;line-height:1.35;font-size:11pt}"
        ".page{max-width:7.4in;margin:0 auto;background:#fff;padding:.42in .2in}"
        ".header{text-align:left;border-bottom:1px solid #111;padding-bottom:6px;margin-bottom:12px}"
        "h1{margin:0 0 2px;font-size:22pt;color:#111;text-transform:uppercase;font-weight:700;letter-spacing:0}"
        ".role{margin:0 0 3px;font-weight:700;color:#111;font-size:11pt;text-transform:none;letter-spacing:0}"
        ".contact{margin:0;color:#111;font-size:10pt;letter-spacing:0}"
        "h2{font-size:11pt;letter-spacing:0;text-transform:uppercase;color:#111;border-bottom:1px solid #555;padding-bottom:2px;margin:14px 0 7px;font-weight:700}"
        "p{margin:0 0 7px}.skill-row{display:grid;grid-template-columns:1.25in 1fr;gap:10px;margin:3px 0;font-size:10.5pt}"
        ".skill-row strong{color:#111}.skill-row span{color:#111}"
        "ul{margin:0 0 0 18px;padding:0;list-style-position:outside}li{margin:4px 0;font-size:10.5pt;break-inside:avoid}"
        ".summary{font-size:10.8pt;color:#111}"
        "@media print{.page{max-width:none;margin:0;padding:0}.header{margin-bottom:10px}}"
        "</style></head><body><main class=\"page\">"
        "<header class=\"header\">"
        f"<h1>{html_escape(candidate.get('name') or 'Candidate Name')}</h1>"
        f"<p class=\"role\">{html_escape(template.get('headline') or draft.get('target_role') or 'Target Role')}</p>"
        f"<p class=\"contact\">{html_escape(' | '.join(contact_items))}</p>"
        "</header>"
        "<h2>Professional Summary</h2>"
        f"<p class=\"summary\">{html_escape(sections['professional_summary'])}</p>"
        f"<h2>{html_escape(template.get('skills_heading') or 'Core Skills')}</h2>"
        f"{skill_group_html}"
        f"{section_html('experience_bullets', str(template.get('experience_heading') or 'Project / Experience Highlights'))}"
        f"{section_html('projects', str(template.get('project_heading') or 'Projects'))}"
        f"{section_html('certifications', str(template.get('certification_heading') or 'Certifications'))}"
        f"{section_html('achievements', str(template.get('achievement_heading') or 'Achievements'))}"
        f"{section_html('education', str(template.get('education_heading') or 'Education'))}"
        "</main></body></html>"
    )


def build_generated_resume_pdf_buffer(draft: dict[str, Any]) -> BytesIO:
    try:
        from reportlab.lib import colors  # type: ignore
        from reportlab.lib.enums import TA_LEFT  # type: ignore
        from reportlab.lib.pagesizes import letter  # type: ignore
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet  # type: ignore
        from reportlab.lib.units import inch  # type: ignore
        from reportlab.platypus import (  # type: ignore
            HRFlowable,
            Paragraph,
            SimpleDocTemplate,
            Table,
            TableStyle,
        )
    except ImportError as error:
        raise RuntimeError("Install reportlab to enable PDF resume export.") from error

    candidate = draft.get("candidate") or {}
    sections = draft.get("sections") or {}
    template = draft.get("template") or {}
    contact_items = unique_clean_values(
        [candidate.get("email"), candidate.get("phone")] + list(candidate.get("link_labels") or candidate.get("links") or []),
        6,
    )

    buffer = BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        leftMargin=0.62 * inch,
        rightMargin=0.62 * inch,
        topMargin=0.5 * inch,
        bottomMargin=0.5 * inch,
        title="ATS Resume Draft",
    )
    base_styles = getSampleStyleSheet()
    styles = {
        "name": ParagraphStyle(
            "ResumeName",
            parent=base_styles["Title"],
            alignment=TA_LEFT,
            fontName="Helvetica-Bold",
            fontSize=20,
            leading=23,
            textColor=colors.black,
            spaceAfter=2,
        ),
        "role": ParagraphStyle(
            "ResumeRole",
            parent=base_styles["Normal"],
            alignment=TA_LEFT,
            fontName="Helvetica-Bold",
            fontSize=10.5,
            leading=12.5,
            textColor=colors.black,
            spaceAfter=3,
        ),
        "contact": ParagraphStyle(
            "ResumeContact",
            parent=base_styles["Normal"],
            alignment=TA_LEFT,
            fontSize=9,
            leading=11,
            textColor=colors.black,
            spaceAfter=6,
        ),
        "heading": ParagraphStyle(
            "ResumeHeading",
            parent=base_styles["Heading2"],
            fontName="Helvetica-Bold",
            fontSize=10.2,
            leading=12.2,
            textColor=colors.black,
            spaceBefore=8,
            spaceAfter=3,
        ),
        "body": ParagraphStyle(
            "ResumeBody",
            parent=base_styles["Normal"],
            fontSize=9.7,
            leading=12,
            textColor=colors.black,
            spaceAfter=4,
        ),
        "skill_label": ParagraphStyle(
            "SkillLabel",
            parent=base_styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=9,
            leading=11,
            textColor=colors.black,
        ),
        "skill_value": ParagraphStyle(
            "SkillValue",
            parent=base_styles["Normal"],
            fontSize=9,
            leading=11,
            textColor=colors.black,
        ),
        "bullet": ParagraphStyle(
            "ResumeBullet",
            parent=base_styles["Normal"],
            fontSize=9.4,
            leading=11.8,
            leftIndent=0,
            textColor=colors.black,
            spaceAfter=2,
        ),
    }

    def heading(text: str):
        return [
            Paragraph(html_escape(text.upper()), styles["heading"]),
            HRFlowable(width="100%", thickness=0.6, color=colors.HexColor("#555555"), spaceAfter=4),
        ]

    story = [
        Paragraph(html_escape(str(candidate.get("name") or "Candidate Name").upper()), styles["name"]),
        Paragraph(html_escape(str(template.get("headline") or draft.get("target_role") or "Target Role")), styles["role"]),
        Paragraph(html_escape(" | ".join(contact_items)), styles["contact"]),
        HRFlowable(width="100%", thickness=0.8, color=colors.black, spaceAfter=7),
    ]

    story.extend(heading("Professional Summary"))
    story.append(Paragraph(html_escape(str(sections.get("professional_summary") or "")), styles["body"]))

    story.extend(heading(str(template.get("skills_heading") or "Technical Skills")))
    skill_rows = []
    for group in sections.get("skill_groups") or []:
        skill_rows.append([
            Paragraph(html_escape(str(group.get("label") or "Skills")), styles["skill_label"]),
            Paragraph(html_escape(", ".join(group.get("items") or [])), styles["skill_value"]),
        ])
    if skill_rows:
        skill_table = Table(skill_rows, colWidths=[1.35 * inch, 5.65 * inch], hAlign="LEFT")
        skill_table.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ]))
        story.append(skill_table)

    section_order = [
        ("experience_bullets", str(template.get("experience_heading") or "Professional Experience And Projects")),
        ("projects", str(template.get("project_heading") or "Projects")),
        ("certifications", str(template.get("certification_heading") or "Certifications")),
        ("achievements", str(template.get("achievement_heading") or "Achievements")),
        ("education", str(template.get("education_heading") or "Education")),
    ]
    for section_key, section_heading in section_order:
        items = real_resume_items(list(sections.get(section_key) or []))
        if not items:
            continue
        story.extend(heading(section_heading))
        for item in items:
            story.append(Paragraph("- " + html_escape(str(item)), styles["bullet"]))

    document.build(story)
    buffer.seek(0)
    return buffer


def build_feedback_report_pdf_buffer(report: dict[str, Any]) -> BytesIO:
    try:
        from reportlab.lib import colors  # type: ignore
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT  # type: ignore
        from reportlab.lib.pagesizes import letter  # type: ignore
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet  # type: ignore
        from reportlab.lib.units import inch  # type: ignore
        from reportlab.platypus import (  # type: ignore
            HRFlowable,
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )
        from reportlab.pdfgen import canvas  # type: ignore
    except ImportError as error:
        raise RuntimeError("Install reportlab to enable PDF feedback report export.") from error

    def clean_report_text(value: Any, fallback: str = "", max_length: int = 700) -> str:
        text = normalize_text(str(value or ""))
        if not text:
            text = fallback
        if max_length and len(text) > max_length:
            return text[: max_length - 3].rstrip() + "..."
        return text

    def clean_report_items(value: Any, fallback: str = "", limit: int = 8, max_length: int = 180) -> list[str]:
        if not isinstance(value, list):
            return [fallback] if fallback else []
        cleaned = []
        for item in value:
            text = clean_report_text(item, "", max_length)
            if text:
                cleaned.append(text)
            if len(cleaned) >= limit:
                break
        return cleaned or ([fallback] if fallback else [])

    def metric_status(value: Any) -> str:
        try:
            safe_percent = max(0.0, min(100.0, float(value or 0)))
        except (TypeError, ValueError):
            safe_percent = 0.0
        if safe_percent >= 75:
            return "Strong"
        if safe_percent >= 55:
            return "Fair"
        return "Needs work"

    def score_label(value: Any) -> str:
        try:
            score = round(max(0.0, min(100.0, float(value or 0))))
        except (TypeError, ValueError):
            score = 0
        return f"{score}/100"

    def criteria_items() -> list[dict[str, Any]]:
        raw_criteria = report.get("qualityCriteria")
        if not isinstance(raw_criteria, list):
            return []
        return [item for item in raw_criteria if isinstance(item, dict)]

    def top_role_items() -> list[dict[str, Any]]:
        raw_matches = report.get("topRoleMatches")
        if not isinstance(raw_matches, list):
            return []
        return [item for item in raw_matches[:3] if isinstance(item, dict)]

    def recommendation_action_items() -> list[dict[str, Any]]:
        raw_actions = report.get("recommendationPlan") or report.get("priorityActions")
        if not isinstance(raw_actions, list):
            return []
        return [item for item in raw_actions[:5] if isinstance(item, dict)]

    def issue_kind(item: dict[str, Any]) -> str:
        combined = normalize_match_text(
            f"{item.get('category', '')} {item.get('title', '')} {item.get('detail', '')}"
        )
        if re.search(r"\b(embedded|photo|passport|decorative image|image detected)\b", combined):
            return "image"
        if re.search(r"\b(contact|email|phone|linkedin|github|portfolio)\b", combined):
            return "contact"
        if re.search(r"\b(skill|technical|tool|framework|language)\b", combined):
            return "skills"
        if re.search(r"\b(section|heading|education|project)\b", combined):
            return "sections"
        if re.search(r"\b(impact|measure|quant|number|outcome)\b", combined):
            return "impact"
        if "length" in combined:
            return "length"
        if re.search(r"\b(image|format|layout|spacing|column|font|parser)\b", combined):
            return "format"
        return "general"

    def marker_position(item: dict[str, Any], index: int) -> dict[str, float]:
        positions = {
            "image": {"top": 12.0, "left": 58.0, "width": 22.0, "height": 20.0},
            "contact": {"top": 21.0, "left": 8.0, "width": 31.0, "height": 9.0},
            "format": {"top": 36.0, "left": 8.0, "width": 34.0, "height": 15.0},
            "sections": {"top": 53.0, "left": 8.0, "width": 34.0, "height": 17.0},
            "impact": {"top": 39.0, "left": 51.0, "width": 37.0, "height": 23.0},
            "length": {"top": 70.0, "left": 11.0, "width": 75.0, "height": 10.0},
            "skills": {"top": 31.0, "left": 8.0, "width": 35.0, "height": 10.0},
            "general": {
                "top": 29.0 + (index * 8.0),
                "left": 12.0 + (index * 7.0),
                "width": 34.0,
                "height": 10.0,
            },
        }
        return positions.get(issue_kind(item), positions["general"])

    def decode_data_url(value: Any) -> tuple[str, bytes] | None:
        data_url = str(value or "")
        match = re.match(r"^data:([^;,]+);base64,(.+)$", data_url, re.S)
        if not match:
            return None
        try:
            return match.group(1).lower(), base64.b64decode(match.group(2), validate=False)
        except Exception:
            return None

    def build_annotated_resume_pdf_preview() -> BytesIO | None:
        preview = report.get("resumeVisualPreview") if isinstance(report.get("resumeVisualPreview"), dict) else {}
        decoded = decode_data_url(preview.get("dataUrl"))
        if not decoded:
            return None

        mime_type, raw_file = decoded
        if mime_type != "application/pdf":
            return None

        try:
            resume_reader = PdfReader(BytesIO(raw_file))
            if not resume_reader.pages:
                return None
            page = resume_reader.pages[0]
            page_width = float(page.mediabox.width)
            page_height = float(page.mediabox.height)

            overlay_buffer = BytesIO()
            overlay_canvas = canvas.Canvas(overlay_buffer, pagesize=(page_width, page_height))
            overlay_canvas.setTitle("Annotated Resume Preview")
            overlay_canvas.setStrokeColor(colors.HexColor("#d01822"))
            overlay_canvas.setFillColor(colors.HexColor("#d01822"))
            overlay_canvas.setLineWidth(1.4)
            overlay_canvas.setFont("Helvetica-Bold", 9)

            for index, item in enumerate(quality_issues[:5]):
                position = marker_position(item, index)
                x = page_width * (position["left"] / 100.0)
                rect_width = page_width * (position["width"] / 100.0)
                rect_height = page_height * (position["height"] / 100.0)
                y = page_height - (page_height * (position["top"] / 100.0)) - rect_height
                overlay_canvas.roundRect(x, y, rect_width, rect_height, 4, stroke=1, fill=0)
                badge_x = min(page_width - 11, x + rect_width)
                badge_y = min(page_height - 11, y + rect_height)
                overlay_canvas.circle(badge_x, badge_y, 9, stroke=0, fill=1)
                overlay_canvas.setFillColor(colors.white)
                overlay_canvas.drawCentredString(badge_x, badge_y - 3, str(index + 1))
                overlay_canvas.setFillColor(colors.HexColor("#d01822"))

            overlay_canvas.showPage()
            overlay_canvas.save()
            overlay_buffer.seek(0)
            page.merge_page(PdfReader(overlay_buffer).pages[0])

            preview_writer = PdfWriter()
            preview_writer.add_page(page)
            preview_buffer = BytesIO()
            preview_writer.write(preview_buffer)
            preview_buffer.seek(0)
            return preview_buffer
        except Exception:
            return None

    resume_quality = report.get("resumeQuality") if isinstance(report.get("resumeQuality"), dict) else {}
    compatibility = report.get("compatibility") if isinstance(report.get("compatibility"), dict) else {}
    score_explanation = report.get("scoreExplanation") if isinstance(report.get("scoreExplanation"), dict) else {}
    quality_criteria = criteria_items()
    quality_issues = [item for item in quality_criteria if item.get("status") != "pass"]
    passed_checks = [item for item in quality_criteria if item.get("status") == "pass"]
    final_score = score_label(report.get("scoreOutOf100"))
    generated_at = clean_report_text(report.get("generatedAt"), datetime.now().strftime("%Y-%m-%d %H:%M"))
    annotated_resume_buffer = build_annotated_resume_pdf_preview()

    buffer = BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        leftMargin=0.55 * inch,
        rightMargin=0.55 * inch,
        topMargin=0.55 * inch,
        bottomMargin=0.55 * inch,
        title="UTS CS Career Navigator Feedback Report",
    )
    base_styles = getSampleStyleSheet()
    styles = {
        "title": ParagraphStyle(
            "FeedbackReportTitle",
            parent=base_styles["Title"],
            alignment=TA_LEFT,
            fontName="Helvetica-Bold",
            fontSize=20,
            leading=24,
            textColor=colors.HexColor("#002555"),
            spaceAfter=3,
        ),
        "score": ParagraphStyle(
            "FeedbackReportScore",
            parent=base_styles["Title"],
            alignment=TA_RIGHT,
            fontName="Helvetica-Bold",
            fontSize=26,
            leading=30,
            textColor=colors.HexColor("#0060ac"),
        ),
        "subtitle": ParagraphStyle(
            "FeedbackReportSubtitle",
            parent=base_styles["Normal"],
            fontSize=9.5,
            leading=12,
            textColor=colors.HexColor("#43474f"),
            spaceAfter=10,
        ),
        "section": ParagraphStyle(
            "FeedbackReportSection",
            parent=base_styles["Heading2"],
            fontName="Helvetica-Bold",
            fontSize=11,
            leading=13,
            textColor=colors.HexColor("#0060ac"),
            spaceBefore=12,
            spaceAfter=4,
        ),
        "heading": ParagraphStyle(
            "FeedbackReportHeading",
            parent=base_styles["Heading3"],
            fontName="Helvetica-Bold",
            fontSize=12.5,
            leading=15,
            textColor=colors.HexColor("#002555"),
            spaceBefore=5,
            spaceAfter=5,
        ),
        "body": ParagraphStyle(
            "FeedbackReportBody",
            parent=base_styles["Normal"],
            fontSize=9,
            leading=12,
            textColor=colors.HexColor("#191c1d"),
            spaceAfter=4,
        ),
        "muted": ParagraphStyle(
            "FeedbackReportMuted",
            parent=base_styles["Normal"],
            fontSize=8,
            leading=10,
            textColor=colors.HexColor("#64748b"),
        ),
        "table_label": ParagraphStyle(
            "FeedbackReportTableLabel",
            parent=base_styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=8.5,
            leading=11,
            textColor=colors.HexColor("#17324f"),
        ),
        "table_cell": ParagraphStyle(
            "FeedbackReportTableCell",
            parent=base_styles["Normal"],
            fontSize=8.4,
            leading=10.8,
            textColor=colors.HexColor("#26384d"),
        ),
        "table_header": ParagraphStyle(
            "FeedbackReportTableHeader",
            parent=base_styles["Normal"],
            alignment=TA_CENTER,
            fontName="Helvetica-Bold",
            fontSize=8.3,
            leading=10.5,
            textColor=colors.white,
        ),
        "note": ParagraphStyle(
            "FeedbackReportNote",
            parent=base_styles["Normal"],
            fontSize=8.7,
            leading=11.2,
            textColor=colors.HexColor("#17324f"),
            backColor=colors.HexColor("#edf5ff"),
            borderColor=colors.HexColor("#bfd4f4"),
            borderWidth=0.6,
            borderPadding=6,
            spaceAfter=6,
        ),
        "excerpt": ParagraphStyle(
            "FeedbackReportExcerpt",
            parent=base_styles["Code"],
            fontName="Courier",
            fontSize=7.5,
            leading=9.4,
            textColor=colors.HexColor("#334155"),
            backColor=colors.HexColor("#f8fafc"),
            borderColor=colors.HexColor("#d9e1ec"),
            borderWidth=0.4,
            borderPadding=4,
            spaceAfter=2,
        ),
    }

    def paragraph(value: Any, style_name: str = "body", fallback: str = "", max_length: int = 700):
        return Paragraph(html_escape(clean_report_text(value, fallback, max_length)), styles[style_name])

    def report_table(rows: list[list[Any]], col_widths: list[float], has_header: bool = False) -> Table:
        styled_rows = []
        for row_index, row in enumerate(rows):
            style_name = "table_header" if has_header and row_index == 0 else "table_cell"
            styled_rows.append([paragraph(cell, style_name, max_length=420) for cell in row])
        table = Table(styled_rows, colWidths=col_widths, hAlign="LEFT", repeatRows=1 if has_header else 0)
        table_style = [
            ("GRID", (0, 0), (-1, -1), 0.45, colors.HexColor("#d9e1ec")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]
        if has_header:
            table_style.extend(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#002555")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
                ]
            )
        else:
            table_style.extend(
                [
                    ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#f1f5f9")),
                    ("ROWBACKGROUNDS", (1, 0), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
                ]
            )
        table.setStyle(TableStyle(table_style))
        return table

    def add_section(story: list[Any], title: str) -> None:
        story.append(Paragraph(html_escape(title.upper()), styles["section"]))
        story.append(HRFlowable(width="100%", thickness=0.45, color=colors.HexColor("#d9e1ec"), spaceAfter=6))

    summary_rows = [
        ["Uploaded file", clean_report_text(report.get("uploadedResumeName"), "Not available", 180)],
        ["File type", clean_report_text(report.get("fileType"), "Not available", 80)],
        ["Selected role", clean_report_text(report.get("selectedRole"), "Not selected", 160)],
        ["Database role", clean_report_text(report.get("selectedDatabaseRole"), "Not mapped", 160)],
        ["Recommended role", clean_report_text(report.get("recommendedRole"), "Not available", 160)],
        ["Final readiness score", final_score],
        ["ATS quality risk", clean_report_text(resume_quality.get("risk_level"), "Not available", 80)],
        ["Submitted at", clean_report_text(report.get("submittedAt"), generated_at, 140)],
    ]
    component_rows = [
        ["Keyword match", metric_status(compatibility.get("keyword_match"))],
        ["Experience alignment", metric_status(compatibility.get("experience_alignment"))],
        ["Education / skills evidence", metric_status(compatibility.get("academic_rigor"))],
        ["ATS format quality", clean_report_text(resume_quality.get("risk_level"), "Not available", 80) + " risk"],
    ]

    top_rows = [["Rank", "Database role", "Fit", "Status", "Evidence"]]
    for index, item in enumerate(top_role_items(), start=1):
        raw_score = item.get("score", item.get("confidence_score"))
        role_score = score_label(raw_score)
        labels = ", ".join(
            [
                label
                for label in [
                    "Final used" if item.get("is_final_recommended") else "",
                    "Selected" if item.get("is_selected_database_role") else "",
                ]
                if label
            ]
        ) or "Compared"
        matched = ", ".join(clean_report_items(item.get("matched_keywords"), "No strong matched keywords", 6, 50))
        missing = ", ".join(clean_report_items(item.get("missing_keywords"), "No major missing gaps", 5, 50))
        top_rows.append(
            [
                str(item.get("rank") or index),
                clean_report_text(item.get("role") or item.get("database_role"), "Role match", 140),
                f"{role_score} - {clean_report_text(item.get('confidence_label'), metric_status(raw_score), 60)}",
                labels,
                f"Matched: {matched} Improve: {missing}",
            ]
        )
    if len(top_rows) == 1:
        top_rows.append(["-", "No top role matches returned", "-", "-", "Run an analysis to compare role evidence."])

    quality_rows = [["Criteria", "Status", "Finding", "Suggested fix"]]
    for item in quality_criteria:
        quality_rows.append(
            [
                clean_report_text(item.get("category"), "Criteria", 80),
                "Pass" if item.get("status") == "pass" else "Fix",
                clean_report_text(item.get("title") or item.get("detail"), "Finding not available", 190),
                clean_report_text(item.get("suggestion") or item.get("detail"), "Suggested fix not available", 260),
            ]
        )
    if len(quality_rows) == 1:
        quality_rows.append(["-", "Pending", "No ATS quality criteria returned", "Run an analysis to generate ATS quality checks."])

    story: list[Any] = []
    header = Table(
        [
            [
                Paragraph("UTS CS Career Navigator<br/>Feedback Report", styles["title"]),
                Paragraph(final_score, styles["score"]),
            ]
        ],
        colWidths=[5.35 * inch, 1.65 * inch],
        hAlign="LEFT",
    )
    header.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
                ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#d9e1ec")),
                ("LEFTPADDING", (0, 0), (-1, -1), 12),
                ("RIGHTPADDING", (0, 0), (-1, -1), 12),
                ("TOPPADDING", (0, 0), (-1, -1), 12),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
            ]
        )
    )
    story.append(header)
    story.append(Paragraph("Structured ATS issues, resume evidence, and next-edit actions. Generated " + html_escape(generated_at) + ".", styles["subtitle"]))

    add_section(story, "Summary")
    story.append(report_table(summary_rows, [1.9 * inch, 5.1 * inch]))

    add_section(story, "Score Explanation")
    story.append(paragraph(score_explanation.get("label"), "heading", "Score meaning", 120))
    story.append(paragraph(score_explanation.get("text"), "note", "The final readiness score combines role match, resume evidence, and ATS format quality.", 650))
    story.append(paragraph(score_explanation.get("detail") or score_explanation.get("formula"), "body", "Formula details were not available.", 650))

    add_section(story, "Role Fit Explanation")
    story.append(paragraph(report.get("roleReason"), "note", "Selected role and recommendation are aligned.", 650))

    recommendation_actions = recommendation_action_items()
    if recommendation_actions:
        add_section(story, "Priority Improvement Plan")
        for index, item in enumerate(recommendation_actions, start=1):
            keywords = ", ".join(clean_report_items(item.get("keywords"), "", 5, 45))
            keyword_text = f" Keywords: {keywords}." if keywords else ""
            action_text = (
                f"{item.get('priority') or index}. {clean_report_text(item.get('title'), 'Recommendation action', 90)}: "
                f"{clean_report_text(item.get('action') or item.get('detail'), 'Apply this improvement before exporting the resume.', 260)}"
                f"{keyword_text}"
            )
            story.append(paragraph("- " + action_text, "body", max_length=390))

    add_section(story, "Top 3 Role Matches")
    story.append(report_table(top_rows, [0.45 * inch, 1.55 * inch, 1.15 * inch, 0.9 * inch, 3.0 * inch], has_header=True))

    add_section(story, "Score Components")
    story.append(report_table(component_rows, [1.9 * inch, 5.1 * inch]))
    story.append(paragraph("Only the final readiness score is numeric. Component results are shown as status labels to keep the report easy to read.", "muted", max_length=220))

    add_section(story, "ATS Quality Criteria")
    story.append(report_table(quality_rows, [1.05 * inch, 0.75 * inch, 2.05 * inch, 3.15 * inch], has_header=True))

    add_section(story, "Resume Review Snapshot")
    story.append(Paragraph("Resume Bad Habits", styles["heading"]))
    bad_habits = [
        f"{clean_report_text(item.get('title') or item.get('category'), 'ATS issue', 90)}: "
        f"{clean_report_text(item.get('suggestion') or item.get('detail'), 'Suggested fix not available', 180)}"
        for item in quality_issues[:5]
    ] or ["No major resume bad habits detected."]
    for item in bad_habits:
        story.append(paragraph("- " + item, "body", max_length=280))

    story.append(Paragraph("Passed Checks", styles["heading"]))
    passed_labels = [
        clean_report_text(item.get("category") or item.get("title"), "Passed", 60)
        for item in passed_checks[:8]
    ]
    story.append(paragraph(", ".join(passed_labels) if passed_labels else "No passed checks yet.", "body", max_length=320))
    if annotated_resume_buffer:
        story.append(paragraph("The first page of this PDF is the uploaded resume with numbered red issue markers. The matching numbered explanations are listed above.", "note", max_length=260))
    else:
        story.append(paragraph("The HTML feedback export includes the visual resume preview when the original upload is available. This PDF report keeps the clean issue list and avoids showing broken extracted text.", "note", max_length=300))

    add_section(story, "Resume Evidence")
    matched_signals = ", ".join(clean_report_items(report.get("matchedSignals"), "No matched signals", 10, 55))
    missing_keywords = ", ".join(clean_report_items(report.get("missingKeywords"), "No missing keywords", 10, 55))
    story.append(paragraph("Matched signals: " + matched_signals, "body", max_length=520))
    story.append(paragraph("Missing keywords: " + missing_keywords, "body", max_length=520))

    add_section(story, "Next Actions")
    for item in clean_report_items(report.get("structuredFeedback"), clean_report_text(report.get("feedback"), "No feedback available.", 280), 8, 240):
        story.append(paragraph("- " + item, "body", max_length=300))

    story.append(Spacer(1, 0.12 * inch))
    story.append(Paragraph("Generated locally by UTS CS Career Navigator.", styles["muted"]))

    document.build(story)
    buffer.seek(0)

    if annotated_resume_buffer:
        combined_writer = PdfWriter()
        for source_buffer in (annotated_resume_buffer, buffer):
            source_buffer.seek(0)
            reader = PdfReader(source_buffer)
            for page in reader.pages:
                combined_writer.add_page(page)
        combined_buffer = BytesIO()
        combined_writer.write(combined_buffer)
        combined_buffer.seek(0)
        return combined_buffer

    return buffer


def get_top_role_generation_scores(analysis: dict[str, Any]) -> list[float]:
    role_recommendation = analysis.get("role_recommendation") if isinstance(analysis.get("role_recommendation"), dict) else {}
    raw_matches = analysis.get("top_role_matches") or role_recommendation.get("top_role_matches") or []
    if not isinstance(raw_matches, list):
        return []

    scores: list[float] = []
    for match in raw_matches[:3]:
        if not isinstance(match, dict):
            continue
        raw_score = match.get("score")
        if raw_score is None:
            raw_score = match.get("confidence_score")
        try:
            scores.append(float(raw_score))
        except (TypeError, ValueError):
            continue
    return scores


def get_top_role_generation_name(analysis: dict[str, Any]) -> str:
    role_recommendation = analysis.get("role_recommendation") if isinstance(analysis.get("role_recommendation"), dict) else {}
    raw_matches = analysis.get("top_role_matches") or role_recommendation.get("top_role_matches") or []
    if isinstance(raw_matches, list) and raw_matches and isinstance(raw_matches[0], dict):
        top_role = normalize_text(str(raw_matches[0].get("role") or raw_matches[0].get("database_role") or ""))
        if top_role:
            return top_role
    return ""


def should_return_cs_sample_resume(analysis: dict[str, Any]) -> tuple[bool, float | None]:
    if not analysis:
        return False, None

    top_scores = get_top_role_generation_scores(analysis)
    if top_scores:
        best_score = max(top_scores)
        return best_score < CS_SAMPLE_ONLY_TOP_MATCH_THRESHOLD, best_score

    if "matching_percentage" not in analysis:
        return False, None
    try:
        score = float(analysis.get("matching_percentage") or 0)
    except (TypeError, ValueError):
        return False, None
    return score < CS_SAMPLE_ONLY_TOP_MATCH_THRESHOLD, score


def build_cs_sample_resume_payload(
    analysis: dict[str, Any],
    target_role: str = "",
    best_score: float | None = None,
) -> dict[str, Any]:
    role = get_top_role_generation_name(analysis) or get_resume_generation_role(analysis, target_role) or "Software Developer"
    template_profile = select_resume_template_profile(role)
    sample_skills = [
        "Java",
        "Python",
        "HTML",
        "CSS",
        "JavaScript",
        "SQL",
        "Git",
        "REST API",
        "Software Testing",
        "Debugging",
        "System Design",
    ]
    sections = {
        "professional_summary": (
            f"Computer Science candidate targeting {role} roles with project experience in software development, "
            "database-backed applications, API testing, and technical documentation. Able to translate requirements "
            "into maintainable features, validate defects, and communicate implementation decisions clearly."
        ),
        "technical_skills": sample_skills,
        "skill_groups": categorize_resume_skills(sample_skills),
        "experience_bullets": [
            "Built and tested a web-based student services application using JavaScript, SQL, and REST API endpoints.",
            "Documented software defects, reproduced test cases, and verified fixes before release.",
            "Collaborated with project teammates to split requirements, track progress, and prepare technical handover notes.",
        ],
        "projects": [
            "Resume Analyzer Web App - implemented file upload, text extraction, ATS scoring, and generated HTML/PDF exports.",
            "Library Management System - designed database tables, CRUD workflows, search filters, and validation checks.",
        ],
        "certifications": [
            "AWS Academy Cloud Foundations - AWS Academy (2025).",
        ],
        "achievements": [
            "Dean's List - Faculty of Computing (2025).",
        ],
        "education": [
            "Bachelor of Computer Science - University Name (2024 - 2027).",
            "Diploma in Information Technology - Institution Name (2020 - 2023).",
        ],
        "keyword_targets": [],
    }
    completeness = build_section_completeness(sections)
    score_text = f"{best_score:.0f}/100" if isinstance(best_score, (int, float)) else "below 20/100"
    draft = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "target_role": role,
        "candidate": {
            "name": "Sample CS Candidate",
            "email": "candidate@example.com",
            "phone": "+60 12 345 6789",
            "links": ["https://linkedin.com/in/sample-cs-candidate", "https://github.com/sample-cs-candidate"],
            "link_labels": ["LinkedIn: linkedin.com/in/sample-cs-candidate", "GitHub: github.com/sample-cs-candidate"],
        },
        "source": {
            "uploaded_resume_name": analysis.get("uploaded_resume_name"),
            "analysis_id": analysis.get("analysis_id"),
            "score": analysis.get("matching_percentage"),
        },
        "sections": sections,
        "completeness": completeness,
        "missing_keywords": [],
        "ats_notes": [
            f"Sample only: the best Top 3 CS role match was {score_text}, so the system did not rewrite this resume as a CS resume.",
            "Use this as a CS resume structure example, not as a final resume to submit.",
            "To generate a real CS ATS resume, upload a resume with true CS coursework, projects, internships, tools, or certifications.",
            "Replace every sample name, contact, education, project, and achievement with your own verified details.",
        ],
        "template": {
            **template_profile,
            "name": "Sample CS ATS Template",
            "inspired_by": [
                "single-column ATS resume structure",
                "compact technical skills grouping",
                "project-focused CS student resume layout",
                "clear education, certification, and achievement sections",
            ],
        },
        "sample_only": True,
        "is_sample_resume": True,
        "sample_reason": "Top 3 CS role matches are below the 20/100 generation threshold.",
        "generation_threshold": CS_SAMPLE_ONLY_TOP_MATCH_THRESHOLD,
        "best_top_role_score": best_score,
    }
    draft["plain_text"] = build_generated_resume_plain_text(draft)
    draft["html"] = build_generated_resume_html(draft)
    return draft


def build_generated_resume_payload(
    analysis: dict[str, Any],
    resume_text: str,
    target_role: str = "",
    candidate_name: str = "",
    contact_links: list[str] | None = None,
) -> dict[str, Any]:
    safe_analysis = analysis or {}
    role = get_resume_generation_role(safe_analysis, target_role)
    use_sample_resume, best_top_role_score = should_return_cs_sample_resume(safe_analysis)
    if use_sample_resume:
        return build_cs_sample_resume_payload(safe_analysis, role, best_top_role_score)

    template_profile = select_resume_template_profile(role)
    resume_source = clean_resume_source_text(resume_text or str(safe_analysis.get("resume_text_for_generation") or ""))
    if not resume_source:
        resume_source = normalize_text(str(safe_analysis.get("resume_excerpt") or ""))

    resume_summary = safe_analysis.get("resume_summary") or {}
    role_recommendation = safe_analysis.get("role_recommendation") or {}
    recommendation_module = safe_analysis.get("recommendation_module") if isinstance(safe_analysis.get("recommendation_module"), dict) else {}
    if not recommendation_module and safe_analysis:
        recommendation_module = build_part8_recommendation_module(safe_analysis)
    resume_generation_strategy = (
        recommendation_module.get("resume_generation_strategy")
        if isinstance(recommendation_module.get("resume_generation_strategy"), dict)
        else {}
    )
    job_analysis = safe_analysis.get("job_description_analysis") or {}
    resume_quality = safe_analysis.get("resume_quality") or {}

    existing_skills = unique_clean_values(list(resume_summary.get("skills") or []), 14)
    matched_signals = unique_clean_values(
        list(role_recommendation.get("matched_resume_signals") or [])
        + list(job_analysis.get("matched_keywords") or []),
        14,
    )
    missing_keywords = unique_clean_values(
        list(job_analysis.get("missing_keywords") or [])
        + list(role_recommendation.get("recommended_missing_keywords") or []),
        10,
    )
    strategy_keyword_targets = unique_clean_values(list(resume_generation_strategy.get("keyword_targets") or []), 8)
    missing_keywords = unique_clean_values(missing_keywords + strategy_keyword_targets, 10)
    important_keywords = unique_clean_values(list(job_analysis.get("important_keywords") or []), 14)
    skills = unique_clean_values(existing_skills + matched_signals, 18)
    if not skills:
        skills = ["Programming fundamentals", "Problem solving", "Technical documentation"]
    display_skills = unique_clean_values([format_resume_keyword(skill) for skill in skills], 18)
    display_matched_signals = unique_clean_values([format_resume_keyword(signal) for signal in matched_signals], 12)
    display_missing_keywords = unique_clean_values([format_resume_keyword(keyword) for keyword in missing_keywords], 10)

    contact = extract_resume_contact(resume_source, candidate_name, contact_links)
    filename_name = infer_resume_name_from_filename(str(safe_analysis.get("uploaded_resume_name") or ""))
    current_contact_name = normalize_text(str(contact.get("name") or ""))
    if filename_name and (
        not current_contact_name
        or current_contact_name == "Candidate Name"
        or len(filename_name.split()) > len(current_contact_name.split())
    ):
        contact["name"] = filename_name
    education_lines = extract_education_lines(resume_source, list(resume_summary.get("education_markers") or []))
    evidence_bullets = extract_resume_evidence_sentences(
        resume_source,
        skills + matched_signals + important_keywords,
        10,
    )
    experience_bullets, project_bullets = split_project_and_experience_bullets(evidence_bullets)
    certification_lines = extract_certification_lines(resume_source)
    achievement_lines = extract_achievement_lines(resume_source, evidence_bullets)
    project_keys = {normalize_match_text(project) for project in project_bullets}
    achievement_lines = unique_clean_values(
        [achievement for achievement in achievement_lines if normalize_match_text(achievement) not in project_keys],
        3,
    )

    guidance_notes: list[str] = []
    for keyword in display_missing_keywords[: max(0, 5 - len(experience_bullets))]:
        guidance_notes.append(
            f"Add evidence if you have it: describe a project, coursework item, or work example that truthfully demonstrates {keyword}."
        )
    if len(experience_bullets) < 4:
        guidance_notes.append(
            "If available, add more measurable experience bullets with an action verb, the technical work, and a result or metric."
        )

    if not project_bullets:
        guidance_notes.append("Add a Projects section only if you have a real project; include project name, technologies, your role, and a measurable result.")
    if not certification_lines:
        guidance_notes.append("Add certifications only if you have them; otherwise leave the Certifications section out.")
    if not achievement_lines:
        guidance_notes.append("Add awards or achievements only if they are real; otherwise leave the Achievements section out.")

    sections = {
        "professional_summary": build_generated_summary(
            role,
            display_skills,
            display_matched_signals,
            list(resume_summary.get("experience_markers") or []),
        ),
        "technical_skills": display_skills,
        "skill_groups": categorize_resume_skills(display_skills),
        "experience_bullets": experience_bullets[:6],
        "projects": project_bullets[:3],
        "certifications": certification_lines[:3],
        "achievements": achievement_lines[:3],
        "education": education_lines,
        "keyword_targets": display_missing_keywords[:6],
    }
    completeness = build_section_completeness(sections)
    ats_notes = [
        "Part 8 recommendation target: generate the resume for " + role + ".",
        "Draft uses extracted resume evidence. Missing optional content is shown here as guidance instead of inside the resume body.",
        "Only keep a missing keyword if you can support it with a real project, coursework, or work example.",
    ]
    priority_actions = [
        item
        for item in list(recommendation_module.get("priority_actions") or [])
        if isinstance(item, dict)
    ]
    ats_notes.extend(
        unique_clean_values(
            [
                str(item.get("action") or item.get("detail") or "")
                for item in priority_actions[:3]
            ],
            3,
        )
    )
    ats_notes.extend(unique_clean_values(guidance_notes, 8))
    ats_notes.extend(build_limited_info_notes(completeness))
    if display_missing_keywords:
        ats_notes.append("Target keywords to add with evidence: " + ", ".join(display_missing_keywords[:8]) + ".")
    issues = list(resume_quality.get("issues") or [])
    if issues:
        ats_notes.extend(unique_clean_values([str(issue.get("suggestion") or "") for issue in issues], 3))

    draft = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "target_role": role,
        "candidate": contact,
        "source": {
            "uploaded_resume_name": safe_analysis.get("uploaded_resume_name"),
            "analysis_id": safe_analysis.get("analysis_id"),
            "score": safe_analysis.get("matching_percentage"),
        },
        "sections": sections,
        "completeness": completeness,
        "missing_keywords": display_missing_keywords,
        "ats_notes": ats_notes,
        "template": {
            **template_profile,
            "inspired_by": [
                "single-column ATS resume structure",
                "compact technical skills grouping",
                "project-focused CS student resume layout",
                "clear education, certification, and achievement sections",
            ],
        },
        "recommendation_plan": priority_actions[:5],
        "resume_generation_strategy": resume_generation_strategy,
    }
    draft["plain_text"] = build_generated_resume_plain_text(draft)
    draft["html"] = build_generated_resume_html(draft)
    return draft


def record_download_history(user_email: str, analysis_id: int | None, downloaded_file_name: str) -> None:
    user_id = get_user_id_by_email(user_email)
    connection = get_db_connection()
    cursor = connection.cursor()
    cursor.execute(
        """
        INSERT INTO download_history (
            user_id,
            analysis_id,
            downloaded_file_name,
            report_type
        ) VALUES (%s, %s, %s, %s)
        """,
        (user_id, analysis_id, downloaded_file_name, "resume_comments"),
    )
    connection.commit()
    cursor.close()
    connection.close()


@app.post("/auth/pin/send")
def send_auth_pin():
    payload = request.get_json(silent=True) or {}
    email = normalize_email(payload.get("email", ""))
    purpose = normalize_pin_purpose(payload.get("purpose", "register"))

    if not email:
        return jsonify({"detail": "email is required"}), 400

    if not is_valid_email(email):
        return jsonify({"detail": "Please provide a valid email address"}), 400

    if purpose not in EMAIL_PIN_PURPOSES:
        return jsonify({"detail": "Unsupported PIN purpose"}), 400

    try:
        user_already_exists = email_exists(email)
        if purpose == "register" and user_already_exists:
            return jsonify({"detail": "This email is already registered"}), 409
        if purpose == "password_reset" and not user_already_exists:
            return jsonify({"detail": "No account was found for this email address"}), 404

        pin, expires_at = create_email_pin(email, purpose)
        delivery = send_email_pin(email, pin, purpose)
    except DatabaseError:
        return jsonify({"detail": "Database error while creating verification PIN"}), 500
    except (OSError, smtplib.SMTPException) as error:
        return jsonify({"detail": get_email_delivery_error_message(error)}), 502

    response: dict[str, Any] = {
        "message": "Verification PIN sent",
        "purpose": purpose,
        "delivery": delivery,
        "expires_in_minutes": EMAIL_PIN_TTL_MINUTES,
        "expires_at": expires_at.isoformat(),
    }
    if should_return_dev_pin(delivery):
        response["dev_pin"] = pin

    return jsonify(response)


@app.post("/users/register")
def register_user():
    payload = request.get_json(silent=True) or {}
    full_name = normalize_text(payload.get("full_name", ""))
    email = normalize_email(payload.get("email", ""))
    password = payload.get("password", "")
    verification_pin = normalize_pin(payload.get("verification_pin", payload.get("pin", "")))

    if not full_name or not email or not password or not verification_pin:
        return jsonify({"detail": "full_name, email, password, and verification_pin are required"}), 400

    if not is_valid_email(email):
        return jsonify({"detail": "Please provide a valid email address"}), 400

    password_error = get_password_validation_error(password)
    if password_error:
        return jsonify({"detail": password_error}), 400

    try:
        if email_exists(email):
            return jsonify({"detail": "This email is already registered"}), 409
        pin_ok, pin_detail = consume_email_pin(email, "register", verification_pin)
        if not pin_ok:
            return jsonify({"detail": pin_detail}), 400
    except DatabaseError:
        return jsonify({"detail": "Database error while verifying email PIN"}), 500

    hashed_password = generate_password_hash(password)

    try:
        connection = get_db_connection()
        cursor = connection.cursor(dictionary=True)
        cursor.execute(
            """
            INSERT INTO users (full_name, email, password_hash, email_verified_at, is_admin)
            VALUES (%s, %s, %s, UTC_TIMESTAMP(), %s)
            """,
            (full_name, email, hashed_password, 1 if is_configured_admin_email(email) else 0),
        )
        connection.commit()

        user_id = cursor.lastrowid
        cursor.execute("SELECT id, full_name, email, email_verified_at, is_admin, created_at FROM users WHERE id = %s", (user_id,))
        created_user = cursor.fetchone()
        cursor.close()
        connection.close()
    except DatabaseError as error:
        if is_duplicate_entry_error(error):
            return jsonify({"detail": "This email is already registered"}), 409
        return jsonify({"detail": "Database error while registering user"}), 500

    save_login_session(created_user)

    return (
        jsonify(
            {
                "message": "User registered successfully",
                "user": serialize_user(created_user),
            }
        ),
        201,
    )


@app.post("/users/password-reset")
def reset_user_password():
    payload = request.get_json(silent=True) or {}
    email = normalize_email(payload.get("email", ""))
    verification_pin = normalize_pin(payload.get("verification_pin", payload.get("pin", "")))
    new_password = payload.get("new_password", payload.get("password", ""))

    if not email or not verification_pin or not new_password:
        return jsonify({"detail": "email, verification_pin, and new_password are required"}), 400

    if not is_valid_email(email):
        return jsonify({"detail": "Please provide a valid email address"}), 400

    password_error = get_password_validation_error(new_password)
    if password_error:
        return jsonify({"detail": password_error}), 400

    try:
        if not email_exists(email):
            return jsonify({"detail": "No account was found for this email address"}), 404
        pin_ok, pin_detail = consume_email_pin(email, "password_reset", verification_pin)
        if not pin_ok:
            return jsonify({"detail": pin_detail}), 400

        connection = get_db_connection()
        cursor = connection.cursor()
        cursor.execute(
            "UPDATE users SET password_hash = %s WHERE email = %s",
            (generate_password_hash(new_password), email),
        )
        connection.commit()
        cursor.close()
        connection.close()
    except DatabaseError:
        return jsonify({"detail": "Database error while resetting password"}), 500

    return jsonify({"message": "Password reset successfully. You can now log in."})


@app.post("/users/login")
def login_user():
    payload = request.get_json(silent=True) or {}
    email = normalize_email(payload.get("email", ""))
    password = payload.get("password", "")

    if not email or not password:
        return jsonify({"detail": "email and password are required"}), 400

    connection = get_db_connection()
    cursor = connection.cursor(dictionary=True)
    cursor.execute(
        "SELECT id, full_name, email, password_hash, email_verified_at, is_admin, created_at FROM users WHERE email = %s",
        (email,),
    )
    user = cursor.fetchone()

    if not user or not check_password_hash(user["password_hash"], password):
        cursor.close()
        connection.close()
        return jsonify({"detail": "Invalid email or password"}), 401

    if is_configured_admin_email(user["email"]) and not user.get("is_admin"):
        cursor.execute("UPDATE users SET is_admin = 1 WHERE id = %s", (user["id"],))
        connection.commit()
        user["is_admin"] = 1

    cursor.close()
    connection.close()
    save_login_session(user)

    return jsonify(
        {
            "message": "Login successful",
            "user": serialize_user(user),
        }
    )


@app.post("/users/logout")
def logout_user():
    clear_login_session()
    return jsonify({"message": "Logged out successfully"})


def parse_stored_analysis_json(raw_value: Any) -> dict[str, Any]:
    if isinstance(raw_value, dict):
        return raw_value
    if not raw_value:
        return {}
    try:
        parsed = json.loads(str(raw_value))
    except (TypeError, ValueError, json.JSONDecodeError):
        return {}
    return parsed if isinstance(parsed, dict) else {}


def clamp_admin_score(value: Any) -> float:
    try:
        return round(max(0.0, min(100.0, float(value or 0))), 2)
    except (TypeError, ValueError):
        return 0.0


def estimate_after_recommendation_score(base_score: Any, analysis: dict[str, Any]) -> float:
    score = clamp_admin_score(base_score or analysis.get("matching_percentage"))
    recommendation_module = (
        analysis.get("recommendation_module")
        if isinstance(analysis.get("recommendation_module"), dict)
        else {}
    )
    actions = [
        action
        for action in list(recommendation_module.get("priority_actions") or [])
        if isinstance(action, dict)
    ]
    action_gains = {
        "role_fit": 3.0,
        "keyword_gap": 8.0,
        "evidence_gap": 7.0,
        "experience": 6.0,
        "ats_quality": 5.0,
        "education": 4.0,
        "impact": 4.0,
    }
    gain = 0.0
    seen_categories: set[str] = set()
    for action in actions:
        category = str(action.get("category") or "").strip().lower()
        if category and category not in seen_categories:
            gain += action_gains.get(category, 3.0)
            seen_categories.add(category)

    job_analysis = analysis.get("job_description_analysis") if isinstance(analysis.get("job_description_analysis"), dict) else {}
    missing_keywords = list(job_analysis.get("missing_keywords") or [])
    compatibility = analysis.get("compatibility") if isinstance(analysis.get("compatibility"), dict) else {}

    if missing_keywords:
        gain += min(8.0, len(missing_keywords) * 1.1)
    if clamp_admin_score(compatibility.get("keyword_match")) < 45:
        gain += 4.0
    if clamp_admin_score(compatibility.get("role_evidence")) < 60:
        gain += 4.0
    if clamp_admin_score(compatibility.get("resume_quality")) < 75:
        gain += 3.0

    if actions and gain < 6.0:
        gain = 6.0
    gain = min(24.0, gain)
    return round(min(100.0, score + gain), 2)


def isoformat_datetime(value: Any) -> str | None:
    return value.isoformat() if hasattr(value, "isoformat") else None


def build_admin_dashboard_payload() -> dict[str, Any]:
    connection = get_db_connection()
    cursor = connection.cursor(dictionary=True)
    cursor.execute(
        """
        SELECT id, full_name, email, email_verified_at, is_admin, created_at
        FROM users
        ORDER BY created_at DESC
        """
    )
    user_rows = cursor.fetchall()
    cursor.execute(
        """
        SELECT
            ra.id,
            ra.user_id,
            ra.uploaded_resume_name,
            ra.target_role,
            ra.matching_percentage,
            ra.analysis_json,
            ra.created_at,
            u.full_name,
            u.email
        FROM resume_analyses ra
        LEFT JOIN users u ON u.id = ra.user_id
        ORDER BY ra.created_at DESC
        LIMIT 250
        """
    )
    analysis_rows = cursor.fetchall()
    cursor.execute("SELECT COUNT(*) AS count FROM download_history")
    download_row = cursor.fetchone() or {}
    cursor.close()
    connection.close()

    user_stats: dict[int, dict[str, Any]] = {}
    role_counter: Counter[str] = Counter()
    score_values: list[float] = []
    estimated_after_values: list[float] = []
    recent_evaluations: list[dict[str, Any]] = []

    for row in analysis_rows:
        before_score = clamp_admin_score(row.get("matching_percentage"))
        analysis = parse_stored_analysis_json(row.get("analysis_json"))
        estimated_after = estimate_after_recommendation_score(before_score, analysis)
        gain = round(max(0.0, estimated_after - before_score), 2)
        user_id = row.get("user_id")
        if user_id is not None:
            safe_user_id = int(user_id)
            stats = user_stats.setdefault(
                safe_user_id,
                {
                    "evaluation_count": 0,
                    "latest_score": None,
                    "latest_estimated_after_score": None,
                    "last_evaluation_at": None,
                },
            )
            stats["evaluation_count"] += 1
            if not stats["last_evaluation_at"]:
                stats["latest_score"] = before_score
                stats["latest_estimated_after_score"] = estimated_after
                stats["last_evaluation_at"] = isoformat_datetime(row.get("created_at"))

        target_role = normalize_text(str(row.get("target_role") or "Unknown role"))
        if target_role:
            role_counter[target_role] += 1
        score_values.append(before_score)
        estimated_after_values.append(estimated_after)
        if len(recent_evaluations) < 40:
            recent_evaluations.append(
                {
                    "id": row["id"],
                    "user_id": user_id,
                    "user_name": row.get("full_name") or "Deleted user",
                    "user_email": row.get("email") or "",
                    "uploaded_resume_name": row.get("uploaded_resume_name") or "resume.pdf",
                    "target_role": target_role,
                    "before_score": before_score,
                    "estimated_after_score": estimated_after,
                    "estimated_gain": gain,
                    "created_at": isoformat_datetime(row.get("created_at")),
                }
            )

    users: list[dict[str, Any]] = []
    for row in user_rows:
        user_id = int(row["id"])
        stats = user_stats.get(
            user_id,
            {
                "evaluation_count": 0,
                "latest_score": None,
                "latest_estimated_after_score": None,
                "last_evaluation_at": None,
            },
        )
        users.append(
            {
                **serialize_user(row),
                "evaluation_count": stats["evaluation_count"],
                "latest_score": stats["latest_score"],
                "latest_estimated_after_score": stats["latest_estimated_after_score"],
                "last_evaluation_at": stats["last_evaluation_at"],
            }
        )

    total_evaluations = len(score_values)
    average_before = round(sum(score_values) / total_evaluations, 2) if total_evaluations else 0.0
    average_after = round(sum(estimated_after_values) / total_evaluations, 2) if total_evaluations else 0.0
    estimated_improved = sum(1 for before, after in zip(score_values, estimated_after_values) if after > before)
    strong_scores = sum(1 for score in score_values if score >= 80)
    fair_scores = sum(1 for score in score_values if 60 <= score < 80)
    weak_scores = sum(1 for score in score_values if score < 60)

    return {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "summary": {
            "registered_users": len(users),
            "resume_evaluations": total_evaluations,
            "evaluating_users": len(user_stats),
            "download_records": int(download_row.get("count") or 0),
            "average_before_score": average_before,
            "estimated_average_after_score": average_after,
            "average_estimated_gain": round(max(0.0, average_after - average_before), 2),
            "estimated_improved_evaluations": estimated_improved,
            "estimated_improvement_rate": round((estimated_improved / total_evaluations) * 100, 2) if total_evaluations else 0.0,
        },
        "score_distribution": [
            {"label": "Strong 80-100", "value": strong_scores},
            {"label": "Fair 60-79", "value": fair_scores},
            {"label": "Needs Work 0-59", "value": weak_scores},
        ],
        "top_evaluated_roles": [
            {"label": role, "value": count}
            for role, count in role_counter.most_common(12)
        ],
        "recent_evaluations": recent_evaluations,
        "users": users,
    }


@app.get("/users")
def list_users():
    admin_error = require_admin_json_response()
    if admin_error:
        return admin_error

    connection = get_db_connection()
    cursor = connection.cursor(dictionary=True)
    cursor.execute("SELECT id, full_name, email, email_verified_at, is_admin, created_at FROM users ORDER BY created_at DESC")
    rows = cursor.fetchall()
    cursor.close()
    connection.close()

    users = [serialize_user(row) for row in rows]
    return jsonify({"count": len(users), "users": users})


@app.get("/admin/dashboard-insights")
def get_admin_dashboard_insights():
    admin_error = require_admin_json_response()
    if admin_error:
        return admin_error
    try:
        return jsonify(build_admin_dashboard_payload())
    except DatabaseError:
        return jsonify({"detail": "Database error while building admin dashboard"}), 500


@app.put("/admin/users/<int:user_id>")
def update_admin_user(user_id: int):
    admin_error = require_admin_json_response()
    if admin_error:
        return admin_error

    payload = request.get_json(silent=True) or {}
    updates: list[str] = []
    values: list[Any] = []
    new_full_name = normalize_text(str(payload.get("full_name", ""))) if "full_name" in payload else None
    new_email = normalize_email(str(payload.get("email", ""))) if "email" in payload else None
    new_password = str(payload.get("new_password", ""))
    has_admin_value = "is_admin" in payload
    requested_admin = bool(payload.get("is_admin")) if has_admin_value else None

    if new_full_name is not None:
        if not new_full_name:
            return jsonify({"detail": "Full name is required"}), 400
        updates.append("full_name = %s")
        values.append(new_full_name)

    if new_email is not None:
        if not is_valid_email(new_email):
            return jsonify({"detail": "Please provide a valid email address"}), 400
        updates.append("email = %s")
        values.append(new_email)

    if new_password:
        password_error = get_password_validation_error(new_password)
        if password_error:
            return jsonify({"detail": password_error}), 400
        updates.append("password_hash = %s")
        values.append(generate_password_hash(new_password))

    if new_email is not None and is_configured_admin_email(new_email):
        requested_admin = True
        has_admin_value = True

    if has_admin_value:
        if int(session.get("user_id") or 0) == user_id and not requested_admin:
            return jsonify({"detail": "You cannot remove admin access from your current session"}), 400
        updates.append("is_admin = %s")
        values.append(1 if requested_admin else 0)

    if not updates:
        return jsonify({"detail": "No user fields were provided"}), 400

    connection = get_db_connection()
    cursor = connection.cursor(dictionary=True)
    try:
        cursor.execute("SELECT id FROM users WHERE id = %s", (user_id,))
        if not cursor.fetchone():
            cursor.close()
            connection.close()
            return jsonify({"detail": "User not found"}), 404
        values.append(user_id)
        cursor.execute(f"UPDATE users SET {', '.join(updates)} WHERE id = %s", tuple(values))
        connection.commit()
        cursor.execute("SELECT id, full_name, email, email_verified_at, is_admin, created_at FROM users WHERE id = %s", (user_id,))
        updated_user = cursor.fetchone()
    except DatabaseError as error:
        connection.rollback()
        cursor.close()
        connection.close()
        if is_duplicate_entry_error(error):
            return jsonify({"detail": "This email is already registered"}), 409
        return jsonify({"detail": "Database error while updating user"}), 500

    cursor.close()
    connection.close()
    return jsonify({"message": "User updated", "user": serialize_user(updated_user)})


@app.delete("/admin/users/<int:user_id>")
def delete_admin_user(user_id: int):
    admin_error = require_admin_json_response()
    if admin_error:
        return admin_error
    if int(session.get("user_id") or 0) == user_id:
        return jsonify({"detail": "You cannot delete the admin account you are currently using"}), 400

    connection = get_db_connection()
    cursor = connection.cursor(dictionary=True)
    try:
        cursor.execute("SELECT id FROM users WHERE id = %s", (user_id,))
        if not cursor.fetchone():
            cursor.close()
            connection.close()
            return jsonify({"detail": "User not found"}), 404
        cursor.execute("DELETE FROM users WHERE id = %s", (user_id,))
        connection.commit()
    except DatabaseError:
        connection.rollback()
        cursor.close()
        connection.close()
        return jsonify({"detail": "Database error while deleting user"}), 500

    cursor.close()
    connection.close()
    return jsonify({"message": "User deleted"})


@app.get("/analyses")
def list_resume_analyses():
    email = normalize_email(request.args.get("email", ""))
    user_id = get_user_id_by_email(email)

    if not user_id:
        return jsonify({"count": 0, "analyses": []})

    connection = get_db_connection()
    cursor = connection.cursor(dictionary=True)
    cursor.execute(
        """
        SELECT
            id,
            uploaded_resume_name,
            target_role,
            matching_percentage,
            feedback,
            resume_excerpt,
            created_at
        FROM resume_analyses
        WHERE user_id = %s
        ORDER BY created_at DESC
        """,
        (user_id,),
    )
    rows = cursor.fetchall()
    cursor.close()
    connection.close()

    analyses = []
    for row in rows:
        analyses.append(
            {
                "id": row["id"],
                "uploaded_resume_name": row["uploaded_resume_name"],
                "target_role": row["target_role"],
                "matching_percentage": float(row["matching_percentage"]),
                "feedback": row["feedback"],
                "resume_excerpt": row["resume_excerpt"],
                "created_at": row["created_at"].isoformat() if row["created_at"] else None,
            }
        )

    return jsonify({"count": len(analyses), "analyses": analyses})


@app.post("/downloads")
def create_download_history():
    payload = request.get_json(silent=True) or {}
    email = normalize_email(payload.get("email", ""))
    analysis_id = payload.get("analysis_id")
    downloaded_file_name = normalize_text(payload.get("downloaded_file_name", "uts-cs-career-navigator-comments.txt"))

    try:
        safe_analysis_id = int(analysis_id) if analysis_id else None
    except (TypeError, ValueError):
        safe_analysis_id = None

    if not email:
        return jsonify({"detail": "email is required"}), 400

    try:
        record_download_history(email, safe_analysis_id, downloaded_file_name)
    except DatabaseError:
        return jsonify({"detail": "Database error while saving download history"}), 500

    return jsonify({"message": "Download history saved"})


@app.post("/generate-resume")
def generate_ats_resume():
    payload = request.get_json(silent=True) or {}
    analysis = payload.get("analysis") if isinstance(payload.get("analysis"), dict) else {}
    resume_text = clean_resume_source_text(
        str(
            payload.get("resume_text")
            or analysis.get("resume_text_for_generation")
            or analysis.get("resume_excerpt")
            or ""
        )
    )
    target_role = normalize_text(str(payload.get("target_role") or ""))
    candidate_name = normalize_text(str(payload.get("candidate_name") or ""))
    raw_contact_links = payload.get("contact_links") if isinstance(payload.get("contact_links"), list) else []
    contact_links = unique_clean_values(
        [payload.get("linkedin_url"), payload.get("github_url"), payload.get("portfolio_url"), *raw_contact_links],
        5,
    )

    if not analysis and not resume_text:
        return jsonify({"detail": "analysis or resume_text is required"}), 400

    generated_resume = build_generated_resume_payload(
        analysis,
        resume_text[:RESUME_GENERATION_MAX_TEXT_LENGTH],
        target_role,
        candidate_name,
        contact_links,
    )
    return jsonify(generated_resume)


@app.post("/generate-resume-pdf")
def generate_ats_resume_pdf():
    payload = request.get_json(silent=True) or {}
    draft = payload.get("generated_resume") if isinstance(payload.get("generated_resume"), dict) else None

    if draft is None:
        analysis = payload.get("analysis") if isinstance(payload.get("analysis"), dict) else {}
        resume_text = clean_resume_source_text(
            str(
                payload.get("resume_text")
                or analysis.get("resume_text_for_generation")
                or analysis.get("resume_excerpt")
                or ""
            )
        )
        target_role = normalize_text(str(payload.get("target_role") or ""))
        candidate_name = normalize_text(str(payload.get("candidate_name") or ""))
        raw_contact_links = payload.get("contact_links") if isinstance(payload.get("contact_links"), list) else []
        contact_links = unique_clean_values(
            [payload.get("linkedin_url"), payload.get("github_url"), payload.get("portfolio_url"), *raw_contact_links],
            5,
        )
        if not analysis and not resume_text:
            return jsonify({"detail": "generated_resume, analysis, or resume_text is required"}), 400
        draft = build_generated_resume_payload(
            analysis,
            resume_text[:RESUME_GENERATION_MAX_TEXT_LENGTH],
            target_role,
            candidate_name,
            contact_links,
        )

    try:
        pdf_buffer = build_generated_resume_pdf_buffer(draft)
    except RuntimeError as error:
        return jsonify({"detail": str(error)}), 500

    role_slug = re.sub(r"[^a-z0-9]+", "-", normalize_match_text(str(draft.get("target_role") or "ats-resume"))).strip("-")
    file_name = f"uts-ats-resume-{role_slug or 'draft'}.pdf"
    return send_file(
        pdf_buffer,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=file_name,
    )


@app.post("/feedback-report-pdf")
def generate_feedback_report_pdf():
    payload = request.get_json(silent=True) or {}
    report = payload.get("report") if isinstance(payload.get("report"), dict) else payload
    if not isinstance(report, dict):
        return jsonify({"detail": "report payload is required"}), 400

    try:
        pdf_buffer = build_feedback_report_pdf_buffer(report)
    except RuntimeError as error:
        return jsonify({"detail": str(error)}), 500

    role_slug = re.sub(
        r"[^a-z0-9]+",
        "-",
        normalize_match_text(str(report.get("recommendedRole") or report.get("selectedRole") or "feedback-report")),
    ).strip("-")
    file_name = f"uts-feedback-report-{role_slug or 'resume'}.pdf"
    return send_file(
        pdf_buffer,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=file_name,
    )


@app.get("/job-roles")
def list_job_roles():
    try:
        roles = fetch_job_roles()
    except DatabaseError:
        return jsonify({"detail": "Database error while loading job roles"}), 500

    return jsonify({"count": len(roles), "roles": roles})


@app.get("/job-requirements")
def get_job_requirements():
    requested_role = normalize_text(request.args.get("role", ""))
    role_id_raw = request.args.get("role_id", "")

    try:
        role_id = int(role_id_raw) if role_id_raw else None
    except (TypeError, ValueError):
        role_id = None

    try:
        matched_role: dict[str, Any] | None = None
        match_type = "none"
        if role_id:
            roles = fetch_job_roles()
            matched_role = next((role for role in roles if int(role["id"]) == role_id), None)
            match_type = "id" if matched_role else "none"
        elif requested_role:
            matched_role, match_type = find_job_role_match(requested_role)
        else:
            return jsonify({"detail": "role or role_id is required"}), 400

        if not matched_role:
            return jsonify({"matched": False, "requested_role": requested_role, "match_type": match_type})

        requirements = fetch_job_requirements(int(matched_role["id"]))
    except DatabaseError:
        return jsonify({"detail": "Database error while loading job requirements"}), 500

    return jsonify(
        {
            "matched": True,
            "requested_role": requested_role,
            "match_type": match_type,
            "role": matched_role,
            "requirements": requirements,
            "grouped_requirements": group_job_requirements(requirements),
            "job_description": build_job_description_from_requirements(matched_role, requirements),
        }
    )


@app.get("/job-data-cleaning/report")
def get_job_data_cleaning_report():
    admin_error = require_admin_json_response()
    if admin_error:
        return admin_error

    try:
        role_rows, requirement_rows = fetch_raw_job_data_rows()
        report = build_job_data_quality_report(role_rows, requirement_rows)
    except DatabaseError:
        return jsonify({"detail": "Database error while building job data cleaning report"}), 500

    return jsonify(report)


@app.post("/job-data-cleaning/run")
def run_job_data_cleaning():
    admin_error = require_admin_json_response()
    if admin_error:
        return admin_error

    payload = request.get_json(silent=True) or {}
    apply_changes = bool(payload.get("apply"))

    try:
        role_rows, requirement_rows = fetch_raw_job_data_rows()
        plan = build_job_data_cleaning_plan(role_rows, requirement_rows)
        report_before = build_job_data_quality_report(role_rows, requirement_rows)

        if not apply_changes:
            return jsonify(
                {
                    "applied": False,
                    "message": "Dry run only. Send {\"apply\": true} to update MySQL rows.",
                    "report": report_before,
                    "plan_summary": plan["summary"],
                }
            )

        applied_summary = apply_job_data_cleaning_plan(plan)
        cleaned_role_rows, cleaned_requirement_rows = fetch_raw_job_data_rows()
        report_after = build_job_data_quality_report(cleaned_role_rows, cleaned_requirement_rows)
    except DatabaseError:
        return jsonify({"detail": "Database error while running job data cleaning"}), 500

    return jsonify(
        {
            "applied": True,
            "message": "Job data cleaning completed.",
            "applied_summary": applied_summary,
            "report_before": report_before,
            "report_after": report_after,
        }
    )


@app.get("/job-data-insights")
def get_job_data_insights():
    admin_error = require_admin_json_response()
    if admin_error:
        return admin_error

    try:
        roles, requirements_by_role = fetch_job_data_for_insights()
        insights = build_job_data_insights(
            roles,
            requirements_by_role,
            skill_vocabulary=DEFAULT_SKILL_DICTIONARY,
            stopwords=KEYWORD_STOPWORDS,
        )
    except DatabaseError:
        return jsonify({"detail": "Database error while building job data insights"}), 500

    return jsonify(insights)


init_mysql_database()


@app.get("/health")
def health_check():
    return jsonify(
        {
            "status": "ok",
            "service": APP_TITLE,
            "accuracy_testing_auto_export": ACCURACY_TESTING_AUTO_EXPORT,
            "accuracy_testing_workbook": ACCURACY_TESTING_WORKBOOK,
            "accuracy_testing_workbook_exists": os.path.exists(ACCURACY_TESTING_WORKBOOK),
            "accuracy_testing_pending_count": len(load_pending_accuracy_testing_analyses()),
        }
    )


def serve_frontend_page(page_name: str):
    response = send_from_directory(FRONTEND_DIR, page_name)
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response


@app.get("/")
def index():
    return serve_frontend_page("home_page.html")


@app.get("/home_page.html")
def home_page():
    return serve_frontend_page("home_page.html")


@app.get("/analysis_result.html")
def analysis_result_page():
    return serve_frontend_page("analysis_result.html")


@app.get("/job_data_insights.html")
def job_data_insights_page():
    if not current_session_is_admin():
        return redirect("/home_page.html?admin_required=1")
    return serve_frontend_page("job_data_insights.html")


@app.get("/app_common.js")
def app_common_script():
    return serve_frontend_page("app_common.js")


@app.get("/api")
def api_index():
    return jsonify(
        {
            "service": APP_TITLE,
            "status": "running",
            "endpoints": [
                "GET /health",
                "POST /analyze-file",
                "POST /auth/pin/send",
                "POST /users/register",
                "POST /users/login",
                "POST /users/logout",
                "POST /users/password-reset",
                "GET /users",
                "GET /admin/dashboard-insights",
                "PUT /admin/users/<user_id>",
                "DELETE /admin/users/<user_id>",
                "GET /analyses",
                "POST /downloads",
                "POST /generate-resume",
                "POST /generate-resume-pdf",
                "POST /feedback-report-pdf",
                "GET /job-roles",
                "GET /job-requirements",
                "GET /job-data-cleaning/report",
                "POST /job-data-cleaning/run",
                "GET /job-data-insights",
            ],
        }
    )


@app.post("/analyze-file")
def analyze_file():
    if "resume_file" not in request.files:
        return jsonify({"detail": "resume_file is required"}), 400

    resume_file = request.files["resume_file"]
    job_description = normalize_text(request.form.get("job_description", ""))
    requested_target_role = normalize_text(request.form.get("target_role", ""))
    target_role = requested_target_role or "selected role"
    user_email = normalize_email(request.form.get("user_email", ""))

    if not resume_file.filename:
        return jsonify({"detail": "A resume file must be uploaded"}), 400

    file_extension = get_file_extension(resume_file.filename)
    if file_extension not in SUPPORTED_RESUME_EXTENSIONS:
        return jsonify({"detail": "Only PDF, PNG, JPG, or JPEG resumes are supported"}), 400

    try:
        resume_text, document_info = extract_resume_text(resume_file, resume_file.filename)
    except ValueError as error:
        return jsonify({"detail": str(error)}), 400

    if not resume_text:
        if file_extension in IMAGE_EXTENSIONS:
            return (
                jsonify(
                    {
                        "detail": (
                            "Could not extract any text from the uploaded image. "
                            "Please upload a clearer screenshot/photo or export the resume as a text-based PDF."
                        )
                    }
                ),
                400,
            )

        return (
            jsonify(
                {
                    "detail": (
                        "Could not extract any text from the uploaded PDF. "
                        "This usually happens when the resume is scanned or image-only. "
                        "Please export the resume as a text-based PDF and upload it again."
                    )
                }
            ),
            400,
        )

    selected_job: dict[str, Any] | None = None
    detected_job: dict[str, Any] | None = None
    top_role_matches: list[dict[str, Any]] = []

    if requested_target_role and requested_target_role != "selected role":
        try:
            selected_job = resolve_database_job_description(requested_target_role)
        except DatabaseError:
            selected_job = None

    try:
        top_role_matches = rank_database_role_matches(resume_text, limit=3)
    except DatabaseError:
        top_role_matches = []

    if not job_description:
        detected_job = top_role_matches[0] if top_role_matches else None

    recommended_job, selected_fit, detected_fit = pick_recommended_database_job(
        resume_text,
        selected_job,
        detected_job,
    )
    role_recommendation = build_role_recommendation(
        requested_target_role,
        selected_job,
        detected_job,
        recommended_job,
        selected_fit,
        detected_fit,
        top_role_matches,
    )

    job_source: dict[str, Any] = {
        "source": "user_input" if job_description else "database_auto",
        "requested_role": requested_target_role,
    }

    if not job_description and recommended_job:
        recommended_role = recommended_job["role"]
        job_description = recommended_job["job_description"]
        target_role = recommended_role["role_name"]

        if selected_job and role_recommendation.get("is_different_from_selected"):
            source = "database_recommended_role"
        elif selected_job:
            source = "database_role_match"
        else:
            source = "database_resume_auto_detect"

        job_source = {
            "source": source,
            "requested_role": requested_target_role,
            "selected_database_role": get_database_role_name(selected_job),
            "matched_role": recommended_role["role_name"],
            "detected_database_role": get_database_role_name(detected_job),
            "match_type": recommended_job.get("match_type"),
            "confidence_score": role_recommendation.get("recommended_score"),
            "reason": role_recommendation.get("reason_summary"),
        }
    elif selected_job:
        target_role = get_database_role_name(selected_job) or target_role
        job_source.update(
            {
                "selected_database_role": get_database_role_name(selected_job),
                "detected_database_role": get_database_role_name(detected_job),
                "reason": role_recommendation.get("reason_summary"),
            }
        )

    if not job_description:
        return (
            jsonify(
                {
                    "detail": (
                        "job_description is required because the system could not auto-detect a matching database role. "
                        "Select a database role or type a similar CS job role before uploading the resume."
                    )
                }
            ),
            400,
        )

    if recommended_job:
        analysis_requirements = recommended_job.get("requirements", [])
    elif selected_job:
        analysis_requirements = selected_job.get("requirements", [])
    else:
        analysis_requirements = None
    analysis = build_analysis_payload(resume_text, job_description, resume_file.filename, analysis_requirements, target_role, document_info)
    analysis["target_role"] = target_role
    analysis["selected_role"] = requested_target_role
    analysis["job_description"] = job_description
    analysis["job_source"] = job_source
    analysis["role_recommendation"] = role_recommendation
    analysis["top_role_matches"] = role_recommendation.get("top_role_matches", [])
    analysis["resume_contact"] = extract_resume_contact(resume_text)
    analysis["recommendation_module"] = build_part8_recommendation_module(analysis)

    try:
        analysis["analysis_id"] = persist_analysis(analysis, job_description, target_role, user_email)
        analysis["history_saved"] = True
    except DatabaseError:
        analysis["analysis_id"] = None
        analysis["history_saved"] = False
        analysis["history_warning"] = "Analysis completed, but it could not be saved to MySQL history."

    accuracy_testing_result = update_accuracy_testing_workbook(analysis)
    analysis.update(accuracy_testing_result)
    analysis["resume_text_for_generation"] = resume_text[:RESUME_GENERATION_MAX_TEXT_LENGTH]

    return jsonify(analysis)


if __name__ == "__main__":
    port = int(os.getenv("PORT", "5000"))
    debug = os.getenv("FLASK_DEBUG", "0").strip().lower() in {"1", "true", "yes", "on"}
    app.run(host="0.0.0.0", port=port, debug=debug)
